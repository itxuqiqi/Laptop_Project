# fill_1_2.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Any, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet


# ==========================================================
# 基础工具
# ==========================================================
def to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def norm_name(x: Any) -> str:
    """
    ✅统一规范化：去空格/全角空格/换行
    ✅统一括号：中文括号→英文括号
    ✅统一连接符：各种短横线→-
    """
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\r", "").replace("\n", "").replace("\t", "")
    s = s.replace("\u3000", "").replace(" ", "").strip()

    # 括号统一
    s = s.replace("（", "(").replace("）", ")")

    # 横线统一
    s = (
        s.replace("－", "-")
        .replace("—", "-")
        .replace("–", "-")
        .replace("−", "-")
    )
    return s


def safe_div(a: float, b: float) -> Optional[float]:
    if b == 0:
        return None
    return a / b


# 1-1 金额口径：万元；1-1 原料数量口径：吨（已调整为吨）
# 1-2 吨油指标希望输出：元/吨
# 因此：单位成本 = (万元 * 10000) / 吨
WANYUAN_TO_YUAN = 10000.0


def cost_yuan_per_ton(amount_wanyuan: float, qty_ton: float) -> Optional[float]:
    """万元 + 吨 -> 元/吨"""
    return safe_div(amount_wanyuan * WANYUAN_TO_YUAN, qty_ton)


# ==========================================================
# 1-2 表定位：A列锚点，从第6行开始
# ==========================================================
@dataclass
class RowRef:
    row: int
    anchor: str


def build_index_1_2(ws: Worksheet, start_row: int = 6, anchor_col: int = 1) -> Dict[str, list[RowRef]]:
    """
    1-2 锚点索引（支持同名锚点出现多次）：
    - key: 规范化后的锚点名
    - value: 该锚点对应的所有行（按出现顺序）
    """
    idx: Dict[str, list[RowRef]] = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, anchor_col).value
        if not v:
            continue
        k = norm_name(v)
        if not k:
            continue
        idx.setdefault(k, []).append(RowRef(row=r, anchor=str(v)))
    return idx


def write_1_2(ws: Worksheet, idx: Dict[str, list[RowRef]], anchor: str,
             e: Optional[float] = None, f: Optional[float] = None, g: Optional[float] = None):
    k = norm_name(anchor)
    if k not in idx:
        print(f"DEBUG ❌ 1-2 未找到锚点行: {anchor}")
        return

    # ✅ 支持同名锚点：把值写入所有匹配行（例如“财务费用”出现两次）
    for ref in idx[k]:
        r = ref.row
        # E=5, F=6, G=7
        if e is not None:
            ws.cell(r, 5).value = e
        if f is not None:
            ws.cell(r, 6).value = f
        if g is not None:
            ws.cell(r, 7).value = g


def build_index_1_1(ws11: Worksheet, start_row: int = 6, anchor_col: int = 1) -> Dict[str, int]:
    idx = {}
    for r in range(start_row, ws11.max_row + 1):
        v = ws11.cell(r, anchor_col).value
        if not v:
            continue
        k = norm_name(v)
        if k and k not in idx:
            idx[k] = r
    return idx


def read_1_1_vals(ws11: Worksheet, idx11: Dict[str, int], anchor: str) -> Tuple[float, float, float, float]:
    """
    返回：(本月数量D, 本月金额F, 累计数量G, 累计金额I)
    """
    k = norm_name(anchor)
    if k not in idx11:
        return 0.0, 0.0, 0.0, 0.0

    r = idx11[k]
    mq = to_float(ws11.cell(r, 4).value)  # D
    ma = to_float(ws11.cell(r, 6).value)  # F
    yq = to_float(ws11.cell(r, 7).value)  # G
    ya = to_float(ws11.cell(r, 9).value)  # I
    return mq, ma, yq, ya


# ==========================================================
# 主函数：填充 1-2
# ==========================================================
def fill_cost_sheet_1_2(
    ws_12: Worksheet,
    ws_11: Worksheet,
    start_row_12: int = 6,
    start_row_11: int = 6,
):
    """
    ws_12: 成本表 sheet 1-2
    ws_11: 成本表 sheet 1-1
    """

    idx12 = build_index_1_2(ws_12, start_row=start_row_12, anchor_col=1)
    idx11 = build_index_1_1(ws_11, start_row=start_row_11, anchor_col=1)

    # ==========================================================
    # 1) 原料数量口径：必须存在
    # ==========================================================
    raw_m_qty, _, raw_y_qty, _ = read_1_1_vals(ws_11, idx11, "原料")
    if raw_m_qty == 0 or raw_y_qty == 0:
        print("DEBUG ⚠️ 1-1 原料数量为0，1-2将无法计算吨油指标")
        return

    # ==========================================================
    # 2) 通用逻辑：同名锚点金额 / 原料数量
    # ==========================================================
    def fill_by_same_anchor(anchor: str):
        _, ma, _, ya = read_1_1_vals(ws_11, idx11, anchor)
        if ma == 0 and ya == 0:
            # 不强制写0，避免覆盖模板
            return

        e = cost_yuan_per_ton(ma, raw_m_qty)
        f = cost_yuan_per_ton(ya, raw_y_qty)
        g = e  # 默认与 E 一致
        write_1_2(ws_12, idx12, anchor, e=e, f=f, g=g)

    # ==========================================================
    # 3) 全量锚点：你描述的全部锚点都应当存在
    # ==========================================================
    SAME_ANCHORS = [
        # 完全费用体系
        "完全费用",
        "现金操作成本",

        # 变动费用体系
        "变动费用小计",
        "变动费用-外购辅助材料",
        "变动费用-外购动力",
        "变动费用-外购动力-新鲜水（吨）",
        "变动费用-外购动力-电（千瓦时、元/千瓦时）",
        "变动费用-外购动力-蒸汽（吨)",
        "变动费用-外购动力-氮气（标立)",
        "变动费用-外购动力-其它",
        "变动费用-外购燃料",

        # 固定费用体系
        "不含折旧、财务费用的固定费用小计",
        "不含折旧、财务费用的固定费用-修理费",
        "不含折旧、财务费用的固定费用-职工薪酬",
        "不含折旧、财务费用的固定费用-其他管理销售费用",

        # 外供劳务及动力
        "外供劳务及动力（减项）",

        # 折旧体系
        "折旧费及摊销",
        "折旧费及摊销-折旧费",
        "折旧费及摊销-无形资产摊销",
        "折旧费及摊销-长期待摊费用摊销",

        # 财务费用
        "财务费用",

        # 合计类
        "成本费用合计（剔除芳烃II）",
        "芳烃联合II成本",
        "成本费用合计",

        # 三大费用（这些直接取1-1对应锚点）
        "管理费用",
        "财务费用",
        "销售费用",
    ]

    for a in SAME_ANCHORS:
        fill_by_same_anchor(a)

    # ==========================================================
    # 4) 三大费用：明确来自 1-1 的特定锚点
    # ==========================================================
    def fill_direct(anchor_12: str, anchor_11: str):
        _, ma, _, ya = read_1_1_vals(ws_11, idx11, anchor_11)
        if ma == 0 and ya == 0:
            return
        e = cost_yuan_per_ton(ma, raw_m_qty)
        f = cost_yuan_per_ton(ya, raw_y_qty)
        g = e
        write_1_2(ws_12, idx12, anchor_12, e=e, f=f, g=g)

    fill_direct("管理费用", "成本费用-减：管理费用")
    fill_direct("财务费用", "成本费用-财务费用")
    fill_direct("销售费用", "成本费用-销售费用")

    # ==========================================================
    # 5) 指标-吨油期间费用 = (管理+财务+销售)/原料数量
    # ==========================================================
    def calc_period_fee():
        _, mg_ma, _, mg_ya = read_1_1_vals(ws_11, idx11, "成本费用-减：管理费用")
        _, fin_ma, _, fin_ya = read_1_1_vals(ws_11, idx11, "成本费用-财务费用")
        _, rd_ma, _, rd_ya = read_1_1_vals(ws_11, idx11, "成本费用-销售费用")

        m_total = mg_ma + fin_ma + rd_ma
        y_total = mg_ya + fin_ya + rd_ya

        e = cost_yuan_per_ton(m_total, raw_m_qty)
        f = cost_yuan_per_ton(y_total, raw_y_qty)
        g = e
        write_1_2(ws_12, idx12, "指标-吨油期间费用", e=e, f=f, g=g)
        return e, f

    e_pf, f_pf = calc_period_fee()

    # ==========================================================
    # 6) 指标-研发费用 = 研发费用 / 原料数量
    # ==========================================================
    def calc_rd_fee():
        _, rd_ma, _, rd_ya = read_1_1_vals(ws_11, idx11, "成本费用-研发费用")

        e = cost_yuan_per_ton(rd_ma, raw_m_qty)
        f = cost_yuan_per_ton(rd_ya, raw_y_qty)
        g = e
        write_1_2(ws_12, idx12, "指标-研发费用", e=e, f=f, g=g)
        return e, f

    e_rd, f_rd = calc_rd_fee()

    # ==========================================================
    # 7) 指标-吨油加工成本 = 完全费用 - 吨油期间费用 - 吨油研发费用
    # ==========================================================
    def calc_process_cost():
        # 完全费用吨油：从 1-2 已经写过（完全费用锚点）
        _, full_ma, _, full_ya = read_1_1_vals(ws_11, idx11, "完全费用")
        e_full = cost_yuan_per_ton(full_ma, raw_m_qty)
        f_full = cost_yuan_per_ton(full_ya, raw_y_qty)

        if e_full is None or e_pf is None or e_rd is None:
            e = None
        else:
            e = e_full - e_pf - e_rd

        if f_full is None or f_pf is None or f_rd is None:
            f = None
        else:
            f = f_full - f_pf - f_rd

        g = e
        write_1_2(ws_12, idx12, "指标-吨油加工成本", e=e, f=f, g=g)

    calc_process_cost()

    print("✅ fill_cost_sheet_1_2 完成")
