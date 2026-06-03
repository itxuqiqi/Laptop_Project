# ton_barrel.py
from openpyxl import load_workbook
from tkinter import messagebox
from utils import norm, to_float, find_row_by_a, find_row_by_a_contains

def read_ton_barrel_ratio_map(file_expenses):
    """
    生产经营费用表 -> 吨桶比：
    sheet: 原油(万)
    U列=物料号(21) V列=物料名称(22) Y列=吨桶比(25)
    第8行开始
    返回 dict[(物料号, 物料名norm)] = ratio
    """
    wb = load_workbook(file_expenses, data_only=True)
    if "原油(万)" not in wb.sheetnames:
        raise ValueError("生产经营费用表缺少 sheet：原油(万)")
    ws = wb["原油(万)"]

    ratio_map = {}
    for r in range(8, ws.max_row + 1):
        mat = ws.cell(r, 21).value
        name = ws.cell(r, 22).value
        ratio = ws.cell(r, 25).value

        mat_s = "" if mat is None else str(mat).strip().lstrip("0")
        name_s = norm(name)
        ratio_f = to_float(ratio)

        if not mat_s or not name_s:
            continue
        if ratio_f is None:
            continue

        ratio_map[(mat_s, name_s)] = ratio_f

    return ratio_map

def calc_weighted_ratio(ws_cost, detail_rows, ratio_map, qty_col):
    num = 0.0
    den = 0.0
    for r in detail_rows:
        qty = to_float(ws_cost.cell(r, qty_col).value)
        if qty is None or qty == 0:
            continue

        mat = ws_cost.cell(r, 3).value
        name = ws_cost.cell(r, 4).value
        mat_s = "" if mat is None else str(mat).strip().lstrip("0")
        name_s = norm(name)

        ratio = ratio_map.get((mat_s, name_s))
        if ratio is None:
            continue

        num += qty * ratio
        den += qty

    return None if den == 0 else (num / den)

def apply_ton_barrel_and_usd_per_bbl(ws_cost, ratio_map, fx_rate):
    """
    - 扫描一般贸易明细（A=进口原油/海洋原油，且D非空）
    - 缺吨桶比 -> 弹窗提示（按 物料号C + 物料名D 匹配）
    - 计算四阶段吨桶比（期初/采购/加工/期末）加权平均
    - 读取一般贸易合计行单价（采购I/加工L/期末R）
    - 运杂费预留三个口：锚点“运杂费”行 E/F/G，目前默认0
    - 写入锚点“美元/桶”行 H~P
    """
    import_rows, ocean_rows, missing = [], [], []

    for r in range(8, ws_cost.max_row + 1):
        cat = ws_cost.cell(r, 1).value
        name = ws_cost.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue

        cat_s = "" if cat is None else str(cat).strip()
        if cat_s not in ("进口原油", "海洋原油"):
            continue

        mat = ws_cost.cell(r, 3).value
        mat_s = "" if mat is None else str(mat).strip().lstrip("0")
        name_s = norm(name)

        if (mat_s, name_s) not in ratio_map:
            missing.append(f"{cat_s} | {mat_s} | {str(name).strip()}")

        (import_rows if cat_s == "进口原油" else ocean_rows).append(r)

    if missing:
        preview = "\n".join(missing[:30])
        more = "" if len(missing) <= 30 else f"\n... 还有 {len(missing)-30} 条未显示"
        messagebox.showwarning(
            "吨桶比缺失提示",
            "以下一般贸易明细油品未匹配吨桶比（按物料号+物料名匹配）：\n\n" + preview + more
        )

    stages = {"begin": 5, "purchase": 8, "process": 11, "end": 17}
    all_rows = import_rows + ocean_rows

    ratio_imp, ratio_oce, ratio_gen = {}, {}, {}
    for key, qty_col in stages.items():
        ratio_imp[key] = calc_weighted_ratio(ws_cost, import_rows, ratio_map, qty_col)
        ratio_oce[key] = calc_weighted_ratio(ws_cost, ocean_rows, ratio_map, qty_col)
        ratio_gen[key] = calc_weighted_ratio(ws_cost, all_rows, ratio_map, qty_col)

    # 运杂费口：锚点“运杂费”行 E/F/G；当前默认0
    row_freight = find_row_by_a_contains(ws_cost, "运杂费")
    freight_gen = freight_imp = freight_oce = 0.0
    # 后续接入时再读 row_freight 的 E/F/G

    row_gen = find_row_by_a(ws_cost, "原油(一般贸易)")
    row_imp_sum = find_row_by_a(ws_cost, "原油(一般贸易)-进口原油")
    row_oce_sum = find_row_by_a(ws_cost, "原油(一般贸易)-海洋原油")

    if row_gen is None or row_imp_sum is None or row_oce_sum is None:
        messagebox.showerror(
            "缺少合计行",
            "找不到一般贸易合计行：原油(一般贸易)/原油(一般贸易)-进口原油/原油(一般贸易)-海洋原油\n无法计算美元/桶。"
        )
        return

    def get_price(row, col):
        return to_float(ws_cost.cell(row, col).value)

    p_purchase_gen = get_price(row_gen, 9)
    p_purchase_imp = get_price(row_imp_sum, 9)
    p_purchase_oce = get_price(row_oce_sum, 9)

    p_process_gen = get_price(row_gen, 12)
    p_process_imp = get_price(row_imp_sum, 12)
    p_process_oce = get_price(row_oce_sum, 12)

    p_end_gen = get_price(row_gen, 18)
    p_end_imp = get_price(row_imp_sum, 18)
    p_end_oce = get_price(row_oce_sum, 18)

    row_usd = find_row_by_a_contains(ws_cost, "美元/桶")
    if row_usd is None:
        messagebox.showerror("缺少锚点行", "找不到A列包含“美元/桶”的锚点行，无法写入美元/桶结果。")
        return

    def safe_usd(price_rmb_per_ton, ratio_ton_to_bbl, fx, freight=0.0):
        if price_rmb_per_ton is None or ratio_ton_to_bbl is None or ratio_ton_to_bbl == 0 or fx is None or fx == 0:
            return "-"
        v = price_rmb_per_ton - (0.0 if freight is None else freight)
        return v / ratio_ton_to_bbl / fx

    ws_cost.cell(row_usd, 8,  value=safe_usd(p_purchase_gen, ratio_gen["purchase"], fx_rate, freight_gen))
    ws_cost.cell(row_usd, 9,  value=safe_usd(p_purchase_imp, ratio_imp["purchase"], fx_rate, freight_imp))
    ws_cost.cell(row_usd, 10, value=safe_usd(p_purchase_oce, ratio_oce["purchase"], fx_rate, freight_oce))

    ws_cost.cell(row_usd, 11, value=safe_usd(p_process_gen, ratio_gen["process"], fx_rate, freight_gen))
    ws_cost.cell(row_usd, 12, value=safe_usd(p_process_imp, ratio_imp["process"], fx_rate, freight_imp))
    ws_cost.cell(row_usd, 13, value=safe_usd(p_process_oce, ratio_oce["process"], fx_rate, freight_oce))

    ws_cost.cell(row_usd, 14, value=safe_usd(p_end_gen, ratio_gen["end"], fx_rate, freight_gen))
    ws_cost.cell(row_usd, 15, value=safe_usd(p_end_imp, ratio_imp["end"], fx_rate, freight_imp))
    ws_cost.cell(row_usd, 16, value=safe_usd(p_end_oce, ratio_oce["end"], fx_rate, freight_oce))
