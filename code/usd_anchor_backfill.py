# usd_anchor_backfill.py
from openpyxl.worksheet.worksheet import Worksheet
from utils import find_row_by_a, find_row_by_a_contains, to_float


def _find_usd_row(ws: Worksheet) -> int | None:
    # 你成本表里“美元/桶”是 A 列包含
    r = find_row_by_a_contains(ws, "美元/桶")
    return r


def _read_23_oil_anchor_prices(ws_23: Worksheet):
    """
    从 2-3 读取三个锚点的：
    - F列：期初美元/桶
    - N列：年初美元/桶
    返回：
      (f_gen, f_imp, f_oce, n_gen, n_imp, n_oce)
    """
    r_gen = find_row_by_a(ws_23, "原油(一般贸易)")
    r_imp = find_row_by_a(ws_23, "原油(一般贸易)-进口原油")
    r_oce = find_row_by_a(ws_23, "原油(一般贸易)-海洋原油")

    def read_cell(row, col):
        if row is None:
            return None
        return to_float(ws_23.cell(row, col).value)

    # 2-3: F=6, N=14
    f_gen = read_cell(r_gen, 6)
    f_imp = read_cell(r_imp, 6)
    f_oce = read_cell(r_oce, 6)

    n_gen = read_cell(r_gen, 14)
    n_imp = read_cell(r_imp, 14)
    n_oce = read_cell(r_oce, 14)

    return f_gen, f_imp, f_oce, n_gen, n_imp, n_oce


def backfill_usd_anchor_from_23(ws_21: Worksheet, ws_22: Worksheet, ws_23: Worksheet):
    """
    ✅你要求的口径：
    - 2-3 三个锚点 F 列 = 期初美元/桶 -> 回填 2-1 “美元/桶”锚点行 E/F/G
    - 2-3 三个锚点 N 列 = 年初美元/桶 -> 回填 2-2 “美元/桶”锚点行 E/F/G
    """
    f_gen, f_imp, f_oce, n_gen, n_imp, n_oce = _read_23_oil_anchor_prices(ws_23)

    r21 = _find_usd_row(ws_21)
    if r21 is not None:
        ws_21.cell(r21, 5, value=0 if f_gen is None else f_gen)  # E
        ws_21.cell(r21, 6, value=0 if f_imp is None else f_imp)  # F
        ws_21.cell(r21, 7, value=0 if f_oce is None else f_oce)  # G

    r22 = _find_usd_row(ws_22)
    if r22 is not None:
        ws_22.cell(r22, 5, value=0 if n_gen is None else n_gen)  # E
        ws_22.cell(r22, 6, value=0 if n_imp is None else n_imp)  # F
        ws_22.cell(r22, 7, value=0 if n_oce is None else n_oce)  # G
