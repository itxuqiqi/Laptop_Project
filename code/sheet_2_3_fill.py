# sheet_2_3_fill.py
# Python 3.7 compatible (no "X | Y" union syntax)

from typing import Optional, Dict, Tuple, Any, Set, List

from openpyxl.worksheet.worksheet import Worksheet
from utils import norm, to_float, find_row_by_a
from row_locator import find_insert_row_after_category_anchor
import re
# =========================
# 分类集合
# =========================
INTRANSIT_CATS = {
    "在途原油-进口原油",
    "在途原油-海洋原油",
    "DES/DAT/DAP在途（货权未转移）",
}
# ✅2-3 不涉及“外购原料油”，只保留这三类原油/加工
OIL_CATS = {"进口原油", "海洋原油", "来料加工"}

DES_CAT = "DES/DAT/DAP在途（货权未转移）"
INTRANSIT_TOTAL_TITLE = "在途原油-合计"
INTRANSIT_IMP_TITLE = "在途原油-进口原油-合计"
INTRANSIT_OCE_TITLE = "在途原油-海洋原油-合计"
DES_TOTAL_TITLE = "DES/DAT/DAP在途（货权未转移）-合计"


# =========================
# 工具
# =========================
def _set(ws: Worksheet, r: int, c: int, v):
    ws.cell(r, c, value="0" if v is None else v)


def _clear_row(ws: Worksheet, r: int, c1: int, c2: int):
    for c in range(c1, c2 + 1):
        ws.cell(r, c, value="0")


def _find_row_by_col_text(ws: Worksheet, col: int, text: str) -> Optional[int]:
    t = "" if text is None else str(text).strip()
    if t == "":
        return None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        if str(v).strip() == t:
            return r
    return None


def _detect_s23_layout(ws_23: Worksheet):
    # ✅你已确认累计区只有 MT 布局
    mode = "MT"

    # 当月区 E-L
    S23_M = {
        "begin_qty": 5,       # E
        "begin_price": 6,     # F
        "purchase_qty": 7,    # G
        "purchase_price": 8,  # H
        "process_qty": 9,     # I
        "process_price": 10,  # J
        "end_qty": 11,        # K
        "end_price": 12,      # L
    }

    # ✅累计区 M-T：M/N 年初，O/P 采购，Q/R 加工，S/T 期末
    S23_Y = {
        "begin_qty": 13,       # M
        "begin_price": 14,     # N
        "purchase_qty": 15,    # O
        "purchase_price": 16,  # P
        "process_qty": 17,     # Q
        "process_price": 18,   # R
        "end_qty": 19,         # S
        "end_price": 20,       # T
    }
    y_clear_end = 20
    return mode, S23_M, S23_Y, y_clear_end


# =========================
# 上月2-3：标题定位取 K / JK / LM
# =========================
def _prev_k_by_rule(ws_prev_23: Optional[Worksheet], rule: str) -> Optional[float]:
    """上月2-3 K列（当月期末单价）某行（锚点用）"""
    if ws_prev_23 is None:
        return None

    def get_k(row: int) -> Optional[float]:
        return to_float(ws_prev_23.cell(row, 11).value)  # K=11

    if rule == "oil_main":
        r = _find_row_by_col_text(ws_prev_23, 3, "一、原油(一般贸易)")
        return None if r is None else get_k(r)

    if rule == "oil_import":
        r = _find_row_by_col_text(ws_prev_23, 3, "一、原油(一般贸易)")
        return None if r is None else get_k(r + 1)

    if rule == "oil_ocean":
        r_pl = _find_row_by_col_text(ws_prev_23, 3, "蓬莱原油\\中国")
        return None if r_pl is None else get_k(r_pl - 1)

    return None


def _prev_jk_qty_price(ws_prev_23: Optional[Worksheet], rule: str) -> Tuple[Optional[float], Optional[float]]:
    """上月2-3 J/K（当月期末数量/单价）用于回填本月 EF（锚点行用）"""
    if ws_prev_23 is None:
        return (None, None)

    def get_jk(row: int) -> Tuple[Optional[float], Optional[float]]:
        j = to_float(ws_prev_23.cell(row, 10).value)  # J=10
        k = to_float(ws_prev_23.cell(row, 11).value)  # K=11
        return (j, k)

    if rule == "intransit_main":
        r = _find_row_by_col_text(ws_prev_23, 3, "二、在途原油")
        return (None, None) if r is None else get_jk(r)

    if rule == "intransit_import":
        r = _find_row_by_col_text(ws_prev_23, 3, "二、在途原油")
        return (None, None) if r is None else get_jk(r + 1)

    if rule == "intransit_ocean":
        r_pl = _find_row_by_col_text(ws_prev_23, 3, "蓬莱原油\\中国")
        return (None, None) if r_pl is None else get_jk(r_pl - 1)

    if rule == "des_main":
        r = _find_row_by_col_text(ws_prev_23, 3, "三、DES/DAT/DAP在途（货权未转移）")
        return (None, None) if r is None else get_jk(r)

    if rule == "proc_main":
        r = _find_row_by_col_text(ws_prev_23, 3, "四、来料加工")
        return (None, None) if r is None else get_jk(r)

    return (None, None)


def _prev_year_begin(
    ws_prev_23: Optional[Worksheet],
    row_prev: Optional[int],
    current_month: Optional[int] = None
) -> Tuple[Optional[float], Optional[float]]:
    """上月2-3 -> 本月2-3 年初库存（数量/单价）（锚点用）

    逻辑：
      - 当本月为 2~12 月：取上月 L/M（年初库存）
      - 当本月为 1 月：上月应视为上一年度 12 月，取上月 J/K（期末库存）
    """
    if ws_prev_23 is None or row_prev is None:
        return (None, None)

    if current_month == 1:
        j = to_float(ws_prev_23.cell(row_prev, 10).value)  # J=10
        k = to_float(ws_prev_23.cell(row_prev, 11).value)  # K=11
        return (j, k)

    l = to_float(ws_prev_23.cell(row_prev, 12).value)  # L=12
    m = to_float(ws_prev_23.cell(row_prev, 13).value)  # M=13
    return (l, m)


# =========================
# 2-1/2-2：找“美元/桶”锚点行并读取
# =========================
def _read_usd_row(ws_cost: Worksheet) -> Optional[int]:
    for i in range(1, ws_cost.max_row + 1):
        v = ws_cost.cell(i, 1).value
        if v is None:
            continue
        if "美元/桶" in str(v).strip():
            return i
    return None


def _usd_map_for_oil_anchors(ws_cost: Worksheet) -> Optional[Dict[str, Dict[str, Optional[float]]]]:
    row_usd = _read_usd_row(ws_cost)
    if row_usd is None:
        return None

    def f(col: int) -> Optional[float]:
        return to_float(ws_cost.cell(row_usd, col).value)

    # H..P: H/I/J = 采购(一般/进/海), K/L/M = 加工, N/O/P = 期末
    return {
        "gen": {"purchase": f(8),  "process": f(11), "end": f(14)},
        "imp": {"purchase": f(9),  "process": f(12), "end": f(15)},
        "oce": {"purchase": f(10), "process": f(13), "end": f(16)},
    }


# =========================
# 2-1/2-2：读取锚点行数量（E/H/K/Q）
# =========================
def _read_anchor_qty(ws_cost: Worksheet, title: str) -> Dict[str, Optional[float]]:
    r = find_row_by_a(ws_cost, title)
    if r is None:
        return {"begin": None, "purchase": None, "process": None, "end": None}
    return {
        "begin": to_float(ws_cost.cell(r, 5).value),     # E
        "purchase": to_float(ws_cost.cell(r, 8).value),  # H
        "process": to_float(ws_cost.cell(r, 11).value),  # K
        "end": to_float(ws_cost.cell(r, 17).value),      # Q
    }


# =========================
# 明细行定位 (cat + name_norm) -> row
# =========================
def _build_detail_map(ws: Worksheet, allow_cats: Set[str]) -> Dict[Tuple[str, str], int]:
    m = {}
    for r in range(8, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue
        cat_s = "" if cat is None else str(cat).strip()
        if cat_s not in allow_cats:
            continue
        m[(cat_s, norm(name))] = r
    return m


def _collect_needed_detail_info(ws_21: Worksheet, ws_22: Worksheet, allow_cats: Set[str]) -> Dict[Tuple[str, str], Dict[str, Optional[str]]]:
    """从2-1/2-2收集“应存在于2-3”的明细物料集合（cat + name_norm）。

    返回：dict[(cat, name_norm)] = {"name": name_raw, "code": code_raw}
    优先保留 2-1 的 name/code；2-1没有时再用2-2。
    """
    need: Dict[Tuple[str, str], Dict[str, Optional[str]]] = {}

    def _take(ws: Worksheet):
        for r in range(8, ws.max_row + 1):
            cat = ws.cell(r, 1).value
            name = ws.cell(r, 4).value
            if name is None or str(name).strip() == "":
                continue
            cat_s = "" if cat is None else str(cat).strip()
            if cat_s not in allow_cats:
                continue
            name_raw = str(name).strip()
            key = (cat_s, norm(name_raw))
            if key in need:
                continue
            code = ws.cell(r, 3).value
            code_raw = None if code is None else str(code).strip().lstrip("0")
            need[key] = {"name": name_raw, "code": code_raw}

    _take(ws_21)
    _take(ws_22)
    return need


def _build_existing_detail_keys(ws_23: Worksheet, allow_cats: Set[str], anchor_title_set: Set[str]) -> Set[Tuple[str, str]]:
    out: Set[Tuple[str, str]] = set()
    for r in range(8, ws_23.max_row + 1):
        a_title = ws_23.cell(r, 1).value
        if a_title is not None and str(a_title).strip() in anchor_title_set:
            continue

        cat = ws_23.cell(r, 1).value
        name = ws_23.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue
        cat_s = "" if cat is None else str(cat).strip()
        if cat_s not in allow_cats:
            continue
        name_raw = str(name).strip()
        out.add((cat_s, norm(name_raw)))
    return out


def _ensure_new_details_in_23(ws_23: Worksheet, needed: Dict[Tuple[str, str], Dict[str, Optional[str]]],
                             allow_cats: Set[str], anchor_title_set: Set[str]):
    """✅修复：2-3里没有的新物料，需要先插入行，否则后续只会“对已有行填充”。"""

    # 2-3 各分类对应的“合计/锚点行”标题（A列）
    anchor_map_23 = {
        # 原油明细（A=进口原油/海洋原油）所在块的合计行
        "进口原油": "原油(一般贸易)-进口原油",
        "海洋原油": "原油(一般贸易)-海洋原油",
        "来料加工": "来料加工-合计",
        # 在途明细
        "在途原油-进口原油": "在途原油-进口原油-合计",
        "在途原油-海洋原油": "在途原油-海洋原油-合计",
        "DES/DAT/DAP在途（货权未转移）": "DES/DAT/DAP在途（货权未转移）-合计",
    }

    existed = _build_existing_detail_keys(ws_23, allow_cats, anchor_title_set)

    # 排序：保证插入稳定（先按分类，再按名称）
    todo: List[Tuple[str, str]] = []
    for (cat_s, name_n), info in needed.items():
        if cat_s not in allow_cats:
            continue
        if (cat_s, name_n) in existed:
            continue
        todo.append((cat_s, info.get("name") or ""))
    todo.sort(key=lambda x: (x[0], x[1]))

    # 每次插入都实时计算 ptr，避免行位移影响
    for cat_s, name_raw in todo:
        anchor_title = anchor_map_23.get(cat_s)
        if not anchor_title:
            continue

        # ✅模板结构：合计行在上、明细在下 => 插到该类别明细块的末尾
        ptr = find_insert_row_after_category_anchor(ws_23, anchor_title, detail_category_value=cat_s)
        if ptr is None:
            continue

        ws_23.insert_rows(ptr)
        ws_23.cell(ptr, 1, value=cat_s)

        # C列尽量写物料号（如果2-1/2-2里有）
        info = needed.get((cat_s, norm(name_raw))) or {}
        code_raw = info.get("code")
        if code_raw:
            ws_23.cell(ptr, 3, value=code_raw)

        ws_23.cell(ptr, 4, value=name_raw)

        # 其余列先填"-"，后续逻辑会再次清空并回填
        for c in range(5, ws_23.max_column + 1):
            ws_23.cell(ptr, c, value="-")

        existed.add((cat_s, norm(name_raw)))


# =========================
# 单价换算：元/吨 -> 美元/桶（按物料吨桶比）
# =========================
def _ratio_for_row(ws: Worksheet, row: int, ratio_map: Dict[Tuple[str, str], float]) -> Optional[float]:
    code = ws.cell(row, 3).value
    name = ws.cell(row, 4).value
    code_s = "" if code is None else str(code).strip().lstrip("0")
    name_s = norm(name)
    if not name_s:
        return None
    return ratio_map.get((code_s, name_s))


def _usd_per_bbl(price_rmb_per_ton: Optional[float], ratio: Optional[float], fx: Optional[float]) -> Optional[float]:
    if price_rmb_per_ton is None or ratio is None or fx is None:
        return None
    if ratio == 0 or fx == 0:
        return None
    return price_rmb_per_ton / ratio / fx


def _is_valid_qty(v) -> bool:
    x = to_float(v)
    return x is not None and x != 0


def _usd_per_bbl_safe(price_rmb_per_ton, qty, ratio, fx):
    """
    ✅新口径：数量为空/为'-'/为0，就不计算单价，返回 None -> 最终写 '-'
    """
    if not _is_valid_qty(qty):
        return None
    return _usd_per_bbl(price_rmb_per_ton, ratio, fx)


def _read_row_stage_from_cost(ws_cost: Worksheet, row: int, cat_for_intransit: str) -> Dict[str, Optional[float]]:
    """
    从 2-1/2-2 明细行读取：
    普通原油：E/F, H/I, K/L, Q/R
    在途：只取 Q/R
    """
    if cat_for_intransit in INTRANSIT_CATS:
        end_qty = to_float(ws_cost.cell(row, 17).value)  # Q
        end_prc = to_float(ws_cost.cell(row, 18).value)  # R
        return {
            "begin_qty": None, "begin_price": None,
            "purchase_qty": None, "purchase_price": None,
            "process_qty": None, "process_price": None,
            "end_qty": end_qty, "end_price": end_prc,   # ✅这里 end_price 是“元/吨”
        }

    return {
        "begin_qty": to_float(ws_cost.cell(row, 5).value),
        "begin_price": to_float(ws_cost.cell(row, 6).value),
        "purchase_qty": to_float(ws_cost.cell(row, 8).value),
        "purchase_price": to_float(ws_cost.cell(row, 9).value),
        "process_qty": to_float(ws_cost.cell(row, 11).value),
        "process_price": to_float(ws_cost.cell(row, 12).value),
        "end_qty": to_float(ws_cost.cell(row, 17).value),
        "end_price": to_float(ws_cost.cell(row, 18).value),
    }


# =========================
# 上月2-3：明细行定位（按C列产品名称精确匹配）
# =========================
def _find_prev_23_row_by_name(ws_prev_23: Optional[Worksheet], name_raw: str) -> Optional[int]:
    """
    上月2-3：C列为产品名称（你口径：锚点/明细都在C列）
    """
    if ws_prev_23 is None:
        return None
    t = "" if name_raw is None else str(name_raw).strip()
    if not t:
        return None
    for r in range(1, ws_prev_23.max_row + 1):
        v = ws_prev_23.cell(r, 3).value  # C列
        if v is None:
            continue
        if str(v).strip() == t:
            return r
    return None


# =========================
# 现场加权工具：Σ(qty*price)/Σqty
# =========================
def _weighted_price_from_rows(ws: Worksheet, rows: List[int], qty_col: int, price_col: int) -> Optional[float]:
    num = 0.0
    den = 0.0
    for r in rows:
        q = to_float(ws.cell(r, qty_col).value)
        p = to_float(ws.cell(r, price_col).value)
        if q is None or q == 0 or p is None:
            continue
        num += q * p
        den += q
    return None if den == 0 else (num / den)


def _collect_detail_rows_by_cat(ws: Worksheet, cat: str, anchor_title_set: Set[str]) -> List[int]:
    """
    收集 ws 中属于某 category 的明细行（排除锚点行/标题行）
    规则：A列=cat 且 D列有值；并且 A列不是锚点标题
    """
    out: List[int] = []
    for r in range(8, ws.max_row + 1):
        a = ws.cell(r, 1).value
        d = ws.cell(r, 4).value
        if d is None or str(d).strip() == "":
            continue
        a_s = "" if a is None else str(a).strip()
        if a_s in anchor_title_set:
            continue
        if a_s == cat:
            out.append(r)
    return out


# =========================
# 主函数
# =========================
def fill_sheet_2_3(
    ws_23: Worksheet,
    ws_21: Worksheet,
    ws_22: Worksheet,
    ws_prev_23: Optional[Worksheet],
    ratio_map: Dict[Tuple[str, str], float],
    fx_month: float,
    fx_ytd: float,
    current_month: Optional[int] = None,
    intransit_price_pack: Optional[dict] = None
):
    """
    intransit_price_pack:
      {
        "detail_price": {(cat_norm, name_norm): usd_per_bbl},  # 在途明细行期末单价(美元/桶)
        "cat_price": {cat_norm: usd_per_bbl},                  # 分类合计行期末单价(美元/桶)
        "intransit_total_price": usd_per_bbl                   # 旧字段保留，但✅不再用于2-3在途合计锚点（改为现场加权子合计）
      }

    ✅ 明细行回填口径（你最新要求）：
      - 本月2-3 EF ← 上月2-3 JK（同一产品名称行）
      - 本月2-3 MN：
          - current_month==1 => 取上月 JK（期末库存，作为本年年初）
          - 其他月份 => 取上月 LM（年初库存）

    ✅新增修复：
      - DES 明细美元/桶：pack找不到时 fallback 用(元/吨 ÷ 吨桶比 ÷ 汇率)
      - DES 合计美元/桶：汇总明细现场加权
      - 在途原油-合计美元/桶：必须按2-3两个子合计行现场加权
    """
    mode, S23_M, S23_Y, y_clear_end = _detect_s23_layout(ws_23)

    pack = intransit_price_pack or {"detail_price": {}, "cat_price": {}, "intransit_total_price": None}
    detail_price = pack.get("detail_price", {}) or {}
    cat_price = pack.get("cat_price", {}) or {}

    usd_21 = _usd_map_for_oil_anchors(ws_21)
    usd_22 = _usd_map_for_oil_anchors(ws_22)

    # =============== 1) 锚点合计行 ===============
    anchor_titles = [
        "原油(一般贸易)",
        "原油(一般贸易)-进口原油",
        "原油(一般贸易)-海洋原油",
        "在途原油-合计",
        "在途原油-进口原油-合计",
        "在途原油-海洋原油-合计",
        "DES/DAT/DAP在途（货权未转移）-合计",
        "来料加工-合计",
    ]
    anchor_title_set = set(anchor_titles)

    # =========================
    # 0) ✅先补齐2-3缺失的新物料行
    # =========================
    allow_cats = set(OIL_CATS) | set(INTRANSIT_CATS)
    needed_detail = _collect_needed_detail_info(ws_21, ws_22, allow_cats)
    _ensure_new_details_in_23(ws_23, needed_detail, allow_cats, anchor_title_set)

    prev_row_oil_main = _find_row_by_col_text(ws_prev_23, 3, "一、原油(一般贸易)") if ws_prev_23 else None
    prev_row_oil_imp = None if prev_row_oil_main is None else prev_row_oil_main + 1
    r_pl_prev = _find_row_by_col_text(ws_prev_23, 3, "蓬莱原油\\中国") if ws_prev_23 else None
    prev_row_oil_oce = None if r_pl_prev is None else r_pl_prev - 1

    for title in anchor_titles:
        r23 = find_row_by_a(ws_23, title)
        if r23 is None:
            continue

        # 清空当月 E..L，累计 M..T
        _clear_row(ws_23, r23, 5, 12)
        _clear_row(ws_23, r23, 13, y_clear_end)

        # ---- 原油三条锚点 ----
        if title in {"原油(一般贸易)", "原油(一般贸易)-进口原油", "原油(一般贸易)-海洋原油"}:
            # 当月数量：来自2-1锚点行（E/G/I/K）
            q21 = _read_anchor_qty(ws_21, title)
            _set(ws_23, r23, S23_M["begin_qty"], q21["begin"])
            _set(ws_23, r23, S23_M["purchase_qty"], q21["purchase"])
            _set(ws_23, r23, S23_M["process_qty"], q21["process"])
            _set(ws_23, r23, S23_M["end_qty"], q21["end"])

            # 当月期初单价F：上月2-3 K列规则
            if title == "原油(一般贸易)":
                f = _prev_k_by_rule(ws_prev_23, "oil_main")
            elif title == "原油(一般贸易)-进口原油":
                f = _prev_k_by_rule(ws_prev_23, "oil_import")
            else:
                f = _prev_k_by_rule(ws_prev_23, "oil_ocean")
            _set(ws_23, r23, S23_M["begin_price"], f)

            # 当月采购/加工/期末单价：取2-1“美元/桶”锚点行（H/J/L）
            if usd_21 is not None:
                if title == "原油(一般贸易)":
                    _set(ws_23, r23, S23_M["purchase_price"], usd_21["gen"]["purchase"])
                    _set(ws_23, r23, S23_M["process_price"],  usd_21["gen"]["process"])
                    _set(ws_23, r23, S23_M["end_price"],      usd_21["gen"]["end"])
                elif title == "原油(一般贸易)-进口原油":
                    _set(ws_23, r23, S23_M["purchase_price"], usd_21["imp"]["purchase"])
                    _set(ws_23, r23, S23_M["process_price"],  usd_21["imp"]["process"])
                    _set(ws_23, r23, S23_M["end_price"],      usd_21["imp"]["end"])
                else:
                    _set(ws_23, r23, S23_M["purchase_price"], usd_21["oce"]["purchase"])
                    _set(ws_23, r23, S23_M["process_price"],  usd_21["oce"]["process"])
                    _set(ws_23, r23, S23_M["end_price"],      usd_21["oce"]["end"])

            # 累计年初 M/N：上月2-3 L/M；若本月=1，则用上月 J/K（期末）当年初
            if title == "原油(一般贸易)" and prev_row_oil_main is not None:
                m, n = _prev_year_begin(ws_prev_23, prev_row_oil_main, current_month)
            elif title == "原油(一般贸易)-进口原油" and prev_row_oil_imp is not None:
                m, n = _prev_year_begin(ws_prev_23, prev_row_oil_imp, current_month)
            elif title == "原油(一般贸易)-海洋原油" and prev_row_oil_oce is not None:
                m, n = _prev_year_begin(ws_prev_23, prev_row_oil_oce, current_month)
            else:
                m, n = (None, None)
            _set(ws_23, r23, S23_Y["begin_qty"], m)
            _set(ws_23, r23, S23_Y["begin_price"], n)

            # 累计采购/加工/期末数量：来自2-2锚点行（O/Q/S）
            q22 = _read_anchor_qty(ws_22, title)
            _set(ws_23, r23, S23_Y["purchase_qty"], q22["purchase"])  # O <- 2-2 H
            _set(ws_23, r23, S23_Y["process_qty"],  q22["process"])   # Q <- 2-2 K
            _set(ws_23, r23, S23_Y["end_qty"],      q22["end"])       # S <- 2-2 Q

            # 累计采购/加工/期末单价：来自2-2“美元/桶”锚点行（P/R/T）
            if usd_22 is not None:
                if title == "原油(一般贸易)":
                    _set(ws_23, r23, S23_Y["purchase_price"], usd_22["gen"]["purchase"])
                    _set(ws_23, r23, S23_Y["process_price"],  usd_22["gen"]["process"])
                    _set(ws_23, r23, S23_Y["end_price"],      usd_22["gen"]["end"])
                elif title == "原油(一般贸易)-进口原油":
                    _set(ws_23, r23, S23_Y["purchase_price"], usd_22["imp"]["purchase"])
                    _set(ws_23, r23, S23_Y["process_price"],  usd_22["imp"]["process"])
                    _set(ws_23, r23, S23_Y["end_price"],      usd_22["imp"]["end"])
                else:
                    _set(ws_23, r23, S23_Y["purchase_price"], usd_22["oce"]["purchase"])
                    _set(ws_23, r23, S23_Y["process_price"],  usd_22["oce"]["process"])
                    _set(ws_23, r23, S23_Y["end_price"],      usd_22["oce"]["end"])
            continue

        # ---- 在途原油-合计 ----
        # ✅注意：这里不再直接写 end_price（L/T），后面会按子合计行现场加权计算
        if title == "在途原油-合计":
            # EF 回填：来自上月2-3的JK（锚点用规则定位）
            j, k = _prev_jk_qty_price(ws_prev_23, "intransit_main")
            _set(ws_23, r23, S23_M["begin_qty"], j)
            _set(ws_23, r23, S23_M["begin_price"], k)

            # 当月期末数量 K：来自2-1锚点 Q
            q21 = _read_anchor_qty(ws_21, title)
            _set(ws_23, r23, S23_M["end_qty"], q21["end"])

            # 累计期末数量 S：来自2-2锚点 Q
            q22 = _read_anchor_qty(ws_22, title)
            _set(ws_23, r23, S23_Y["end_qty"], q22["end"])
            continue

        if title == "在途原油-进口原油-合计":
            j, k = _prev_jk_qty_price(ws_prev_23, "intransit_import")
            _set(ws_23, r23, S23_M["begin_qty"], j)
            _set(ws_23, r23, S23_M["begin_price"], k)

            q21 = _read_anchor_qty(ws_21, title)
            _set(ws_23, r23, S23_M["end_qty"], q21["end"])
            _set(ws_23, r23, S23_M["end_price"], cat_price.get(norm("在途原油-进口原油")))

            q22 = _read_anchor_qty(ws_22, title)
            _set(ws_23, r23, S23_Y["end_qty"], q22["end"])
            _set(ws_23, r23, S23_Y["end_price"], cat_price.get(norm("在途原油-进口原油")))
            continue

        if title == "在途原油-海洋原油-合计":
            j, k = _prev_jk_qty_price(ws_prev_23, "intransit_ocean")
            _set(ws_23, r23, S23_M["begin_qty"], j)
            _set(ws_23, r23, S23_M["begin_price"], k)

            q21 = _read_anchor_qty(ws_21, title)
            _set(ws_23, r23, S23_M["end_qty"], q21["end"])
            _set(ws_23, r23, S23_M["end_price"], cat_price.get(norm("在途原油-海洋原油")))

            q22 = _read_anchor_qty(ws_22, title)
            _set(ws_23, r23, S23_Y["end_qty"], q22["end"])
            _set(ws_23, r23, S23_Y["end_price"], cat_price.get(norm("在途原油-海洋原油")))
            continue

        # ---- DES 合计：✅改为后面“汇总明细现场加权”来写 end_price ----
        if title == DES_TOTAL_TITLE:
            j, k = _prev_jk_qty_price(ws_prev_23, "des_main")
            _set(ws_23, r23, S23_M["begin_qty"], j)
            _set(ws_23, r23, S23_M["begin_price"], k)

            q21 = _read_anchor_qty(ws_21, title)
            _set(ws_23, r23, S23_M["end_qty"], q21["end"])

            q22 = _read_anchor_qty(ws_22, title)
            _set(ws_23, r23, S23_Y["end_qty"], q22["end"])
            continue

        # ---- 来料加工-合计：保持原逻辑 ----
        if title == "来料加工-合计":
            j, k = _prev_jk_qty_price(ws_prev_23, "proc_main")
            _set(ws_23, r23, S23_M["begin_qty"], j)
            _set(ws_23, r23, S23_M["begin_price"], k)

            q21 = _read_anchor_qty(ws_21, title)
            _set(ws_23, r23, S23_M["end_qty"], q21["end"])

            proc_end_price = cat_price.get(norm("来料加工"))
            _set(ws_23, r23, S23_M["end_price"], proc_end_price)

            q22 = _read_anchor_qty(ws_22, title)
            _set(ws_23, r23, S23_Y["end_qty"], q22["end"])
            _set(ws_23, r23, S23_Y["end_price"], proc_end_price)
            continue

    # =============== 2) 明细行：EF/MN从上月回填；其余按2-1/2-2填入 ===============
    map21 = _build_detail_map(ws_21, allow_cats)
    map22 = _build_detail_map(ws_22, allow_cats)

    for r in range(8, ws_23.max_row + 1):
        # 跳过锚点行，避免被当明细重填
        a_title = ws_23.cell(r, 1).value
        if a_title is not None and str(a_title).strip() in anchor_title_set:
            continue

        cat = ws_23.cell(r, 1).value
        name = ws_23.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue

        cat_s = "" if cat is None else str(cat).strip()
        if cat_s not in allow_cats:
            continue

        name_raw = str(name).strip()
        name_norm = norm(name_raw)

        # ---------- 2.1 回填 EF / MN（上月2-3：JK / LM or JK） ----------
        prev_r = _find_prev_23_row_by_name(ws_prev_23, name_raw)

        # 先把 EF/MN 设为 0（内部用 0/None 区分写入），能匹配上月再覆盖
        _set(ws_23, r, S23_M["begin_qty"], None)     # E
        _set(ws_23, r, S23_M["begin_price"], None)   # F
        _set(ws_23, r, S23_Y["begin_qty"], None)     # M
        _set(ws_23, r, S23_Y["begin_price"], None)   # N

        if prev_r is not None:
            # 本月 EF ← 上月 JK
            _set(ws_23, r, S23_M["begin_qty"], to_float(ws_prev_23.cell(prev_r, 10).value))    # J
            _set(ws_23, r, S23_M["begin_price"], to_float(ws_prev_23.cell(prev_r, 11).value))  # K

            # 本月 MN：默认取上月 LM（年初库存）；但若当前月=1，则取上月 JK（期末库存）
            if current_month == 1:
                _set(ws_23, r, S23_Y["begin_qty"], to_float(ws_prev_23.cell(prev_r, 10).value))    # J
                _set(ws_23, r, S23_Y["begin_price"], to_float(ws_prev_23.cell(prev_r, 11).value))  # K
            else:
                _set(ws_23, r, S23_Y["begin_qty"], to_float(ws_prev_23.cell(prev_r, 12).value))    # L
                _set(ws_23, r, S23_Y["begin_price"], to_float(ws_prev_23.cell(prev_r, 13).value))  # M

        # ---------- 2.2 清其它区域，但保留 EF/MN ----------
        _clear_row(ws_23, r, 7, 12)             # 当月：清 G..L
        _clear_row(ws_23, r, 15, y_clear_end)   # 累计：清 O..T

        # ---------- 2.3 当月：来自 2-1 明细 ----------
        r21 = map21.get((cat_s, name_norm))
        d21 = _read_row_stage_from_cost(ws_21, r21, cat_s) if r21 is not None else None

        if d21 is not None:
            _set(ws_23, r, S23_M["purchase_qty"], d21["purchase_qty"])
            _set(ws_23, r, S23_M["process_qty"], d21["process_qty"])
            _set(ws_23, r, S23_M["end_qty"], d21["end_qty"])

        if cat_s in INTRANSIT_CATS:
            prc = detail_price.get((norm(cat_s), name_norm))

            if prc is None and cat_s == "DES/DAT/DAP在途（货权未转移）":
                base_name = re.sub(r"（.*?）$", "", str(name_raw).strip())
                prc = detail_price.get((norm(cat_s), norm(base_name)))

            # 只有在 pack 没取到时，才依赖 2-1 做 fallback
            if prc is None:
                if d21 is None or not _is_valid_qty(d21["end_qty"]):
                    prc = None
                else:
                    ratio = _ratio_for_row(ws_21, r21, ratio_map)
                    prc = _usd_per_bbl(d21["end_price"], ratio, fx_month)

            _set(ws_23, r, S23_M["end_price"], prc)
        else:
            if d21 is not None:
                ratio = _ratio_for_row(ws_21, r21, ratio_map)
                _set(ws_23, r, S23_M["purchase_price"],
                     _usd_per_bbl_safe(d21["purchase_price"], d21["purchase_qty"], ratio, fx_month))
                _set(ws_23, r, S23_M["process_price"],
                     _usd_per_bbl_safe(d21["process_price"], d21["process_qty"], ratio, fx_month))
                _set(ws_23, r, S23_M["end_price"],
                     _usd_per_bbl_safe(d21["end_price"], d21["end_qty"], ratio, fx_month))

        # ---------- 2.4 累计：来自 2-2 明细 ----------
        r22 = map22.get((cat_s, name_norm))
        d22 = _read_row_stage_from_cost(ws_22, r22, cat_s) if r22 is not None else None

        if d22 is not None:
            _set(ws_23, r, S23_Y["purchase_qty"], d22["purchase_qty"])
            _set(ws_23, r, S23_Y["process_qty"], d22["process_qty"])
            _set(ws_23, r, S23_Y["end_qty"], d22["end_qty"])

        if cat_s in INTRANSIT_CATS:
            prc = detail_price.get((norm(cat_s), name_norm))

            if prc is None and cat_s == "DES/DAT/DAP在途（货权未转移）":
                base_name = re.sub(r"（.*?）$", "", str(name_raw).strip())
                prc = detail_price.get((norm(cat_s), norm(base_name)))

            if prc is None:
                if d22 is None or not _is_valid_qty(d22["end_qty"]):
                    prc = None
                else:
                    ratio = _ratio_for_row(ws_22, r22, ratio_map)
                    prc = _usd_per_bbl(d22["end_price"], ratio, fx_ytd)

            _set(ws_23, r, S23_Y["end_price"], prc)
        else:
            if d22 is not None:
                ratio = _ratio_for_row(ws_22, r22, ratio_map)
                _set(ws_23, r, S23_Y["purchase_price"],
                     _usd_per_bbl_safe(d22["purchase_price"], d22["purchase_qty"], ratio, fx_ytd))
                _set(ws_23, r, S23_Y["process_price"],
                     _usd_per_bbl_safe(d22["process_price"], d22["process_qty"], ratio, fx_ytd))
                _set(ws_23, r, S23_Y["end_price"],
                     _usd_per_bbl_safe(d22["end_price"], d22["end_qty"], ratio, fx_ytd))

    # =============== 3) ✅现场加权修复区 ===============
    # 3.1 在途原油-合计：必须按2-3两个子合计行现场加权（当月K/L + 累计S/T）
    r_it_total = find_row_by_a(ws_23, INTRANSIT_TOTAL_TITLE)
    r_it_imp = find_row_by_a(ws_23, INTRANSIT_IMP_TITLE)
    r_it_oce = find_row_by_a(ws_23, INTRANSIT_OCE_TITLE)

    if r_it_total is not None and r_it_imp is not None and r_it_oce is not None:
        # ---- 当月（K/L）----
        k_imp = to_float(ws_23.cell(r_it_imp, S23_M["end_qty"]).value)
        l_imp = to_float(ws_23.cell(r_it_imp, S23_M["end_price"]).value)

        k_oce = to_float(ws_23.cell(r_it_oce, S23_M["end_qty"]).value)
        l_oce = to_float(ws_23.cell(r_it_oce, S23_M["end_price"]).value)

        k_imp = 0.0 if k_imp is None else k_imp
        k_oce = 0.0 if k_oce is None else k_oce
        k_total = k_imp + k_oce

        if k_total == 0:
            l_total = None
        else:
            num = 0.0
            has = False
            if k_imp != 0 and l_imp is not None:
                num += k_imp * l_imp
                has = True
            if k_oce != 0 and l_oce is not None:
                num += k_oce * l_oce
                has = True
            l_total = None if not has else (num / k_total)

        _set(ws_23, r_it_total, S23_M["end_qty"], k_total)
        _set(ws_23, r_it_total, S23_M["end_price"], l_total)

        # ---- 累计（S/T）----
        s_imp = to_float(ws_23.cell(r_it_imp, S23_Y["end_qty"]).value)
        t_imp = to_float(ws_23.cell(r_it_imp, S23_Y["end_price"]).value)

        s_oce = to_float(ws_23.cell(r_it_oce, S23_Y["end_qty"]).value)
        t_oce = to_float(ws_23.cell(r_it_oce, S23_Y["end_price"]).value)

        s_imp = 0.0 if s_imp is None else s_imp
        s_oce = 0.0 if s_oce is None else s_oce
        s_total = s_imp + s_oce

        if s_total == 0:
            t_total = None
        else:
            num = 0.0
            has = False
            if s_imp != 0 and t_imp is not None:
                num += s_imp * t_imp
                has = True
            if s_oce != 0 and t_oce is not None:
                num += s_oce * t_oce
                has = True
            t_total = None if not has else (num / s_total)

        _set(ws_23, r_it_total, S23_Y["end_qty"], s_total)
        _set(ws_23, r_it_total, S23_Y["end_price"], t_total)

    # 3.2 DES 合计：✅汇总明细现场加权（当月K/L + 累计S/T）
    r_des_total = find_row_by_a(ws_23, DES_TOTAL_TITLE)
    if r_des_total is not None:
        des_detail_rows = _collect_detail_rows_by_cat(ws_23, DES_CAT, anchor_title_set)

        # 当月：L = Σ(K*L)/ΣK
        des_l = _weighted_price_from_rows(ws_23, des_detail_rows, S23_M["end_qty"], S23_M["end_price"])
        _set(ws_23, r_des_total, S23_M["end_price"], des_l)

        # 累计：T = Σ(S*T)/ΣS
        des_t = _weighted_price_from_rows(ws_23, des_detail_rows, S23_Y["end_qty"], S23_Y["end_price"])
        _set(ws_23, r_des_total, S23_Y["end_price"], des_t)

    # =============== 4) “合计”锚点行：当月K/L + 累计S/T ===============
    r_total = find_row_by_a(ws_23, "合计")
    r_anchor_general = find_row_by_a(ws_23, "原油(一般贸易)")
    r_anchor_intransit = find_row_by_a(ws_23, "在途原油-合计")

    if r_total is not None and r_anchor_general is not None and r_anchor_intransit is not None:
        gen_k = to_float(ws_23.cell(r_anchor_general, S23_M["end_qty"]).value)
        gen_l = to_float(ws_23.cell(r_anchor_general, S23_M["end_price"]).value)

        it_k = to_float(ws_23.cell(r_anchor_intransit, S23_M["end_qty"]).value)
        it_l = to_float(ws_23.cell(r_anchor_intransit, S23_M["end_price"]).value)

        gen_k = 0.0 if gen_k is None else gen_k
        it_k = 0.0 if it_k is None else it_k
        k_total = gen_k + it_k

        if k_total == 0:
            l_total = None
        else:
            numerator = 0.0
            has_term = False
            if gen_k != 0 and gen_l is not None:
                numerator += gen_k * gen_l
                has_term = True
            if it_k != 0 and it_l is not None:
                numerator += it_k * it_l
                has_term = True
            l_total = None if not has_term else (numerator / k_total)

        _set(ws_23, r_total, S23_M["end_qty"], k_total)     # 当月K
        _set(ws_23, r_total, S23_M["end_price"], l_total)   # 当月L
        _set(ws_23, r_total, S23_Y["end_qty"], k_total)     # 累计S（你当前模板“合计”累计等同展示）
        _set(ws_23, r_total, S23_Y["end_price"], l_total)   # 累计T
