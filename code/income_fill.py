# income_fill.py
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet
from utils import to_float  # 你原来引用保留（当前文件里没用到不影响）


def guess_bare_tax_sheet_name(sheetnames: list[str]) -> str:
    """在产销存文件中猜测“裸税价/裸价税/祼税价”sheet名。"""
    candidates = [
        "裸价税", "裸税价", "裸价", "裸税", "裸价税价", "裸税价表",
        "祼税价", "祼价税", "祼税",
    ]
    for name in candidates:
        if name in sheetnames:
            return name

    # 容错：包含关键字的sheet
    for n in sheetnames:
        s = str(n)
        if ("税价" in s or "价税" in s) and ("裸" in s or "祼" in s):
            return n
    for n in sheetnames:
        s = str(n)
        if ("裸" in s or "祼" in s) and "税" in s:
            return n

    raise ValueError(f"产销存文件中找不到裸税价/裸价税/祼税价sheet，当前sheet={sheetnames}")


def _clean_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0").strip()


def _set(ws: Worksheet, r: int, c: int, v):
    # 需求：收入sheet缺失值用 0（而不是 "-"），便于后续在Excel中直接用公式计算。
    ws.cell(r, c, value=0 if v is None else v)


def _unit_price(amt_yuan: float | None, qty_ton: float | None):
    # 需求：单价缺失或数量为0时，写0（避免"-"导致公式不可用）
    if amt_yuan is None or qty_ton is None or qty_ton == 0:
        return 0.0
    return amt_yuan / qty_ton


# def _abs_pair(q: float | None, a: float | None):
#     if q is None and a is None:
#         return (None, None)
#     q2 = None if q is None else abs(float(q))
#     a2 = None if a is None else abs(float(a))
#     return (q2, a2)
def _neg_pair(q: float | None, a: float | None):
    if q is None and a is None:
        return (None, None)
    q2 = None if q is None else -float(q)
    a2 = None if a is None else -float(a)
    return (q2, a2)

def _normalize_sales_pack(pack: dict) -> dict:
    by_code = pack.get("by_code") or {}
    if not by_code:
        return pack

    new_by: dict[str, dict] = {}
    for raw_code, rec in by_code.items():
        code = _clean_code(raw_code)
        if not code or rec is None:
            continue

        dst = new_by.get(code)
        if dst is None:
            dst = {"desc": rec.get("desc", ""), "monthly": {}}
            new_by[code] = dst

        if not dst.get("desc") and rec.get("desc"):
            dst["desc"] = rec.get("desc", "")

        m1 = dst.setdefault("monthly", {})
        m2 = rec.get("monthly") or {}
        for ym, mm in m2.items():
            if not mm:
                continue
            q = float(mm.get("qty_ton", 0.0) or 0.0)
            a = float(mm.get("amt_yuan", 0.0) or 0.0)
            cur = m1.get(ym)
            if cur is None:
                m1[ym] = {"qty_ton": q, "amt_yuan": a}
            else:
                cur["qty_ton"] = float(cur.get("qty_ton", 0.0) or 0.0) + q
                cur["amt_yuan"] = float(cur.get("amt_yuan", 0.0) or 0.0) + a

    new_pack = dict(pack)
    new_pack["by_code"] = new_by
    return new_pack


def _get_month(pack: dict, code: str, y: int, m: int):
    rec = (pack.get("by_code") or {}).get(code)
    if not rec:
        return (None, None)
    mm = rec.get("monthly", {}).get((y, m))
    if not mm:
        return (None, None)
    return (float(mm.get("qty_ton", 0.0)), float(mm.get("amt_yuan", 0.0)))


def _get_ytd(pack: dict, code: str, y: int, month_to: int):
    qty = 0.0
    amt = 0.0
    has = False
    for m in range(1, month_to + 1):
        q, a = _get_month(pack, code, y, m)
        if q is None and a is None:
            continue
        has = True
        qty += (q or 0.0)
        amt += (a or 0.0)
    return (qty, amt) if has else (None, None)


def _find_last_material_row(ws: Worksheet, start_row: int = 4) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)  # B
        if code:
            last = r
    return last


def _build_row_map(ws: Worksheet, start_row: int = 4) -> dict[str, int]:
    m = {}
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)
        if code:
            m[code] = r
    return m


def _renumber_seq(ws: Worksheet, start_row: int = 4):
    seq = 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)
        if not code:
            continue
        ws.cell(r, 1, value=seq)  # A
        seq += 1


def fill_income_ws(
    ws_income: Worksheet,
    sales_pack: dict,
    product_month_name_map: dict[str, str] | None = None,
):
    """
    产销存-收入表（sheet=收入）列结构（你给的口径）：
      D-F : 当月（D万吨 / E元/吨 / F万元）
      G-I : 当年累计到当月（G万吨 / H元/吨 / I万元）
      J-L : 当年累计到当月（J吨 / K元/吨 / L元）
      M-O : 1月（吨/元/元）
      ...
      AT-AV : 12月（吨/元/元）
      AW-AY : 上年12月（吨/元/元）
    第3行：合计；第4行开始：物料
    新增物料：插到最后一个物料号下方

    ✅ 修复：
    - 底表是支出负数：写入统一转正（包括月度块、累计、合计）
    - by_code key 统一清洗，避免物料行漏填（同 4-1 的 83005952 问题）
    """
    sales_pack = _normalize_sales_pack(sales_pack)

    year = sales_pack.get("year")
    prev_year = sales_pack.get("prev_year")
    cm = int(sales_pack.get("current_month") or 0)
    if not year or not (1 <= cm <= 12):
        return []

    by_code = sales_pack.get("by_code") or {}
    if not by_code:
        return []

    start_row = 4
    row_map = _build_row_map(ws_income, start_row=start_row)
    existed_codes = set(row_map.keys())
    all_codes = sorted(by_code.keys())

    def _name_of(code: str) -> str:
        """新增物料名称：优先用产品当月数据（A=物料号，B=物料名称），兜底用 sales_pack 的 desc。"""
        if product_month_name_map:
            n = product_month_name_map.get(code)
            if n:
                return n
        return by_code.get(code, {}).get("desc", "")

    # 插入新增物料
    new_codes = [c for c in all_codes if c not in existed_codes]
    if new_codes:
        last_row = _find_last_material_row(ws_income, start_row=start_row)
        ptr = last_row + 1 if last_row >= start_row else start_row

        for code in new_codes:
            ws_income.insert_rows(ptr)
            ws_income.cell(ptr, 2, value=code)  # B 物料号
            ws_income.cell(ptr, 3, value=_name_of(code))  # C 描述
            ptr += 1

        row_map = _build_row_map(ws_income, start_row=start_row)

    # 若模板已存在物料但名称为空，也用产品当月数据补齐一次（不影响已有正确名称）
    if product_month_name_map:
        for code, r in row_map.items():
            cur = ws_income.cell(r, 3).value
            if cur is None or str(cur).strip() == "":
                n = product_month_name_map.get(code)
                if n:
                    ws_income.cell(r, 3, value=n)

    # 列定位
    COL = {
        "m_qty_wan": 4, "m_price": 5, "m_amt_wan": 6,     # D/E/F
        "y_qty_wan": 7, "y_price": 8, "y_amt_wan": 9,     # G/H/I
        "y_qty_ton": 10, "y_price_ton": 11, "y_amt_yuan": 12,  # J/K/L
        "m1_start": 13,          # M
        "prev12_start": 49,      # AW
        "end_col": 51,           # AY
    }

    def month_block_start(m: int) -> int:
        # 1月从M=13开始，每月3列
        return COL["m1_start"] + (m - 1) * 3

    # 填明细
    for code, r in row_map.items():
        # 清 D..AY
        for c in range(4, COL["end_col"] + 1):
            ws_income.cell(r, c, value=0)

        # 当月（转正）
        q_m, a_m = _get_month(sales_pack, code, year, cm)
        q_m, a_m = _neg_pair(q_m, a_m)
        if q_m is not None:
            _set(ws_income, r, COL["m_qty_wan"], q_m / 10000.0)
            _set(ws_income, r, COL["m_amt_wan"], a_m / 10000.0)
            _set(ws_income, r, COL["m_price"], _unit_price(a_m, q_m))

        # 累计到当月（转正）
        q_y, a_y = _get_ytd(sales_pack, code, year, cm)
        q_y, a_y = _neg_pair(q_y, a_y)
        if q_y is not None:
            _set(ws_income, r, COL["y_qty_wan"], q_y / 10000.0)
            _set(ws_income, r, COL["y_amt_wan"], a_y / 10000.0)
            _set(ws_income, r, COL["y_price"], _unit_price(a_y, q_y))

            _set(ws_income, r, COL["y_qty_ton"], q_y)
            _set(ws_income, r, COL["y_amt_yuan"], a_y)
            _set(ws_income, r, COL["y_price_ton"], _unit_price(a_y, q_y))

        # 1~12月（月度块：转正）
        for mm in range(1, 13):
            sc = month_block_start(mm)
            q, a = _get_month(sales_pack, code, year, mm)
            q, a = _neg_pair(q, a)
            if q is None:
                continue
            _set(ws_income, r, sc, q)           # qty ton
            _set(ws_income, r, sc + 2, a)       # amt yuan
            _set(ws_income, r, sc + 1, _unit_price(a, q))

        # 上年12月（转正）
        q_p12, a_p12 = _get_month(sales_pack, code, prev_year, 12)
        q_p12, a_p12 = _neg_pair(q_p12, a_p12)
        if q_p12 is not None:
            sc = COL["prev12_start"]
            _set(ws_income, r, sc, q_p12)
            _set(ws_income, r, sc + 2, a_p12)
            _set(ws_income, r, sc + 1, _unit_price(a_p12, q_p12))

    # 合计行（第3行）
    total_row = 3
    for c in range(4, COL["end_col"] + 1):
        ws_income.cell(total_row, c, value=0)

    def sum_codes_month(y: int, m: int):
        tq = 0.0
        ta = 0.0
        has = False
        for code in all_codes:
            q, a = _get_month(sales_pack, code, y, m)
            if q is None:
                continue
            has = True
            tq += q
            ta += a
        return (tq, ta) if has else (None, None)

    # 当月合计（转正）
    tq, ta = sum_codes_month(year, cm)
    tq, ta = _neg_pair(tq, ta)
    if tq is not None:
        _set(ws_income, total_row, COL["m_qty_wan"], tq / 10000.0)
        _set(ws_income, total_row, COL["m_amt_wan"], ta / 10000.0)
        _set(ws_income, total_row, COL["m_price"], _unit_price(ta, tq))

    # 累计合计（转正）
    tqy = 0.0
    tay = 0.0
    hasy = False
    for mm in range(1, cm + 1):
        q, a = sum_codes_month(year, mm)
        if q is None:
            continue
        hasy = True
        tqy += q
        tay += a
    if hasy:
        tqy, tay = _neg_pair(tqy, tay)
        _set(ws_income, total_row, COL["y_qty_wan"], tqy / 10000.0)
        _set(ws_income, total_row, COL["y_amt_wan"], tay / 10000.0)
        _set(ws_income, total_row, COL["y_price"], _unit_price(tay, tqy))

        _set(ws_income, total_row, COL["y_qty_ton"], tqy)
        _set(ws_income, total_row, COL["y_amt_yuan"], tay)
        _set(ws_income, total_row, COL["y_price_ton"], _unit_price(tay, tqy))

    # 每月合计（转正）
    for mm in range(1, 13):
        sc = month_block_start(mm)
        q, a = sum_codes_month(year, mm)
        q, a = _neg_pair(q, a)
        if q is None:
            continue
        _set(ws_income, total_row, sc, q)
        _set(ws_income, total_row, sc + 2, a)
        _set(ws_income, total_row, sc + 1, _unit_price(a, q))

    # 上年12月合计（转正）
    q, a = sum_codes_month(prev_year, 12)
    q, a = _neg_pair(q, a)
    if q is not None:
        sc = COL["prev12_start"]
        _set(ws_income, total_row, sc, q)
        _set(ws_income, total_row, sc + 2, a)
        _set(ws_income, total_row, sc + 1, _unit_price(a, q))

    _renumber_seq(ws_income, start_row=start_row)

    # 返回：本次新增插入到“收入”sheet 的物料号（用于 runner 同步补裸税价sheet）
    return new_codes


def _build_row_map_by_code(ws: Worksheet, *, start_row: int, code_col: int) -> dict[str, int]:
    """扫描到底，按 clean_code 建 row_map（与收入sheet补行口径一致）。"""
    m: dict[str, int] = {}
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, code_col).value)
        if code:
            m[code] = r
    return m


def _find_last_code_row(ws: Worksheet, *, start_row: int, code_col: int) -> int:
    """找最后一个 code 不为空的行（与收入sheet补行口径一致）。"""
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, code_col).value)
        if code:
            last = r
    return last


def ensure_new_codes_in_bare_tax_ws(
    ws_bare: Worksheet,
    new_codes: list[str],
    *,
    start_row: int = 4,
    code_col: int = 2,   # B
    name_col: int = 3,   # C
    vat_col: int = 4,    # D
    cons_col: int = 5,   # E
    name_map: dict[str, str] | None = None,
):
    """将新增物料号同步补到“裸税价/裸价税/祼税价”sheet。

    设计目标：插入逻辑与“收入”sheet一致：
    - 以“物料号清洗后的值”为唯一标准去重
    - 只追加到末尾（最后一个物料号行的下方）
    - VAT/消费税标准留空（None），便于人工在Excel中筛选空值补填
    """

    if not new_codes:
        return

    row_map = _build_row_map_by_code(ws_bare, start_row=start_row, code_col=code_col)

    # new_codes 里也可能有重复/格式差异：这里统一 clean + 去重
    todo: list[str] = []
    seen: set[str] = set()
    for code in new_codes:
        cc = _clean_code(code)
        if not cc or cc in seen:
            continue
        seen.add(cc)
        if cc in row_map:
            continue
        todo.append(cc)

    if not todo:
        return

    last_row = _find_last_code_row(ws_bare, start_row=start_row, code_col=code_col)
    ptr = last_row + 1 if last_row >= start_row else start_row

    for cc in todo:
        if cc in row_map:
            continue
        ws_bare.insert_rows(ptr)
        ws_bare.cell(ptr, code_col, value=cc)

        name = ""
        if name_map:
            name = name_map.get(cc) or ""
        ws_bare.cell(ptr, name_col, value=name)

        ws_bare.cell(ptr, vat_col, value=None)
        ws_bare.cell(ptr, cons_col, value=None)

        row_map[cc] = ptr
        ptr += 1

    # 返回：本次真正新增到“收入”sheet中的物料号（用于runner同步到裸税价sheet）
    return new_codes


def ensure_new_codes_in_bare_tax_ws(
    ws_bare: Worksheet,
    new_codes: list[str],
    *,
    start_row: int = 4,
    code_col: int = 2,  # B
    name_col: int = 3,  # C
    vat_col: int = 4,   # D
    cons_col: int = 5,  # E
    name_map: dict[str, str] | None = None,
):
    """
    将“收入”sheet新增的物料号，同步追加到“裸税价/裸价税/祼税价”sheet。

    规则（按收入sheet插入逻辑一致）：
    - 以“物料号（清洗后）”为唯一标准去重
    - 只在最后一个有物料号的行下方顺序追加
    - 增值税率/消费税标准留空，方便人工打开表格补充
    """

    if not new_codes:
        return

    def _build_row_map_by_code() -> dict[str, int]:
        m = {}
        for r in range(start_row, ws_bare.max_row + 1):
            code = _clean_code(ws_bare.cell(r, code_col).value)
            if code:
                m[code] = r
        return m

    row_map = _build_row_map_by_code()

    # clean + 去重 + 过滤已存在
    todo: list[str] = []
    seen: set[str] = set()
    for code in new_codes:
        cc = _clean_code(code)
        if not cc or cc in seen:
            continue
        seen.add(cc)
        if cc in row_map:
            continue
        todo.append(cc)

    if not todo:
        return

    # 末尾追加
    last = start_row - 1
    for r in range(start_row, ws_bare.max_row + 1):
        c = _clean_code(ws_bare.cell(r, code_col).value)
        if c:
            last = r
    ptr = last + 1 if last >= start_row else start_row

    for cc in todo:
        # 双保险：防止同批次重复/或表内已有重复
        if cc in row_map:
            continue

        ws_bare.insert_rows(ptr)
        ws_bare.cell(ptr, code_col, value=cc)
        if name_map:
            ws_bare.cell(ptr, name_col, value=(name_map.get(cc) or ""))
        else:
            ws_bare.cell(ptr, name_col, value="")

        # 留空，人工补
        ws_bare.cell(ptr, vat_col, value=None)
        ws_bare.cell(ptr, cons_col, value=None)

        row_map[cc] = ptr
        ptr += 1
