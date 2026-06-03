# fill_1_1.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ==========================================================
# ✅ 基础工具
# ==========================================================
def to_float(v: Any) -> float:
    """
    强健数字解析：
    - 支持 "1,234.56"
    - 支持 会计括号 "(1,234.56)" 视为负数
    - 支持 " 1 234 "（中间空格）
    - 支持 "-" / "" / None
    - 支持 str 公式 "=SUM(...)" -> 0（避免报错）
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if not s or s in {"-", "—"}:
        return 0.0
    if s.startswith("="):
        return 0.0

    s = s.replace(" ", "").replace("\u3000", "").replace("\t", "")

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]

    s = s.replace(",", "")
    try:
        x = float(s)
        return -x if neg else x
    except Exception:
        return 0.0


def norm_proj_name(x: Any) -> str:
    """强力去空格 + 统一括号 + 统一破折号"""
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\r", "").replace("\n", "").replace("\t", "")
    s = s.replace("\u3000", "").replace(" ", "").strip()
    s = s.replace("（", "(").replace("）", ")")
    s = (
        s.replace("－", "-")
         .replace("—", "-")
         .replace("–", "-")
         .replace("−", "-")
    )
    return s


def norm_elem_no(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        if "e" in s.lower():
            s = str(int(float(s)))
    except Exception:
        pass
    return s


def yuan_to_wanyuan(v_yuan: float) -> float:
    return v_yuan / 10000.0


def wanyuan_to_yuan(v_wanyuan: float) -> float:
    return v_wanyuan * 10000.0


def safe_div(a: float, b: float) -> Optional[float]:
    if not b:
        return None
    return a / b


# ==========================================================
# ✅ 口径说明（按你最新要求）
# 1-1 sheet 写入列：D~I
# D=本月数量(吨)    E=本月单价(元/吨)    F=本月金额(万元)
# G=累计数量(吨)    H=累计单价(元/吨)    I=累计金额(万元)
#
# “万元”sheet（生产经营文件）：
# - 上月累计数量：吨（J列）
# - 上月累计金额：万元（L列）  [你最新要求：金额是万元]
#
# 成本表（2-1/2-2/3-1/3-2 等）：
# - 数量：万吨
# - 金额：万元
#   -> 写入 1-1 时，数量要从 万吨 转为 吨（*10000）
# ==========================================================
@dataclass
class RowRef:
    row: int
    anchor_raw: str


@dataclass
class SixVals:
    mq: Optional[float] = None  # 本月数量(吨)
    mp: Optional[float] = None  # 本月单价(元/吨)
    ma: Optional[float] = None  # 本月金额(万元)
    yq: Optional[float] = None  # 累计数量(吨)
    yp: Optional[float] = None  # 累计单价(元/吨)
    ya: Optional[float] = None  # 累计金额(万元)


COL_M_QTY = 4    # D
COL_M_PRICE = 5  # E
COL_M_AMT = 6    # F
COL_Y_QTY = 7    # G
COL_Y_PRICE = 8  # H
COL_Y_AMT = 9    # I


def qty_wt_to_ton(q_wt: float) -> float:
    """万吨 -> 吨"""
    return q_wt * 10000.0


def unit_price_yuan_per_ton(amount_wanyuan: float, qty_ton: float) -> Optional[float]:
    """万元 + 吨 -> 元/吨"""
    if not qty_ton:
        return None
    return safe_div(wanyuan_to_yuan(amount_wanyuan), qty_ton)


def build_index_sheet_1_1(ws: Worksheet, start_row: int = 6, anchor_col: int = 1) -> Dict[str, RowRef]:
    """锚点在A列，用锚点建索引（允许A列中间有空）"""
    idx: Dict[str, RowRef] = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, anchor_col).value
        if v is None:
            continue
        key = norm_proj_name(v)
        if not key:
            continue
        if key not in idx:
            idx[key] = RowRef(row=r, anchor_raw=str(v))
    return idx


def write_sixvals(ws: Worksheet, idx: Dict[str, RowRef], anchor_name: str, vals: SixVals, overwrite_none: bool = False):
    key = norm_proj_name(anchor_name)
    if key not in idx:
        return

    r = idx[key].row

    def set_cell(col: int, v: Optional[float]):
        if v is None and not overwrite_none:
            return
        ws.cell(r, col).value = v

    set_cell(COL_M_QTY, vals.mq)
    set_cell(COL_M_PRICE, vals.mp)
    set_cell(COL_M_AMT, vals.ma)
    set_cell(COL_Y_QTY, vals.yq)
    set_cell(COL_Y_PRICE, vals.yp)
    set_cell(COL_Y_AMT, vals.ya)


def read_sixvals(ws: Worksheet, idx: Dict[str, RowRef], anchor_name: str) -> SixVals:
    key = norm_proj_name(anchor_name)
    if key not in idx:
        return SixVals(0, 0, 0, 0, 0, 0)
    r = idx[key].row
    return SixVals(
        mq=to_float(ws.cell(r, COL_M_QTY).value),
        mp=to_float(ws.cell(r, COL_M_PRICE).value),
        ma=to_float(ws.cell(r, COL_M_AMT).value),
        yq=to_float(ws.cell(r, COL_Y_QTY).value),
        yp=to_float(ws.cell(r, COL_Y_PRICE).value),
        ya=to_float(ws.cell(r, COL_Y_AMT).value),
    )


def get_project_name_by_anchor(ws: Worksheet, idx: Dict[str, RowRef], anchor_name: str, project_col: int = 2) -> str:
    """锚点定位行，再取B列项目名"""
    k = norm_proj_name(anchor_name)
    if k not in idx:
        return ""
    r = idx[k].row
    return norm_proj_name(ws.cell(r, project_col).value)


# ==========================================================
# ✅ 通用找行：2-1/2-2/3-1/3-2 锚点都在 A 列
# ==========================================================
def find_row_by_name(ws: Worksheet, name: str, start_row: int = 1, col: int = 1) -> Optional[int]:
    target = norm_proj_name(name)
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        if norm_proj_name(v) == target:
            return r
    return None


# ==========================================================
# ✅ 读取 2-1 / 2-2：取 K-M（本期加工）
# 成本表口径（你要求）：数量=万吨，金额=万元
# ==========================================================
def read_anchor_2x_km(ws: Worksheet, anchor_name: str) -> Tuple[float, float, float]:
    r = find_row_by_name(ws, anchor_name, start_row=1, col=1)
    if r is None:
        return 0.0, 0.0, 0.0
    q_wt = to_float(ws.cell(r, 11).value)  # K 数量(万吨)
    p = to_float(ws.cell(r, 12).value)     # L 单价(口径不强依赖)
    a_wan = to_float(ws.cell(r, 13).value) # M 金额(万元)
    return q_wt, p, a_wan


# ==========================================================
# ✅ 读取 3-1 / 3-2（一般贸易-合计）
# 成本表口径（你要求）：数量=万吨，金额=万元
# ==========================================================
COL3_2 = {"年初库存": 5, "本期生产": 8, "本期销售": 11, "其他减少": 14, "期末库存": 17}
COL3_1 = {"年初库存": 5, "本期生产": 8, "本期销售": 11, "其他减少": 14, "期末库存": 17}


def read_group_3x(ws: Worksheet, anchor_name: str, base_col: int) -> Tuple[float, float, float]:
    r = find_row_by_name(ws, anchor_name, start_row=1, col=1)
    if r is None:
        return 0.0, 0.0, 0.0
    q_wt = to_float(ws.cell(r, base_col).value)          # 数量(万吨)
    p = to_float(ws.cell(r, base_col + 1).value)         # 单价（不强依赖）
    a_wan = to_float(ws.cell(r, base_col + 2).value)     # 金额(万元)
    return q_wt, p, a_wan


# ==========================================================
# ✅ “导” sheet 映射表读取（expense文件）
# ==========================================================
def read_elem_project_map_from_dao(ws_dao: Worksheet,
                                  start_row: int = 3,
                                  col_no: int = 20,
                                  col_proj: int = 22) -> Dict[str, List[str]]:
    mp: Dict[str, List[str]] = {}
    for r in range(start_row, ws_dao.max_row + 1):
        proj = ws_dao.cell(r, col_proj).value
        no = ws_dao.cell(r, col_no).value
        if proj is None or no is None:
            continue
        proj_key = norm_proj_name(proj)
        no_key = norm_elem_no(no)
        if not proj_key or not no_key:
            continue
        mp.setdefault(proj_key, []).append(no_key)
    return mp


# ==========================================================
# ✅ “万元” sheet 上月累计（expense文件）
# 你最新口径要求：
# 第5行开始：
# A=项目名
# J=上月累计数量(吨)
# K=单价(元/吨)  （可不使用）
# L=累计金额(万元)
# ==========================================================
def read_last_month_cum_from_wanyuan(ws_wanyuan: Worksheet,
                                     start_row: int = 5) -> Dict[str, Tuple[float, float, float]]:
    out: Dict[str, Tuple[float, float, float]] = {}
    for r in range(start_row, ws_wanyuan.max_row + 1):
        nm = ws_wanyuan.cell(r, 1).value  # A
        if nm is None:
            continue
        key = norm_proj_name(nm)
        if not key:
            continue
        q_ton = to_float(ws_wanyuan.cell(r, 10).value)  # J 吨
        p_yuan_per_ton = to_float(ws_wanyuan.cell(r, 11).value)  # K 元/吨（可不依赖）
        a_wan = to_float(ws_wanyuan.cell(r, 12).value)  # L 万元
        out[key] = (q_ton, p_yuan_per_ton, a_wan)
    return out


# ==========================================================
# ✅ 手工输入解析（UI给：字符串→float）
# 手工输入在本脚本中按 1-1 口径理解：
# - mq/yq：吨
# - ma/ya：万元
# - mp/yp：元/吨
# ==========================================================
def parse_manual_inputs(manual_dict: Dict[str, Dict[str, Any]]) -> Dict[str, SixVals]:
    out: Dict[str, SixVals] = {}
    for k, d in (manual_dict or {}).items():
        name = norm_proj_name(k)
        if not name:
            continue
        out[name] = SixVals(
            mq=None if d.get("mq") in (None, "") else to_float(d.get("mq")),
            mp=None if d.get("mp") in (None, "") else to_float(d.get("mp")),
            ma=None if d.get("ma") in (None, "") else to_float(d.get("ma")),
            yq=None if d.get("yq") in (None, "") else to_float(d.get("yq")),
            yp=None if d.get("yp") in (None, "") else to_float(d.get("yp")),
            ya=None if d.get("ya") in (None, "") else to_float(d.get("ya")),
        )
    return out


# ==========================================================
# ✅ 读取 sheet6：B列项目名匹配，取 D/E（单位：万元）
# ==========================================================
def build_sheet6_map(ws6: Worksheet, start_row: int = 4, name_col: int = 2) -> Dict[str, Tuple[float, float]]:
    mp = {}
    for r in range(start_row, ws6.max_row + 1):
        nm = ws6.cell(r, name_col).value  # B
        if not nm:
            continue
        key = norm_proj_name(nm)
        if not key:
            continue
        m = to_float(ws6.cell(r, 4).value)  # D 万元
        y = to_float(ws6.cell(r, 5).value)  # E 万元
        mp[key] = (m, y)
    return mp


# ==========================================================
# ✅ 主函数：填充 1-1（完整版 + 财务费用统一口径）
# ==========================================================
def fill_cost_sheet_1_1(
    wb_cost=None,
    expense_wb=None,

    cost_elem_month_map: Optional[Dict[str, float]] = None,  # 元
    cost_elem_ytd_map: Optional[Dict[str, float]] = None,    # 元

    ws_11: Worksheet = None,
    ws_21: Worksheet = None,
    ws_22: Worksheet = None,
    ws_31: Worksheet = None,
    ws_32: Worksheet = None,
    ws_6: Worksheet = None,
    ws_7: Worksheet = None,

    expense_file: str = None,
    manual_1_1_table: Optional[Dict[str, Dict[str, Any]]] = None,

    dao_sheet_name: str = "导",
    sheet_wanyuan_name: str = "万元",
    **_ignored,
):
    # runner 兼容：fill_cost_sheet_1_1(ws_11, ws_21=...)
    if ws_11 is None and isinstance(wb_cost, Worksheet):
        ws_11 = wb_cost

    if ws_11 is None:
        raise ValueError("fill_cost_sheet_1_1: 缺少 ws_11")

    ws = ws_11
    idx = build_index_sheet_1_1(ws, start_row=6, anchor_col=1)

    # ==========================================================
    # ✅补充：1-1 半成品锚点取数（来自成本报表 sheet=7）
    # - 成本费用-加：期初半成品
    #     DEF <- 7表 GHI（找 C列=合计 行）
    #     GHI <- 7表 MNO
    # - 成本费用-减：期末半成品
    #     DEF <- 7表 DEF
    #     GHI <- 7表 DEF（同口径）
    # ==========================================================
    if ws_7 is not None:
        # 找到 7表 C列文字为“合计”的行；若找不到，回退到第5行（通常是合计行）
        total_row = None
        for rr in range(1, ws_7.max_row + 1):
            v = ws_7.cell(rr, 3).value  # C
            if v is None:
                continue
            if str(v).strip() == "合计":
                total_row = rr
                break
        if total_row is None and ws_7.max_row >= 5:
            total_row = 5

        if total_row is not None:
            def _v(col: int) -> float:
                return to_float(ws_7.cell(total_row, col).value)

            # 期初半成品：本月 DEF <- GHI；累计 GHI <- MNO
            # 期初半成品：本月 DEF <- GHI；累计 GHI <- MNO
            if "成本费用-加：期初半成品" in idx:
                write_sixvals(
                    ws, idx, "成本费用-加：期初半成品",
                    SixVals(
                        mq=qty_wt_to_ton(_v(7)), mp=_v(8), ma=_v(9),
                        yq=qty_wt_to_ton(_v(13)), yp=_v(14), ya=_v(15),
                    ),
                )

            # 期末半成品：本月 DEF <- DEF；累计 GHI <- DEF
            if "成本费用-减：期末半成品" in idx:
                write_sixvals(
                    ws, idx, "成本费用-减：期末半成品",
                    SixVals(
                        mq=qty_wt_to_ton(_v(4)), mp=_v(5), ma=_v(6),
                        yq=qty_wt_to_ton(_v(4)), yp=_v(5), ya=_v(6),
                    ),
                )

    # ✅ 解析手工输入：供中间逻辑使用；最终覆盖写入在函数末尾执行（空值不覆盖）
    manual_inputs = parse_manual_inputs(manual_1_1_table or {})
    # （注意：此处不直接写入，避免后续自动取数覆盖）


    # 成本要素 map（元）
    cost_elem_month_map = cost_elem_month_map or {}
    cost_elem_ytd_map = cost_elem_ytd_map or {}

    # expense_wb
    if expense_wb is None:
        if not expense_file:
            raise ValueError("fill_cost_sheet_1_1: 缺少 expense_wb/expense_file")
        expense_wb = load_workbook(expense_file, data_only=True)
    if isinstance(expense_wb, Worksheet):
        expense_wb = expense_wb.parent

    if dao_sheet_name not in expense_wb.sheetnames:
        raise ValueError(f"生产经营文件缺少sheet：{dao_sheet_name}")
    if sheet_wanyuan_name not in expense_wb.sheetnames:
        raise ValueError(f"生产经营文件缺少sheet：{sheet_wanyuan_name}")

    ws_dao = expense_wb[dao_sheet_name]
    ws_wanyuan = expense_wb[sheet_wanyuan_name]

    proj2elems = read_elem_project_map_from_dao(ws_dao)
    last_cum = read_last_month_cum_from_wanyuan(ws_wanyuan)  # 数量吨，金额万元

    # sheet6 map（如果没勾6，可能为空，不报错）
    sheet6_map: Dict[str, Tuple[float, float]] = {}
    if ws_6 is not None:
        sheet6_map = build_sheet6_map(ws_6)
    else:
        try:
            if ws.parent and "6" in ws.parent.sheetnames:
                sheet6_map = build_sheet6_map(ws.parent["6"])
        except Exception:
            sheet6_map = {}

    # ==========================================================
    # 1) 原料三项：来自2-1/2-2 K-M（本期加工）
    # 成本表：数量=万吨 金额=万元 -> 写入1-1：数量=吨 金额=万元
    # 单价按 元/吨 重新计算
    # ==========================================================
    raw_map = {
        "原料-原油": ("原油(一般贸易)", "原油(一般贸易)"),
        "原料-外购其他原料": ("外购原料油-合计", "外购原料油-合计"),
        "原料-来料加工原油": ("来料加工-合计", "来料加工-合计"),
    }

    def set_from_2x(anchor_1_1: str, a21: str, a22: str):
        q_m_wt, _p_m, a_m_wan = read_anchor_2x_km(ws_21, a21)
        q_y_wt, _p_y, a_y_wan = read_anchor_2x_km(ws_22, a22)

        mq_ton = qty_wt_to_ton(q_m_wt)
        yq_ton = qty_wt_to_ton(q_y_wt)

        mp = unit_price_yuan_per_ton(a_m_wan, mq_ton)
        yp = unit_price_yuan_per_ton(a_y_wan, yq_ton)

        write_sixvals(ws, idx, anchor_1_1, SixVals(
            mq=mq_ton, mp=mp, ma=a_m_wan,
            yq=yq_ton, yp=yp, ya=a_y_wan,
        ))

    for anchor_1_1, (a21, a22) in raw_map.items():
        set_from_2x(anchor_1_1, a21, a22)

    # ==========================================================
    # 2) 原料合计行（锚点=原料）：三项逐列求和
    # 1-1：数量吨，金额万元，单价元/吨
    # ==========================================================
    v1 = read_sixvals(ws, idx, "原料-原油")
    v2 = read_sixvals(ws, idx, "原料-外购其他原料")
    v3 = read_sixvals(ws, idx, "原料-来料加工原油")

    raw_total = SixVals(
        mq=(v1.mq or 0) + (v2.mq or 0) + (v3.mq or 0),
        ma=(v1.ma or 0) + (v2.ma or 0) + (v3.ma or 0),
        yq=(v1.yq or 0) + (v2.yq or 0) + (v3.yq or 0),
        ya=(v1.ya or 0) + (v2.ya or 0) + (v3.ya or 0),
    )
    raw_total.mp = unit_price_yuan_per_ton(raw_total.ma or 0.0, raw_total.mq or 0.0)
    raw_total.yp = unit_price_yuan_per_ton(raw_total.ya or 0.0, raw_total.yq or 0.0)
    write_sixvals(ws, idx, "原料", raw_total)

    raw_m_qty_ton = raw_total.mq or 0.0
    raw_y_qty_ton = raw_total.yq or 0.0

    # ==========================================================
    # 3) 成本要素映射求和（元->万元）
    # ==========================================================
    def sum_cost_by_anchor(anchor_name: str) -> Tuple[float, float]:
        proj_name = get_project_name_by_anchor(ws, idx, anchor_name, project_col=2)
        if not proj_name:
            return 0.0, 0.0
        proj_key = norm_proj_name(proj_name)
        elems = proj2elems.get(proj_key, [])

        m_yuan = 0.0
        y_yuan = 0.0
        for e in elems:
            m_yuan += cost_elem_month_map.get(e, 0.0)
            y_yuan += cost_elem_ytd_map.get(e, 0.0)
        return yuan_to_wanyuan(m_yuan), yuan_to_wanyuan(y_yuan)

    def set_amt_only(anchor_name: str):
        m_wan, y_wan = sum_cost_by_anchor(anchor_name)
        write_sixvals(ws, idx, anchor_name, SixVals(ma=m_wan, ya=y_wan))

    amt_only_anchors = [
        "变动费用-外购辅助材料",
        "变动费用-外购动力-其它",
        "不含折旧、财务费用的固定费用-修理费",
        "不含折旧、财务费用的固定费用-职工薪酬",
        "不含折旧、财务费用的固定费用-其他管理销售费用",
        "折旧费及摊销-折旧费",
        "折旧费及摊销-无形资产摊销",
        "折旧费及摊销-长期待摊费用摊销",
    ]
    for a in amt_only_anchors:
        set_amt_only(a)

    # ==========================================================
    # 4) 动力类：数量手工（吨），本月/累计金额=成本要素（万元），上月累计数量=万元sheet（吨）
    # 单价 = (万元*10000)/吨
    # ==========================================================
    power_items = [
        "变动费用-外购动力-新鲜水（吨）",
        "变动费用-外购动力-电（千瓦时、元/千瓦时）",
        "变动费用-外购动力-蒸汽（吨)",
        "变动费用-外购动力-氮气（标立)",
        "变动费用-外购燃料",
    ]

    def fill_power_item(anchor_name: str):
        mk = norm_proj_name(anchor_name)

        # 手工输入：吨
        mq_ton = 0.0
        if mk in manual_inputs and manual_inputs[mk].mq is not None:
            mq_ton = float(manual_inputs[mk].mq)

        # 成本要素：万元
        m_amt_wan, y_amt_wan = sum_cost_by_anchor(anchor_name)

        # 上月累计数量：吨（万元sheet）
        proj_name = get_project_name_by_anchor(ws, idx, anchor_name, project_col=2)
        last_q_ton = last_cum.get(norm_proj_name(proj_name), (0.0, 0.0, 0.0))[0]

        yq_ton = last_q_ton + mq_ton

        mp = unit_price_yuan_per_ton(m_amt_wan, mq_ton)
        yp = unit_price_yuan_per_ton(y_amt_wan, yq_ton)

        write_sixvals(ws, idx, anchor_name, SixVals(
            mq=mq_ton, mp=mp, ma=m_amt_wan,
            yq=yq_ton, yp=yp, ya=y_amt_wan
        ))

    for it in power_items:
        fill_power_item(it)

    # ==========================================================
    # 5) 汇总项 + 现金操作成本 + 完全费用
    # ==========================================================
    def get_amt(anchor: str, col: int) -> float:
        k = norm_proj_name(anchor)
        if k not in idx:
            return 0.0
        return to_float(ws.cell(idx[k].row, col).value)

    # 外购动力合计（金额万元）
    power_parts = [
        "变动费用-外购动力-新鲜水（吨）",
        "变动费用-外购动力-电（千瓦时、元/千瓦时）",
        "变动费用-外购动力-蒸汽（吨)",
        "变动费用-外购动力-氮气（标立)",
        "变动费用-外购动力-其它",
    ]
    power_m = sum(get_amt(x, COL_M_AMT) for x in power_parts)
    power_y = sum(get_amt(x, COL_Y_AMT) for x in power_parts)
    write_sixvals(ws, idx, "变动费用-外购动力", SixVals(ma=power_m, ya=power_y))

    # 变动费用小计（金额万元）
    var_m = (
        get_amt("变动费用-外购辅助材料", COL_M_AMT)
        + get_amt("变动费用-外购动力", COL_M_AMT)
        + get_amt("变动费用-外购燃料", COL_M_AMT)
    )
    var_y = (
        get_amt("变动费用-外购辅助材料", COL_Y_AMT)
        + get_amt("变动费用-外购动力", COL_Y_AMT)
        + get_amt("变动费用-外购燃料", COL_Y_AMT)
    )
    write_sixvals(ws, idx, "变动费用小计", SixVals(ma=var_m, ya=var_y))

    # 固定费用小计（金额万元）
    fix_parts = [
        "不含折旧、财务费用的固定费用-修理费",
        "不含折旧、财务费用的固定费用-职工薪酬",
        "不含折旧、财务费用的固定费用-其他管理销售费用",
    ]
    fix_m = sum(get_amt(x, COL_M_AMT) for x in fix_parts)
    fix_y = sum(get_amt(x, COL_Y_AMT) for x in fix_parts)
    write_sixvals(ws, idx, "不含折旧、财务费用的固定费用小计", SixVals(ma=fix_m, ya=fix_y))

    # 折旧费及摊销（金额万元）
    dep_parts = [
        "折旧费及摊销-折旧费",
        "折旧费及摊销-无形资产摊销",
        "折旧费及摊销-长期待摊费用摊销",
    ]
    dep_m = sum(get_amt(x, COL_M_AMT) for x in dep_parts)
    dep_y = sum(get_amt(x, COL_Y_AMT) for x in dep_parts)
    write_sixvals(ws, idx, "折旧费及摊销", SixVals(ma=dep_m, ya=dep_y))

    # ==========================================================
    # 6) 管理/财务/销售费用：来自 sheet6（万元），单价换算为 元/吨（用原料总量吨做分母）
    # ==========================================================
    def fill_fee_anchor(anchor_name: str, special_remove_prefix: bool = False, sheet6_key_override: str | None = None):
        proj = get_project_name_by_anchor(ws, idx, anchor_name, project_col=2)
        if not proj:
            return
        key = norm_proj_name(proj)
        if special_remove_prefix:
            key = key.replace("减:", "").replace("减：", "")

        # ✅ 可选：强制映射到 sheet6 的项目名（用于 1-1 模板B列写法与sheet6不一致）
        if sheet6_key_override:
            key = norm_proj_name(sheet6_key_override)

        m_amt_wan, y_amt_wan = sheet6_map.get(key, (0.0, 0.0))

        mp = unit_price_yuan_per_ton(m_amt_wan, raw_m_qty_ton) if raw_m_qty_ton else None
        yp = unit_price_yuan_per_ton(y_amt_wan, raw_y_qty_ton) if raw_y_qty_ton else None

        write_sixvals(ws, idx, anchor_name, SixVals(mp=mp, ma=m_amt_wan, yp=yp, ya=y_amt_wan))

    fill_fee_anchor("成本费用-减：管理费用", special_remove_prefix=True)
    fill_fee_anchor("成本费用-财务费用", special_remove_prefix=False)
    fill_fee_anchor("成本费用-销售费用", special_remove_prefix=False)

    # ✅ 研发费用：同销售费用口径，来源 sheet6 的“研发费用”
    fill_fee_anchor("成本费用-研发费用", special_remove_prefix=False, sheet6_key_override="研发费用")

    # ==========================================================
    # ✅ 财务费用统一口径：财务费用锚点 = 复制 “成本费用-财务费用” 锚点（只复制价/额）
    # 这样后续 fin_m/fin_y 读取时一定是统一口径
    # ==========================================================
    src_key = norm_proj_name("成本费用-财务费用")
    dst_key = norm_proj_name("财务费用")
    if src_key in idx and dst_key in idx:
        src = read_sixvals(ws, idx, "成本费用-财务费用")
        write_sixvals(
            ws, idx, "财务费用",
            SixVals(mp=src.mp, ma=src.ma, yp=src.yp, ya=src.ya),
            overwrite_none=False
        )

    fin_m = get_amt("财务费用", COL_M_AMT)  # 万元
    fin_y = get_amt("财务费用", COL_Y_AMT)  # 万元

    # ✅ 现金操作成本 = 变动费用小计 + 固定费用小计 - 外供劳务及动力（减项）
    outsvc_m = get_amt("外供劳务及动力（减项）", COL_M_AMT)
    outsvc_y = get_amt("外供劳务及动力（减项）", COL_Y_AMT)

    cash_m = var_m + fix_m - outsvc_m  # 万元
    cash_y = var_y + fix_y - outsvc_y  # 万元

    cash_mp = unit_price_yuan_per_ton(cash_m, raw_m_qty_ton) if raw_m_qty_ton else None
    cash_yp = unit_price_yuan_per_ton(cash_y, raw_y_qty_ton) if raw_y_qty_ton else None
    write_sixvals(ws, idx, "现金操作成本", SixVals(mp=cash_mp, ma=cash_m, yp=cash_yp, ya=cash_y))

    # ✅ 完全费用 = 现金操作成本 + 折旧费及摊销 + 财务费用
    full_m = cash_m + dep_m + fin_m  # 万元
    full_y = cash_y + dep_y + fin_y  # 万元

    full_mp = unit_price_yuan_per_ton(full_m, raw_m_qty_ton) if raw_m_qty_ton else None
    full_yp = unit_price_yuan_per_ton(full_y, raw_y_qty_ton) if raw_y_qty_ton else None
    write_sixvals(ws, idx, "完全费用", SixVals(mp=full_mp, ma=full_m, yp=full_yp, ya=full_y))

    # 成本费用合计（剔除芳烃II）= 原料 + 完全费用（金额万元）
    raw_amt_m = raw_total.ma or 0.0
    raw_amt_y = raw_total.ya or 0.0
    write_sixvals(ws, idx, "成本费用合计（剔除芳烃II）", SixVals(ma=raw_amt_m + full_m, ya=raw_amt_y + full_y))
    write_sixvals(ws, idx, "成本费用合计(剔除芳烃II)", SixVals(ma=raw_amt_m + full_m, ya=raw_amt_y + full_y))

    # ==========================================================
    # 7) 商品产品总成本 / 销售成本合计：3-1/3-2 一般贸易-合计
    # 成本表：数量=万吨 金额=万元 -> 写入1-1：数量=吨 金额=万元，单价重新算
    # ==========================================================
    anchor_3 = "一般贸易-合计"

    def write_from_3x(target_anchor: str, seg: str):
        q_m_wt, _p_m, a_m_wan = read_group_3x(ws_31, anchor_3, base_col=COL3_1[seg])
        q_y_wt, _p_y, a_y_wan = read_group_3x(ws_32, anchor_3, base_col=COL3_2[seg])

        mq_ton = qty_wt_to_ton(q_m_wt)
        yq_ton = qty_wt_to_ton(q_y_wt)

        mp = unit_price_yuan_per_ton(a_m_wan, mq_ton)
        yp = unit_price_yuan_per_ton(a_y_wan, yq_ton)

        write_sixvals(ws, idx, target_anchor, SixVals(
            mq=mq_ton, mp=mp, ma=a_m_wan,
            yq=yq_ton, yp=yp, ya=a_y_wan,
        ))

    write_from_3x("商品产品总成本", "本期生产")
    write_from_3x("商品产品总成本-加：期初产成品", "年初库存")
    write_from_3x("商品产品总成本-减：期末产成品", "期末库存")
    write_from_3x("商品产品总成本-产成品其他减少", "其他减少")
    write_from_3x("销售成本合计", "本期销售")

    # ✅ 手工输入最终覆盖写入（空值不覆盖）：确保界面手工值优先于自动取数
    for anchor, sv in manual_inputs.items():
        write_sixvals(ws, idx, anchor, sv, overwrite_none=False)

    # ✅ 手工调整可能影响汇总指标（如现金操作成本/完全费用/成本费用合计等），此处统一重算一次
    def _recalc_after_manual():
        # 1) 外购动力合计（金额万元）
        _power_parts = [
            "变动费用-外购动力-新鲜水（吨）",
            "变动费用-外购动力-电（千瓦时、元/千瓦时）",
            "变动费用-外购动力-蒸汽（吨)",
            "变动费用-外购动力-氮气（标立)",
            "变动费用-外购动力-其它",
        ]
        _power_m = sum(get_amt(x, COL_M_AMT) for x in _power_parts)
        _power_y = sum(get_amt(x, COL_Y_AMT) for x in _power_parts)
        write_sixvals(ws, idx, "变动费用-外购动力", SixVals(ma=_power_m, ya=_power_y))

        # 2) 变动费用小计（金额万元）
        _var_m = (
            get_amt("变动费用-外购辅助材料", COL_M_AMT)
            + get_amt("变动费用-外购动力", COL_M_AMT)
            + get_amt("变动费用-外购燃料", COL_M_AMT)
        )
        _var_y = (
            get_amt("变动费用-外购辅助材料", COL_Y_AMT)
            + get_amt("变动费用-外购动力", COL_Y_AMT)
            + get_amt("变动费用-外购燃料", COL_Y_AMT)
        )
        write_sixvals(ws, idx, "变动费用小计", SixVals(ma=_var_m, ya=_var_y))

        # 3) 固定费用小计（金额万元）
        _fix_parts = [
            "不含折旧、财务费用的固定费用-修理费",
            "不含折旧、财务费用的固定费用-职工薪酬",
            "不含折旧、财务费用的固定费用-其他管理销售费用",
        ]
        _fix_m = sum(get_amt(x, COL_M_AMT) for x in _fix_parts)
        _fix_y = sum(get_amt(x, COL_Y_AMT) for x in _fix_parts)
        write_sixvals(ws, idx, "不含折旧、财务费用的固定费用小计", SixVals(ma=_fix_m, ya=_fix_y))

        # 4) 折旧费及摊销（金额万元）
        _dep_parts = [
            "折旧费及摊销-折旧费",
            "折旧费及摊销-无形资产摊销",
            "折旧费及摊销-长期待摊费用摊销",
        ]
        _dep_m = sum(get_amt(x, COL_M_AMT) for x in _dep_parts)
        _dep_y = sum(get_amt(x, COL_Y_AMT) for x in _dep_parts)
        write_sixvals(ws, idx, "折旧费及摊销", SixVals(ma=_dep_m, ya=_dep_y))

        # 5) 财务费用统一口径：财务费用锚点 = 复制 “成本费用-财务费用” 锚点（只复制价/额）
        _src_key = norm_proj_name("成本费用-财务费用")
        _dst_key = norm_proj_name("财务费用")
        if _src_key in idx and _dst_key in idx:
            _src = read_sixvals(ws, idx, "成本费用-财务费用")
            write_sixvals(ws, idx, "财务费用", SixVals(mp=_src.mp, ma=_src.ma, yp=_src.yp, ya=_src.ya), overwrite_none=False)

        _fin_m = get_amt("财务费用", COL_M_AMT)
        _fin_y = get_amt("财务费用", COL_Y_AMT)

        # 6) 现金操作成本 / 完全费用
        _outsvc_m = get_amt("外供劳务及动力（减项）", COL_M_AMT)
        _outsvc_y = get_amt("外供劳务及动力（减项）", COL_Y_AMT)
        _cash_m = _var_m + _fix_m - _outsvc_m
        _cash_y = _var_y + _fix_y - _outsvc_y
        _cash_mp = unit_price_yuan_per_ton(_cash_m, raw_m_qty_ton) if raw_m_qty_ton else None
        _cash_yp = unit_price_yuan_per_ton(_cash_y, raw_y_qty_ton) if raw_y_qty_ton else None
        write_sixvals(ws, idx, "现金操作成本", SixVals(mp=_cash_mp, ma=_cash_m, yp=_cash_yp, ya=_cash_y))

        _full_m = _cash_m + _dep_m + _fin_m
        _full_y = _cash_y + _dep_y + _fin_y
        _full_mp = unit_price_yuan_per_ton(_full_m, raw_m_qty_ton) if raw_m_qty_ton else None
        _full_yp = unit_price_yuan_per_ton(_full_y, raw_y_qty_ton) if raw_y_qty_ton else None
        write_sixvals(ws, idx, "完全费用", SixVals(mp=_full_mp, ma=_full_m, yp=_full_yp, ya=_full_y))

        # 7) 成本费用合计（剔除芳烃II）= 原料 + 完全费用（金额万元）
        _raw_amt_m = (raw_total.ma or 0.0) if "raw_total" in locals() else 0.0
        _raw_amt_y = (raw_total.ya or 0.0) if "raw_total" in locals() else 0.0
        write_sixvals(ws, idx, "成本费用合计（剔除芳烃II）", SixVals(ma=_raw_amt_m + _full_m, ya=_raw_amt_y + _full_y))
        write_sixvals(ws, idx, "成本费用合计(剔除芳烃II)", SixVals(ma=_raw_amt_m + _full_m, ya=_raw_amt_y + _full_y))

    _recalc_after_manual()



# runner 兼容
def fill_cost_sheet_1_10(*args, **kwargs):
    return fill_cost_sheet_1_1(*args, **kwargs)