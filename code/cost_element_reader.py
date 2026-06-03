# cost_element_reader.py
from __future__ import annotations

from typing import Dict, Any, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _norm_elem_no(x: Any) -> str:
    """成本要素号：去空格/去.0/去科学计数法"""
    if x is None:
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        if "e" in s.lower():
            s = str(int(float(s)))
    except Exception:
        pass
    return s


def _to_float(v: Any) -> float:
    """
    强健数字解析：
    - 支持 '1,234.56'
    - 支持 会计括号 '(1,234.56)' 视为负数
    - 支持 '-' / '' / None
    - 支持全角空格
    """
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if not s or s in {"-", "—"}:
        return 0.0
    if s.startswith("="):  # 公式
        return 0.0

    s = s.replace(" ", "").replace("\u3000", "").replace("\t", "")

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]

    s = s.replace(",", "")
    try:
        x = float(s)
        return -x if neg else x
    except Exception:
        return 0.0


def _split_elem_cell(x: Any) -> Tuple[str, str]:
    """
    SAP 成本要素列格式常见为：
    "4001009900  成本费用-原料及主要材料-手工"
    """
    if x is None:
        return "", ""
    s = str(x).strip()
    if not s:
        return "", ""

    parts = s.split()
    if not parts:
        return "", ""
    elem_no = _norm_elem_no(parts[0])
    elem_name = " ".join(parts[1:]).strip() if len(parts) > 1 else ""
    return elem_no, elem_name


def _norm_header(s: Any) -> str:
    """表头归一：去空白、统一括号/横杠/冒号"""
    if s is None:
        return ""
    t = str(s).strip()
    if not t:
        return ""
    t = t.replace("\r", "").replace("\n", "").replace("\t", "")
    t = t.replace("\u3000", "").replace(" ", "")
    t = t.replace("（", "(").replace("）", ")")
    t = t.replace("－", "-").replace("—", "-").replace("–", "-").replace("−", "-")
    t = t.replace("：", ":")
    return t


def _build_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """从 header_row 扫描表头，返回 {归一化表头: 列号(1-based)}"""
    mp: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        key = _norm_header(v)
        if not key:
            continue
        if key not in mp:
            mp[key] = c
    return mp


def _find_col_contains(header_map: Dict[str, int], kw: str) -> Optional[int]:
    for k, c in header_map.items():
        if kw in k:
            return c
    return None


def _guess_header_row(ws: Worksheet, max_scan: int = 40) -> Optional[int]:
    """
    自动找标题行：优先找包含“成本要素”和“实际成本”的那行
    """
    for r in range(1, min(ws.max_row, max_scan) + 1):
        row_cells = []
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            row_cells.append(_norm_header(v))
        joined = "|".join(row_cells)
        if "成本要素" in joined and "实际成本" in joined:
            return r
    return None


def read_cost_element_table(
    file_path: str,
    sheet_name: str | int = 0,
    header_row: Optional[int] = 12,  # ✅你新模板默认12；不确定可传 None 自动识别
) -> Dict[str, float]:
    """
    成本要素表读取成 map：
    {成本要素号: 实际成本(元)}

    新模板（你现在的）：
    - 第12行：标题
    - A列：成本要素
    - B列：实际成本

    本函数会：
    - 扫描 header_row 表头自动定位“成本要素/实际成本”列
    - 聚合同一成本要素号的实际成本
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[int(sheet_name)]

    if header_row is None:
        header_row = _guess_header_row(ws) or 12

    header_map = _build_header_map(ws, header_row)

    # 你给的表头就是：成本要素 / 实际成本
    col_elem = header_map.get("成本要素") or _find_col_contains(header_map, "成本要素") or 1
    col_amt = header_map.get("实际成本") or _find_col_contains(header_map, "实际成本") or 2

    start_row = header_row + 1

    mp: Dict[str, float] = {}
    for r in range(start_row, ws.max_row + 1):
        elem_cell = ws.cell(r, col_elem).value
        amt_cell = ws.cell(r, col_amt).value

        elem_no, _ = _split_elem_cell(elem_cell)
        if not elem_no:
            continue

        amt = _to_float(amt_cell)
        mp[elem_no] = mp.get(elem_no, 0.0) + amt

    return mp
