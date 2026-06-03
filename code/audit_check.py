# audit_check.py
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import IconSetRule


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except Exception:
        return None


def _round7(v):
    """用于等值判断：保留7位小数"""
    if v is None:
        return None
    try:
        return round(float(v), 7)
    except Exception:
        return None


def normalize_material_name(name: str) -> str:
    """
    成本表物料名可能带后缀 '\\中国'，审查匹配时去掉：
    '蓬莱原油\\中国' -> '蓬莱原油'
    """
    if name is None:
        return ""
    s = str(name).strip()
    if s.endswith("\\中国"):
        s = s[:-3]
    return s.strip().lower()


# =========================================================
# 原油平衡表读取（单位 吨 -> 万吨）
# =========================================================
def read_balance_table(balance_file: str, sheet_name: str = "原燃料-合并石化"):
    wb = load_workbook(balance_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"平衡表缺少sheet：{sheet_name}")
    ws = wb[sheet_name]

    EXCLUDE_TITLE_NAMES = {
        "进口原油",
        "一般贸易",
        "来料加工",
        "二、其他外购原料：",
        "二、其他外购原料:",
        "其他外购原料：",
        "其他外购原料:",
    }

    def to_wan_or_none(v):
        f = _to_float(v)
        if f is None:
            return None
        return f / 10000.0  # 吨 -> 万吨

    m = {}
    for r in range(9, ws.max_row + 1):
        name = ws.cell(r, 2).value  # B
        if name is None or str(name).strip() == "":
            continue

        name_raw = str(name).strip()
        if name_raw in EXCLUDE_TITLE_NAMES:
            continue

        begin0 = to_wan_or_none(ws.cell(r, 4).value)       # D
        purchase_m0 = to_wan_or_none(ws.cell(r, 5).value)  # E
        purchase_y0 = to_wan_or_none(ws.cell(r, 6).value)  # F
        process_m0 = to_wan_or_none(ws.cell(r, 7).value)   # G
        process_y0 = to_wan_or_none(ws.cell(r, 8).value)   # H
        end0 = to_wan_or_none(ws.cell(r, 9).value)         # I

        if (name_raw.endswith("：") or name_raw.endswith(":")) and all(
            v is None for v in [begin0, purchase_m0, purchase_y0, process_m0, process_y0, end0]
        ):
            continue

        key = normalize_material_name(name_raw)

        vals = [begin0, purchase_m0, purchase_y0, process_m0, process_y0, end0]

        def is_none_or_zero(x):
            rx = _round7(x)
            return (x is None) or (rx == 0.0)

        all_empty_or_zero = all(is_none_or_zero(x) for x in vals)

        skip_exist_2_1 = (
            is_none_or_zero(begin0) and
            is_none_or_zero(purchase_m0) and
            is_none_or_zero(process_m0) and
            is_none_or_zero(end0)
        )

        skip_exist_2_2 = (
            is_none_or_zero(purchase_y0) and
            is_none_or_zero(process_y0)
        )

        m[key] = {
            "begin": begin0,
            "purchase_m": purchase_m0,
            "purchase_y": purchase_y0,
            "process_m": process_m0,
            "process_y": process_y0,
            "end": end0,

            "balance_name_raw": name_raw,
            "all_empty_or_zero": all_empty_or_zero,
            "skip_exist_2_1": skip_exist_2_1,
            "skip_exist_2_2": skip_exist_2_2,
        }

    return m, {"sheet_name": sheet_name}


def audit_sheet(ws_cost, sheet_name: str, balance_map: dict, balance_sheet_name: str, tol: float = 0.0):
    qty_cols = {"begin": 5, "purchase": 8, "process": 11, "end": 17}  # 成本表：E/H/K/Q
    target_cats = {"进口原油", "海洋原油", "外购原料油"}

    issues = []
    seq = 1

    cost_map = {}
    for r in range(8, ws_cost.max_row + 1):
        cat = ws_cost.cell(r, 1).value
        name = ws_cost.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue

        cat_s = "" if cat is None else str(cat).strip()
        if cat_s not in target_cats:
            continue

        code = ws_cost.cell(r, 3).value
        code_s = "" if code is None else str(code).strip().lstrip("0")

        key = normalize_material_name(name)

        def get_cost(col):
            return _to_float(ws_cost.cell(r, col).value) or 0.0

        cost_map[key] = {
            "row": r,
            "cat": cat_s,
            "code": code_s,
            "name_raw": str(name).strip(),
            "begin": get_cost(qty_cols["begin"]),
            "purchase": get_cost(qty_cols["purchase"]),
            "process": get_cost(qty_cols["process"]),
            "end": get_cost(qty_cols["end"]),
        }

    def add_issue(cat, code, name_raw, cost_row, col_label, cost_v, bal_v, status, bal_name_raw=""):
        nonlocal seq
        diff = (cost_v - bal_v) if (cost_v is not None and bal_v is not None) else None
        issues.append({
            "seq": seq,
            "sheet": sheet_name,
            "balance_sheet": balance_sheet_name,
            "cost_row": cost_row,
            "category": cat,
            "code": code,
            "name": name_raw,
            "balance_name": bal_name_raw,
            "column": col_label,
            "cost_qty": cost_v,
            "balance_qty": bal_v,
            "diff": diff,
            "status": status,
        })
        seq += 1

    if sheet_name == "2-1":
        compare_plan = [
            ("期初库存数量", "begin", "begin"),
            ("本期采购数量", "purchase", "purchase_m"),
            ("本期加工数量", "process", "process_m"),
            ("期末库存数量", "end", "end"),
        ]
        exist_skip_flag = "skip_exist_2_1"
    else:  # 2-2
        compare_plan = [
            ("年累计采购量", "purchase", "purchase_y"),
            ("年累计加工量", "process", "process_y"),
            ("期末库存数量", "end", "end"),
        ]
        exist_skip_flag = "skip_exist_2_2"

    for key, c in cost_map.items():
        if key not in balance_map:
            continue

        b = balance_map[key]
        if b.get("all_empty_or_zero", False):
            continue

        for col_label, cost_field, bal_field in compare_plan:
            cv = c[cost_field]
            bv = b.get(bal_field, None)
            bv = 0.0 if bv is None else float(bv)

            if _round7(cv) == _round7(bv):
                continue
            if tol and abs(cv - bv) <= tol:
                continue

            add_issue(
                c["cat"], c["code"], c["name_raw"], c["row"],
                col_label, cv, bv, "数值误差",
                bal_name_raw=b.get("balance_name_raw", "")
            )

    for key, b in balance_map.items():
        if key in cost_map:
            continue
        if b.get("all_empty_or_zero", False):
            continue
        if b.get(exist_skip_flag, False):
            continue

        add_issue(
            "", "", "", None,
            "物料存在性", None, None,
            "成本表中不存在，请人工核查",
            bal_name_raw=b.get("balance_name_raw", "")
        )

    for key, c in cost_map.items():
        if key in balance_map:
            if balance_map[key].get("all_empty_or_zero", False):
                continue
            continue

        add_issue(
            c["cat"], c["code"], c["name_raw"], c["row"],
            "物料存在性", None, None,
            "物料表中不存在，请人工核查",
            bal_name_raw=""
        )

    return issues


# =========================================================
# ✅产品平衡表读取（单位：吨 -> 万吨）
# =========================================================
def read_product_balance_table(balance_file: str, sheet_name: str = "产品表-合并石化部"):
    wb = load_workbook(balance_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"平衡表缺少sheet：{sheet_name}")
    ws = wb[sheet_name]

    def to_wan_or_none(v):
        f = _to_float(v)
        if f is None:
            return None
        return f / 10000.0

    m = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 3).value  # C
        if name is None or str(name).strip() == "":
            continue
        name_raw = str(name).strip()
        key = name_raw.strip().lower()

        begin = to_wan_or_none(ws.cell(r, 5).value)     # E
        prod_m = to_wan_or_none(ws.cell(r, 7).value)    # G
        sales_m = to_wan_or_none(ws.cell(r, 10).value)  # J
        end = to_wan_or_none(ws.cell(r, 12).value)      # L

        prod_y = to_wan_or_none(ws.cell(r, 8).value)    # H
        sales_y = to_wan_or_none(ws.cell(r, 11).value)  # K

        vals = [begin, prod_m, sales_m, end, prod_y, sales_y]

        def is_none_or_zero(x):
            rx = _round7(x)
            return (x is None) or (rx == 0.0)

        all_empty_or_zero = all(is_none_or_zero(x) for x in vals)

        m[key] = {
            "begin": begin,
            "prod_m": prod_m,
            "sales_m": sales_m,
            "end": end,
            "prod_y": prod_y,
            "sales_y": sales_y,
            "balance_name_raw": name_raw,
            "all_empty_or_zero": all_empty_or_zero,
        }

    return m, {"sheet_name": sheet_name}


def audit_sheet_product(ws_cost, sheet_name: str, product_balance_map: dict, balance_sheet_name: str, tol: float = 0.0):
    issues = []
    seq = 1

    def add_issue(cat, code, name_raw, cost_row, col_label, cost_v, bal_v, status, bal_name_raw=""):
        nonlocal seq
        diff = (cost_v - bal_v) if (cost_v is not None and bal_v is not None) else None
        issues.append({
            "seq": seq,
            "sheet": sheet_name,
            "balance_sheet": balance_sheet_name,
            "cost_row": cost_row,
            "category": cat,
            "code": code,
            "name": name_raw,
            "balance_name": bal_name_raw,
            "column": col_label,
            "cost_qty": cost_v,
            "balance_qty": bal_v,
            "diff": diff,
            "status": status,
        })
        seq += 1

    # 成本表明细 map：name_norm -> row & vals
    cost_map = {}
    for r in range(6, ws_cost.max_row + 1):
        name = ws_cost.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue
        cat = ws_cost.cell(r, 1).value
        cat_s = "" if cat is None else str(cat).strip()
        code = ws_cost.cell(r, 3).value
        code_s = "" if code is None else str(code).strip().lstrip("0")

        key = str(name).strip().lower()

        def get_qty(col):
            return _to_float(ws_cost.cell(r, col).value) or 0.0

        cost_map[key] = {
            "row": r,
            "cat": cat_s,
            "code": code_s,
            "name_raw": str(name).strip(),
            "begin": get_qty(5),    # E
            "prod": get_qty(8),     # H
            "sales": get_qty(11),   # K
            "end": get_qty(17),     # Q
        }

    if sheet_name == "3-1":
        plan = [
            ("期初库存数量", "begin", "begin"),
            ("本期生产数量", "prod", "prod_m"),
            ("本期销售数量", "sales", "sales_m"),
            ("期末库存数量", "end", "end"),
        ]
    else:  # 3-2
        plan = [
            ("期初库存数量", "begin", "begin"),
            ("年累计生产量", "prod", "prod_y"),
            ("年累计销售量", "sales", "sales_y"),
            ("期末库存数量", "end", "end"),
        ]

    # (1) 双方都有：对比数值
    for k, c in cost_map.items():
        if k not in product_balance_map:
            continue
        b = product_balance_map[k]
        if b.get("all_empty_or_zero", False):
            continue

        for col_label, cost_field, bal_field in plan:
            cv = c[cost_field]
            bv = b.get(bal_field, None)
            bv = 0.0 if bv is None else float(bv)

            if _round7(cv) == _round7(bv):
                continue
            if tol and abs(cv - bv) <= tol:
                continue

            add_issue(
                c["cat"], c["code"], c["name_raw"], c["row"],
                col_label, cv, bv, "数值误差",
                bal_name_raw=b.get("balance_name_raw", "")
            )

    # (2) 平衡表有、成本表没有
    for k, b in product_balance_map.items():
        if k in cost_map:
            continue
        if b.get("all_empty_or_zero", False):
            continue
        add_issue(
            "", "", "", None,
            "物料存在性", None, None,
            "成本表中不存在，请人工核查",
            bal_name_raw=b.get("balance_name_raw", "")
        )

    # (3) 成本表有、平衡表没有
    for k, c in cost_map.items():
        if k in product_balance_map:
            continue
        add_issue(
            c["cat"], c["code"], c["name_raw"], c["row"],
            "物料存在性", None, None,
            "物料表中不存在，请人工核查",
            bal_name_raw=""
        )

    return issues


# =========================================================
# ✅输出审查日志（含：其他减少调整校验 + 图标集✅/❌）
# =========================================================
def save_audit_excel(issues: list, out_path: str, adjust_records: list | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "审查日志"

    headers = [
        "对比记录号",
        "成本表sheet",
        "平衡表sheet",
        "成本表行号",
        "物料分类",
        "物料号",
        "成本表物料名",
        "平衡表物料名",
        "问题列",
        "成本表数量(万/万吨)",
        "平衡表数量(万/万吨)",
        "差值(万/万吨)",
        "状态",
    ]
    ws.append(headers)

    for it in issues:
        ws.append([
            it.get("seq"),
            it.get("sheet"),
            it.get("balance_sheet"),
            it.get("cost_row"),
            it.get("category"),
            it.get("code"),
            it.get("name"),
            it.get("balance_name"),
            it.get("column"),
            it.get("cost_qty"),
            it.get("balance_qty"),
            it.get("diff"),
            it.get("status"),
        ])

    # ========== 样式 ==========
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    FILL_DIFF = PatternFill("solid", fgColor="D9E1F2")         # 浅蓝：数值误差
    FILL_COST_MISSING = PatternFill("solid", fgColor="FFF2CC") # 浅黄：成本表中不存在
    FILL_BAL_MISSING = PatternFill("solid", fgColor="F8CBAD")  # 浅红：物料表中不存在

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        status = str(ws.cell(r, len(headers)).value or "").strip()
        if "误差" in status:
            fill = FILL_DIFF
        elif "成本表中不存在" in status:
            fill = FILL_COST_MISSING
        elif "物料表中不存在" in status:
            fill = FILL_BAL_MISSING
        else:
            fill = None

        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill

    widths = [12, 10, 14, 10, 16, 12, 26, 26, 16, 18, 18, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 数字列格式
    for r in range(2, ws.max_row + 1):
        for c in (10, 11, 12):  # 数量/差值列
            ws.cell(r, c).number_format = "0.0000000"

    ws.freeze_panes = "A2"

    # ======================================================
    # ✅新增sheet：其他减少调整校验（加 IconSet 图标）
    # ======================================================
    if adjust_records:
        ws2 = wb.create_sheet("其他减少调整校验")

        # 说明：
        # - “校验图标”列写 1/0，用 IconSet 显示✅/❌，并隐藏数值
        # - 同时保留“校验文本(✅/❌)”列，方便拷贝/查看
        h2 = [
            "成本表sheet", "行号", "分类", "物料号", "产品名",
            "生产数量(调整前)", "生产金额(调整前)",
            "其他减少数量(调整前)", "其他减少金额(调整前)",
            "生产数量(调整后)", "生产金额(调整后)",
            "其他减少数量(调整后)", "其他减少金额(调整后)",
            "校验图标(1=✅,0=❌)", "校验文本(✅/❌)",
            "lhs(期初+生产(调整前)-销售+其他减少(调整前))", "期末数量",
        ]
        ws2.append(h2)

        def _pick_check_text(x: dict) -> str:
            # 当前主口径：调整前 期初+生产-销售+其他减少 = 期末
            for k in (
                "check(before_adjust: begin+prod_before-sales+other_before=end)",
                "check_before_adjust",
                "check(begin+prod_before-sales+other_before=end)",
                # 兼容历史字段
                "check(begin+prod-sales=end)",
                "check_begin_plus_prod_minus_sales_eq_end",
                "check(begin+prod-sales)=end",
                "check",
            ):
                v = x.get(k)
                if v is not None and str(v).strip() != "":
                    return str(v).strip()
            return ""

        for x in adjust_records:
            # 真实校验口径：比较“调整其他减少到生产之前”的数据
            # lhs = 期初 + 生产(调整前) - 销售 + 其他减少(调整前)

            def _pick_num(d: dict, keys):
                for k in keys:
                    v = _to_float(d.get(k))
                    if v is not None:
                        return v
                return None

            begin_q = _pick_num(x, ("begin_qty", "begin_q", "qty_begin"))
            prod_q_before = _pick_num(x, ("prod_qty_before", "prod_before_qty", "prod_qty_before_calc"))
            sales_q = _pick_num(x, ("sales_qty", "sales_q", "qty_sales"))
            other_q_before = _pick_num(x, ("other_qty_before", "other_before_qty", "other_qty_raw"))
            end_q = _pick_num(x, ("end_qty_calc", "end_qty", "end_q", "qty_end"))

            lhs_q = None
            if None not in (begin_q, prod_q_before, sales_q, other_q_before):
                lhs_q = begin_q + prod_q_before - sales_q + other_q_before
            else:
                lhs_q = _pick_num(x, (
                    "lhs(before_adjust: begin+prod_before-sales+other_before)",
                    "lhs_before_adjust",
                    "lhs(begin+prod_before-sales+other_before)",
                    # 兼容历史字段
                    "lhs(begin+prod-sales)", "lhs_begin_plus_prod_minus_sales", "lhs_qty"
                ))

            check_text = _pick_check_text(x)
            ok = (check_text == "✅")
            if check_text not in {"✅", "❌"} and lhs_q is not None and end_q is not None:
                ok = (_round7(lhs_q) == _round7(end_q))
                check_text = "✅" if ok else "❌"

            check_num = 1 if ok else 0

            ws2.append([
                x.get("sheet"), x.get("row"), x.get("category"), x.get("code"), x.get("name"),
                x.get("prod_qty_before"), x.get("prod_amt_before"),
                x.get("other_qty_before"), x.get("other_amt_before"),
                x.get("prod_qty_after"), x.get("prod_amt_after"),
                x.get("other_qty_after", x.get("other_qty_after_sheet")),
                x.get("other_amt_after", x.get("other_amt_after_sheet")),
                check_num, check_text,
                lhs_q,
                end_q,
            ])

        # 头部样式（绿色表头）
        header2_fill = PatternFill("solid", fgColor="006100")
        header2_font = Font(color="FFFFFF", bold=True)
        for c in range(1, len(h2) + 1):
            cell = ws2.cell(1, c)
            cell.fill = header2_fill
            cell.font = header2_font
            cell.alignment = center
            cell.border = border

        for r in range(2, ws2.max_row + 1):
            for c in range(1, len(h2) + 1):
                cell = ws2.cell(r, c)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 数字格式美化
            for c in range(6, 14):  # 数量/金额各列
                ws2.cell(r, c).number_format = "0.0000000"
            ws2.cell(r, 16).number_format = "0.0000000"  # lhs
            ws2.cell(r, 17).number_format = "0.0000000"  # 期末

        ws2.freeze_panes = "A2"

        # 列宽更舒服
        col_widths = [
            10, 8, 10, 12, 26,
            16, 18,
            18, 18,
            16, 18,
            18, 18,
            16, 14,
            20, 12,
        ]
        for i, w in enumerate(col_widths, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # ✅ IconSet：在第14列（校验图标）加图标集
        icon_col = 14
        first_data_row = 2
        last_data_row = ws2.max_row
        if last_data_row >= first_data_row:
            rng = f"{get_column_letter(icon_col)}{first_data_row}:{get_column_letter(icon_col)}{last_data_row}"

            rule = IconSetRule(
                icon_style="3Symbols2",
                type="num",
                values=[0, 1, 2],
                showValue=False,
                percent=False,
                reverse=False
            )
            ws2.conditional_formatting.add(rng, rule)

            for r in range(first_data_row, last_data_row + 1):
                ws2.cell(r, icon_col).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "审查日志"

    headers = [
        "对比记录号",
        "成本表sheet",
        "平衡表sheet",
        "成本表行号",
        "物料分类",
        "物料号",
        "成本表物料名",
        "平衡表物料名",
        "问题列",
        "成本表数量(万/万吨)",
        "平衡表数量(万/万吨)",
        "差值(万/万吨)",
        "状态",
    ]
    ws.append(headers)

    for it in issues:
        ws.append([
            it.get("seq"),
            it.get("sheet"),
            it.get("balance_sheet"),
            it.get("cost_row"),
            it.get("category"),
            it.get("code"),
            it.get("name"),
            it.get("balance_name"),
            it.get("column"),
            it.get("cost_qty"),
            it.get("balance_qty"),
            it.get("diff"),
            it.get("status"),
        ])

    # ========== 样式 ==========
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A0A0A0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    FILL_DIFF = PatternFill("solid", fgColor="D9E1F2")         # 浅蓝：数值误差
    FILL_COST_MISSING = PatternFill("solid", fgColor="FFF2CC") # 浅黄：成本表中不存在
    FILL_BAL_MISSING = PatternFill("solid", fgColor="F8CBAD")  # 浅红：物料表中不存在

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        status = str(ws.cell(r, len(headers)).value or "").strip()
        if "误差" in status:
            fill = FILL_DIFF
        elif "成本表中不存在" in status:
            fill = FILL_COST_MISSING
        elif "物料表中不存在" in status:
            fill = FILL_BAL_MISSING
        else:
            fill = None

        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill

    widths = [12, 10, 14, 10, 16, 12, 26, 26, 16, 18, 18, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 数字列格式
    for r in range(2, ws.max_row + 1):
        for c in (10, 11, 12):  # 数量/差值列
            ws.cell(r, c).number_format = "0.0000000"

    ws.freeze_panes = "A2"

    # ======================================================
    # ✅新增sheet：其他减少调整校验（加 IconSet 图标）
    # ======================================================
    if adjust_records:
        ws2 = wb.create_sheet("其他减少调整校验")

        # 说明：
        # - “校验图标”列写 1/0，用 IconSet 显示✅/❌，并隐藏数值
        # - 同时保留“校验文本(✅/❌)”列，方便拷贝/查看（你不想要可以删掉）
        h2 = [
            "成本表sheet", "行号", "分类", "物料号", "产品名",
            "生产数量(调整前)", "生产金额(调整前)",
            "其他减少数量(调整前)", "其他减少金额(调整前)",
            "生产数量(调整后)", "生产金额(调整后)",
            "其他减少数量(调整后)", "其他减少金额(调整后)",
            "校验图标(1=✅,0=❌)", "校验文本(✅/❌)",
            "lhs(期初+生产-销售)", "期末数量",
        ]
        ws2.append(h2)

        def _pick_check_text(x: dict) -> str:
            # 兼容不同key命名（你product_fill里用的是 check(begin+prod-sales=end)）
            for k in (
                "check(begin+prod-sales=end)",
                "check_begin_plus_prod_minus_sales_eq_end",
                "check(begin+prod-sales)=end",
                "check",
            ):
                v = x.get(k)
                if v is not None and str(v).strip() != "":
                    return str(v).strip()
            return ""

        for x in adjust_records:
            # lhs 校验口径：lhs = 期初数量 +（已将‘其他减少’挪入后的生产数量）- 销量数量
            # 优先用 product_fill 提供的字段重算；字段缺失时再退化使用历史 lhs 字段

            def _pick_num(d: dict, keys):
                for k in keys:
                    v = _to_float(d.get(k))
                    if v is not None:
                        return v
                return None

            begin_q = _pick_num(x, ("begin_qty", "begin_q", "qty_begin"))
            # 生产：必须取‘挪到生产后’的生产数量；优先 prod_qty_after_calc，其次 prod_qty_after
            prod_q_after = _pick_num(x, ("prod_qty_after_calc", "prod_qty_after", "prod_after_qty", "prod_qty"))
            sales_q = _pick_num(x, ("sales_qty", "sales_q", "qty_sales"))
            end_q = _pick_num(x, ("end_qty_calc", "end_qty", "end_q", "qty_end"))

            lhs_q = None
            if begin_q is not None and prod_q_after is not None and sales_q is not None:
                lhs_q = begin_q + prod_q_after - sales_q
            else:
                lhs_q = _pick_num(x, ("lhs(begin+prod-sales)", "lhs_begin_plus_prod_minus_sales", "lhs_qty"))

            ok = False
            if lhs_q is not None and end_q is not None:
                ok = (_round7(lhs_q) == _round7(end_q))

            check_text = "✅" if ok else "❌"
            check_num = 1 if ok else 0

            ws2.append([
                x.get("sheet"), x.get("row"), x.get("category"), x.get("code"), x.get("name"),
                x.get("prod_qty_before"), x.get("prod_amt_before"),
                x.get("other_qty_before"), x.get("other_amt_before"),
                x.get("prod_qty_after"), x.get("prod_amt_after"),
                x.get("other_qty_after"), x.get("other_amt_after"),
                check_num, check_text,
                lhs_q,
                end_q,
            ])

        # 头部样式（绿色表头）
        header2_fill = PatternFill("solid", fgColor="006100")
        header2_font = Font(color="FFFFFF", bold=True)
        for c in range(1, len(h2) + 1):
            cell = ws2.cell(1, c)
            cell.fill = header2_fill
            cell.font = header2_font
            cell.alignment = center
            cell.border = border

        for r in range(2, ws2.max_row + 1):
            for c in range(1, len(h2) + 1):
                cell = ws2.cell(r, c)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 数字格式美化
            for c in range(6, 14):  # 数量/金额各列
                ws2.cell(r, c).number_format = "0.0000000"
            ws2.cell(r, 16).number_format = "0.0000000"  # lhs
            ws2.cell(r, 17).number_format = "0.0000000"  # 期末

        ws2.freeze_panes = "A2"

        # 列宽更舒服
        col_widths = [
            10, 8, 10, 12, 26,
            16, 18,
            18, 18,
            16, 18,
            18, 18,
            16, 14,
            20, 12,
        ]
        for i, w in enumerate(col_widths, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # ✅ IconSet：在第14列（校验图标）加图标集
        icon_col = 14
        first_data_row = 2
        last_data_row = ws2.max_row
        if last_data_row >= first_data_row:
            rng = f"{get_column_letter(icon_col)}{first_data_row}:{get_column_letter(icon_col)}{last_data_row}"

            # 3图标：0 -> 红叉；1 -> 绿勾（中间图标不会用到）
            # openpyxl 的 IconSetRule 接收 icon_style（常见：'3Symbols', '3TrafficLights1', etc.）
            # '3Symbols2' 通常显示为 ✅ / ! / ❌ 风格更接近你截图
            rule = IconSetRule(
                icon_style="3Symbols2",
                type="num",
                values=[0, 1, 2],   # 分段阈值
                showValue=False,
                percent=False,
                reverse=False
            )
            ws2.conditional_formatting.add(rng, rule)

            # 把数值居中（图标更像表格）
            for r in range(first_data_row, last_data_row + 1):
                ws2.cell(r, icon_col).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_path)
def _norm_product_name(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip()
    if s.endswith("\\中国"):
        s = s[:-3]
    return s.strip().lower()


def audit_sheet_sales_41(ws_cost, sheet_name: str, product_balance_map: dict, balance_sheet_name: str, tol: float = 0.0):
    """
    ✅4-1 审查：
      - 用平衡表 产品表-合并石化部：
          J列=当月销售数量 -> read_product_balance_table 已读为 sales_m（万吨）
          K列=年累计销售数量 -> 已读为 sales_y（万吨）
      - 4-1：
          D=当月销售数量（万吨）
          G=累计销售数量（万吨）
      - 匹配：按物料名称（4-1 的 C列）匹配
      - 输出结构与其它审查一致，追加到审查日志
    """
    issues = []
    seq = 1

    def add_issue(name_raw, cost_row, col_label, cost_v, bal_v, status, bal_name_raw=""):
        nonlocal seq
        diff = (cost_v - bal_v) if (cost_v is not None and bal_v is not None) else None
        issues.append({
            "seq": seq,
            "sheet": sheet_name,
            "balance_sheet": balance_sheet_name,
            "cost_row": cost_row,
            "category": "",
            "code": _clean_code(ws_cost.cell(cost_row, 2).value) if cost_row else "",
            "name": name_raw,
            "balance_name": bal_name_raw,
            "column": col_label,
            "cost_qty": cost_v,
            "balance_qty": bal_v,
            "diff": diff,
            "status": status,
        })
        seq += 1

    def _clean_code(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s.lstrip("0").strip()

    # 4-1 成本表 map：name_norm -> row & vals
    cost_map = {}
    for r in range(4, ws_cost.max_row + 1):
        name = ws_cost.cell(r, 3).value  # C 物料描述
        code = ws_cost.cell(r, 2).value  # B 物料号
        if name is None or str(name).strip() == "":
            continue
        code_s = _clean_code(code)
        name_raw = str(name).strip()
        key = _norm_product_name(name_raw)

        def get_qty(col):
            # D=4 当月数量(万吨)；G=7 累计数量(万吨)
            v = ws_cost.cell(r, col).value
            try:
                if v is None:
                    return 0.0
                if isinstance(v, str) and v.strip() in ("", "-"):
                    return 0.0
                return float(str(v).replace(",", "").strip())
            except Exception:
                return 0.0

        cost_map[key] = {
            "row": r,
            "code": code_s,
            "name_raw": name_raw,
            "sales_m": get_qty(4),
            "sales_y": get_qty(7),
        }

    # 对比计划：当月/累计
    plan = [
        ("当月销售数量", "sales_m", "sales_m"),
        ("年累计销售数量", "sales_y", "sales_y"),
    ]

    # (1) 双方都有：对比数值
    for k, c in cost_map.items():
        if k not in product_balance_map:
            continue
        b = product_balance_map[k]
        if b.get("all_empty_or_zero", False):
            continue

        for col_label, cost_field, bal_field in plan:
            cv = float(c[cost_field] or 0.0)
            bv = b.get(bal_field, None)
            bv = 0.0 if bv is None else float(bv)

            if round(cv, 7) == round(bv, 7):
                continue
            if tol and abs(cv - bv) <= tol:
                continue

            add_issue(
                c["name_raw"], c["row"],
                col_label, cv, bv, "销售数量误差",
                bal_name_raw=b.get("balance_name_raw", "")
            )

    # (2) 平衡表有、4-1没有
    for k, b in product_balance_map.items():
        if k in cost_map:
            continue
        if b.get("all_empty_or_zero", False):
            continue
        add_issue(
            "", None,
            "物料存在性", None, None,
            "4-1中不存在，请人工核查",
            bal_name_raw=b.get("balance_name_raw", "")
        )

    # (3) 4-1有、平衡表没有
    for k, c in cost_map.items():
        if k in product_balance_map:
            continue
        add_issue(
            c["name_raw"], c["row"],
            "物料存在性", None, None,
            "平衡表中不存在，请人工核查",
            bal_name_raw=""
        )

    return issues
