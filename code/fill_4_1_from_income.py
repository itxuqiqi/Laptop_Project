# fill_4_1_from_income.py
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet


def _clean_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0").strip()


def _num_or_none(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
    try:
        return float(v)
    except Exception:
        return None


def _set(ws: Worksheet, r: int, c: int, v):
    ws.cell(r, c, value="-" if v is None else v)


def _build_row_map(ws: Worksheet, start_row: int) -> dict[str, int]:
    m = {}
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)  # B
        if code:
            m[code] = r
    return m


def _find_last_material_row(ws: Worksheet, start_row: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)
        if code:
            last = r
    return last


def _renumber_seq(ws: Worksheet, start_row: int):
    seq = 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)
        if not code:
            continue
        ws.cell(r, 1, value=seq)  # A
        seq += 1


def fill_cost_sheet_4_1_from_income(
    ws_41: Worksheet,
    ws_income: Worksheet,
    current_month: int,
    *,
    income_total_row: int = 3,
    income_start_row: int = 4,
    total_row_41: int = 5,         # ✅模板合计行=第5行
    start_row_41: int = 6,         # ✅明细从第6行开始
):
    """
    从【收入表(sheet=收入)】回填【成本表4-1】。

    4-1 合计行口径：
    - K：合计行当月价 - 合计行上月价（不反推）
    - L：逐物料增利求和
    """

    cm = int(current_month)
    if not (1 <= cm <= 12):
        return

    # --- 收入表列 ---
    COL_IN = {
        "m_qty_wan": 4, "m_price": 5, "m_amt_wan": 6,     # D/E/F
        "y_qty_wan": 7, "y_price": 8, "y_amt_wan": 9,     # G/H/I
        "m1_start": 13,                                   # M
        "prev12_start": 49,                               # AW
    }

    def month_block_start(m: int) -> int:
        return COL_IN["m1_start"] + (m - 1) * 3

    def prev_block_start_for_mom() -> int:
        # 上月块：cm>1 用 (cm-1) 月块；cm=1 用上年12月块(AW)
        return month_block_start(cm - 1) if cm > 1 else COL_IN["prev12_start"]

    # --- 4-1列 ---
    COL_41 = {
        "m_qty_wan": 4, "m_price": 5, "m_amt_wan": 6,
        "y_qty_wan": 7, "y_price": 8, "y_amt_wan": 9,
        "mom_qty_diff": 10, "mom_price_diff": 11, "profit": 12,
        "end_col": 12,
    }

    # --- 行映射 ---
    income_row_map = _build_row_map(ws_income, start_row=income_start_row)
    if not income_row_map:
        return
    codes = sorted(income_row_map.keys())

    row_map_41 = _build_row_map(ws_41, start_row=start_row_41)
    existed_41 = set(row_map_41.keys())

    # 4-1 补齐缺失行（以收入表为准）
    new_codes = [c for c in codes if c not in existed_41]
    if new_codes:
        last_row = _find_last_material_row(ws_41, start_row=start_row_41)
        ptr = last_row + 1 if last_row >= start_row_41 else start_row_41
        for code in new_codes:
            ws_41.insert_rows(ptr)
            ws_41.cell(ptr, 2, value=code)  # B
            r_in = income_row_map.get(code)
            if r_in:
                ws_41.cell(ptr, 3, value=ws_income.cell(r_in, 3).value or "")
            ptr += 1
        row_map_41 = _build_row_map(ws_41, start_row=start_row_41)

    def in_num(r: int, c: int):
        return _num_or_none(ws_income.cell(r, c).value)

    prev_sc = prev_block_start_for_mom()
    prev_qty_col = prev_sc          # 吨
    prev_price_col = prev_sc + 1    # 元/吨

    # ✅逐物料增利求和
    profit_sum = 0.0

    # --- 明细回填 ---
    for code in codes:
        r_in = income_row_map[code]
        r_41 = row_map_41.get(code)
        if not r_41:
            continue

        # 清 4-1 D..L
        for c in range(4, COL_41["end_col"] + 1):
            ws_41.cell(r_41, c, value="-")

        # D..I 透传
        m_qty_wan = in_num(r_in, COL_IN["m_qty_wan"])
        m_price = in_num(r_in, COL_IN["m_price"])
        m_amt_wan = in_num(r_in, COL_IN["m_amt_wan"])

        y_qty_wan = in_num(r_in, COL_IN["y_qty_wan"])
        y_price = in_num(r_in, COL_IN["y_price"])
        y_amt_wan = in_num(r_in, COL_IN["y_amt_wan"])

        _set(ws_41, r_41, COL_41["m_qty_wan"], m_qty_wan)
        _set(ws_41, r_41, COL_41["m_price"], m_price)
        _set(ws_41, r_41, COL_41["m_amt_wan"], m_amt_wan)

        _set(ws_41, r_41, COL_41["y_qty_wan"], y_qty_wan)
        _set(ws_41, r_41, COL_41["y_price"], y_price)
        _set(ws_41, r_41, COL_41["y_amt_wan"], y_amt_wan)

        # J 量差：本月量(缺失=0) - 上月量(缺失=0)
        q_m = m_qty_wan if m_qty_wan is not None else 0.0
        q_prev_ton = in_num(r_in, prev_qty_col)
        q_prev = (q_prev_ton / 10000.0) if (q_prev_ton is not None) else 0.0
        _set(ws_41, r_41, COL_41["mom_qty_diff"], q_m - q_prev)

        # K 价差：只有两月都有价才做差，否则=0
        p_prev = in_num(r_in, prev_price_col)
        if (m_price is not None) and (p_prev is not None):
            pdiff = m_price - p_prev
        else:
            pdiff = 0.0
        _set(ws_41, r_41, COL_41["mom_price_diff"], pdiff)

        # L 增利：价差 * 本月量(万吨)
        profit = pdiff * q_m
        _set(ws_41, r_41, COL_41["profit"], profit)

        profit_sum += profit

    # --- 合计行（4-1 第5行）：D..I 取收入表第3行；K=当月价-上月价；L=逐物料增利求和 ---
    for c in range(4, COL_41["end_col"] + 1):
        ws_41.cell(total_row_41, c, value="-")

    m_qty_wan_tot = in_num(income_total_row, COL_IN["m_qty_wan"])
    m_price_tot = in_num(income_total_row, COL_IN["m_price"])
    m_amt_wan_tot = in_num(income_total_row, COL_IN["m_amt_wan"])

    y_qty_wan_tot = in_num(income_total_row, COL_IN["y_qty_wan"])
    y_price_tot = in_num(income_total_row, COL_IN["y_price"])
    y_amt_wan_tot = in_num(income_total_row, COL_IN["y_amt_wan"])

    _set(ws_41, total_row_41, COL_41["m_qty_wan"], m_qty_wan_tot)
    _set(ws_41, total_row_41, COL_41["m_price"], m_price_tot)
    _set(ws_41, total_row_41, COL_41["m_amt_wan"], m_amt_wan_tot)

    _set(ws_41, total_row_41, COL_41["y_qty_wan"], y_qty_wan_tot)
    _set(ws_41, total_row_41, COL_41["y_price"], y_price_tot)
    _set(ws_41, total_row_41, COL_41["y_amt_wan"], y_amt_wan_tot)

    # 合计量差（仍按合计行吨块）
    q_m_tot = m_qty_wan_tot if m_qty_wan_tot is not None else 0.0
    q_prev_ton_tot = in_num(income_total_row, prev_qty_col)
    q_prev_tot = (q_prev_ton_tot / 10000.0) if (q_prev_ton_tot is not None) else 0.0
    _set(ws_41, total_row_41, COL_41["mom_qty_diff"], q_m_tot - q_prev_tot)

    # ✅合计价差 K：合计当月价 - 合计上月价（仅两月都有价才算，否则0）
    p_prev_tot = in_num(income_total_row, prev_price_col)
    if (m_price_tot is not None) and (p_prev_tot is not None):
        pdiff_tot = m_price_tot - p_prev_tot
    else:
        pdiff_tot = 0.0
    _set(ws_41, total_row_41, COL_41["mom_price_diff"], pdiff_tot)

    # ✅合计增利 L：逐物料增利求和
    _set(ws_41, total_row_41, COL_41["profit"], profit_sum)

    _renumber_seq(ws_41, start_row=start_row_41)
