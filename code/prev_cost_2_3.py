# prev_cost_2_3.py
from openpyxl import load_workbook
from utils import norm, to_float

def read_prev_2_3_map(prev_cost_file: str, sheet_name: str = "2-3"):
    """
    上月成本报表 2-3：
    - B列：物料号
    - C列：产品名称
    - R列：期末库存数量
    - S列：期末库存单价
    返回：dict[(code_norm, name_norm)] = (end_qty, end_price)
    说明：
    - code 可能为空（在途类），允许 key=( '', name )
    """
    wb = load_workbook(prev_cost_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"上月成本报表缺少sheet：{sheet_name}")
    ws = wb[sheet_name]

    m = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 2).value  # B
        name = ws.cell(r, 3).value  # C
        if name is None or str(name).strip() == "":
            continue

        code_s = "" if code is None else str(code).strip().lstrip("0")
        name_s = norm(name)

        end_qty = to_float(ws.cell(r, 18).value)  # R=18
        end_price = to_float(ws.cell(r, 19).value)  # S=19

        m[(code_s, name_s)] = (end_qty, end_price)

    return m
