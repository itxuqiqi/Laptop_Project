# sap_profit_read.py
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook

from profit_utils import normalize_item_name,to_number

def read_sap_profit_table(
    file_path: str,
    header_row: int = 12,
    name_col: int = 2,   # B
    m_col: int = 3,      # C 本月数
    y_col: int = 4,      # D 本年金额
    ly_col: int = 5,     # E 上年同期
    sheet_name: Optional[str] = None,
) -> Dict[str, Dict[str, Optional[float]]]:
    """
    读取 SAP 导出利润表（利新/利旧结构一致）：
    - 第12行表头；从13行开始读
    - B=项目名称，C/D/E=本月/本年/上年同期
    - 读数保留符号（不做 abs）
    返回：key(去空格项目名) -> {"raw": 原项目名, "m":..., "y":..., "ly":...}
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    out: Dict[str, Dict[str, Optional[float]]] = {}
    start_row = header_row + 1

    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        key = normalize_item_name(name)
        if not key:
            continue

        m = to_number(ws.cell(r, m_col).value)
        y = to_number(ws.cell(r, y_col).value)
        ly = to_number(ws.cell(r, ly_col).value)

        # 至少有一列有值才收，避免大量空行污染
        if m is None and y is None and ly is None:
            continue

        if key not in out:
            out[key] = {"raw": str(name).strip() if name is not None else "", "m": m, "y": y, "ly": ly}

    return out


def build_needed_map(
    profit_map: Dict[str, Dict[str, Optional[float]]],
    needed_items: List[str],
) -> Dict[str, Dict[str, Optional[float]]]:
    """
    从全量 profit_map 里抓取 needed_items（按 normalize 后的 key 取）。
    如果缺失则给空结构，避免 KeyError。
    """
    out: Dict[str, Dict[str, Optional[float]]] = {}
    for it in needed_items:
        k = normalize_item_name(it)
        out[k] = profit_map.get(k, {"raw": it, "m": None, "y": None, "ly": None})
    return out
