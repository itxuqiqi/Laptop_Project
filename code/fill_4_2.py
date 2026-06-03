from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

DIESEL_CODE = "82964054"


def _clean_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0").strip()


def _num_or_none(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
    try:
        return float(v)
    except Exception:
        return None


def _set(ws: Worksheet, r: int, c: int, v):
    ws.cell(r, c, value=0 if v is None else v)

def _build_row_map(ws: Worksheet, start_row: int, code_col: int) -> dict[str, int]:
    m = {}
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, code_col).value)
        if code:
            m[code] = r
    return m


def _find_last_material_row(ws: Worksheet, start_row: int, code_col: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, code_col).value)
        if code:
            last = r
    return last


def _renumber_seq(ws: Worksheet, start_row: int, seq_col: int, code_col: int):
    seq = 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, code_col).value)
        if not code:
            continue
        ws.cell(r, seq_col, value=seq)
        seq += 1


def _income_month_block_start(m: int) -> int:
    # income_fill.py：M=13 为1月块，每月3列(吨/元吨/元)
    return 13 + (m - 1) * 3


def _income_prev_price_col(cm: int) -> int:
    # 上月单价列：cm>1 用 cm-1 月块的“单价”列（start+1）
    # cm==1 用 上年12月块：AW=49 qty, AX=50 price, AY=51 amt
    if cm > 1:
        return _income_month_block_start(cm - 1) + 1
    return 50


def _income_prev_qty_col(cm: int) -> int:
    # 上月数量列（吨）：cm>1 用 cm-1 月块 qty 列（start）
    # cm==1 用 上年12月块 qty 列：AW=49
    if cm > 1:
        return _income_month_block_start(cm - 1)
    return 49


def _normalize_vat_rate(vat: float | None) -> float:
    """
    vat 可能填 0.13 或 13（%），统一成小数。
    """
    if vat is None:
        return 0.0
    x = float(vat)
    if x > 1.5:
        x = x / 100.0
    return x


def _calc_bare_tax_price(
    income_price: float | None,
    cons: float | None,
    vat_rate: float | None,
    qty: float | None = None,
) -> float | None:
    """
    裸价税 = 收入单价 - 消费税标准*1.12 - 消费税标准*增值税率*0.12

    ✅最稳口径：用数量约束单价有效性
    - 若 qty 为空或为 0：认为当期无销量/无计价基础，返回 None（最终写 0）
    - 若 income_price 为空：返回 None（最终写 0）
    """
    if income_price is None:
        return None
    if qty is None:
        return None
    try:
        q = float(qty)
    except Exception:
        return None
    if q == 0:
        return None
    p = float(income_price)
    c = float(cons) if cons is not None else 0.0
    vat = _normalize_vat_rate(vat_rate)
    return p - c * 1.12 - c * vat * 0.12



def _guess_bare_sheet(wb) -> Worksheet:
    """
    产销存文件里“裸/祼税价”sheet名字可能不同，做别名+模糊匹配兜底。
    你现在实际sheet名是：祼税价（注意“祼”）
    """
    candidates = [
        "裸价税", "裸税价", "裸价", "裸税", "裸价税价", "裸税价表",
        "祼税价", "祼价税", "祼税",
    ]
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]

    for n in wb.sheetnames:
        s = str(n)
        if (("税价" in s) or ("价税" in s)) and (("裸" in s) or ("祼" in s)):
            return wb[n]

    for n in wb.sheetnames:
        s = str(n)
        if (("裸" in s) or ("祼" in s)) and ("税" in s):
            return wb[n]

    raise ValueError(f"产销存文件中找不到“裸价税/裸税价/祼税价”sheet（当前sheet={wb.sheetnames}）")


def fill_cost_sheet_4_2(
    ws_42: Worksheet,           # 成本表 4-2
    ws_income: Worksheet,       # 产销存 收入 sheet
    ws_bare_src: Worksheet,     # 产销存 裸/祼税价 sheet（提供增值税率/消费税标准/物料列表）
    current_month: int,
    product_ytd_name_map: dict[str, str] | None = None,
    *,
    # 成本表4-2模板：第5行合计，第6行开始明细
    total_row_42: int = 5,
    start_row_42: int = 6,

    # 产销存裸/祼税价sheet模板：通常第5行合计，第6行开始明细
    bare_total_row: int = 5,
    bare_start_row: int = 6,
):
    """
    ✅成本表 4-2 填充逻辑（按你最新要求）：

    数据源：
    - 产销存“祼税价/裸税价”sheet：D=增值税率，E=消费税标准（并给出物料号列表）
    - 产销存“收入”sheet：提供当月/累计/上月 单价（你允许手工调整）

    计算：
      裸价税 = 收入单价 - 消费税标准*1.12 - 消费税标准*增值税率*0.12

    成本表4-2列（按你现有代码口径）：
      A序号 B物料号 C产品名称 D消费税标准
      E本月裸价税 F累计裸价税 G上月裸价税 H环比(本月-上月)
      I本月-国六柴油差价 J累计-国六柴油差价 K上月-国六柴油差价

    ✅新增修正：
    - 明细差价(I/J/K)：如果本物料对应裸税价为空或“-”，差价直接写“-”（不当0减柴油）
    - 合计行第5行：按“加权平均”写入本月/累计/上月裸税价；差价合计直接用合计裸税价减柴油裸税价
      * 本月：SUM(本月裸税价_i * 收入本月数量_i) / 收入本月合计数量
      * 累计：SUM(累计裸税价_i * 收入累计数量_i) / 收入累计合计数量
      * 上月：参考本月（用本月数量做权重）
    """
    cm = int(current_month)
    if not (1 <= cm <= 12):
        return

    # ---- 收入 sheet：单价/数量列 ----
    IN_COL = {
        "m_qty_wan": 4,  # D：当月数量（万吨）
        "m_price": 5,    # E：当月单价（元/吨）
        "y_qty_wan": 7,  # G：累计数量（万吨）
        "y_price": 8,    # H：累计单价（元/吨）
    }
    prev_price_col = _income_prev_price_col(cm)
    prev_qty_col = _income_prev_qty_col(cm)  # 上月数量列（单位按收入sheet月块：通常为“吨”）

    # ---- 产销存祼税价sheet：D=增值税率，E=消费税标准 ----
    SRC = {
        "seq": 1,
        "code": 2,
        "name": 3,
        "vat": 4,   # D
        "cons": 5,  # E
    }

    # ---- 成本表4-2列（你的现代码口径） ----
    COL = {
        "seq": 1,
        "code": 2,
        "name": 3,
        "cons": 4,      # D
        "bare_m": 5,    # E
        "bare_y": 6,    # F
        "bare_p": 7,    # G
        "mom": 8,       # H
        "diff_m": 9,    # I
        "diff_y": 10,   # J
        "diff_p": 11,   # K
        "end_col": 11,
    }

    # ---- 行映射：收入 / 祼税价源 / 成本4-2 ----
    income_row_map = _build_row_map(ws_income, start_row=4, code_col=2)  # 收入从第4行开始
    src_row_map = _build_row_map(ws_bare_src, start_row=bare_start_row, code_col=SRC["code"])
    if not src_row_map:
        return

    codes = sorted(src_row_map.keys())

    row_map_42 = _build_row_map(ws_42, start_row=start_row_42, code_col=COL["code"])
    existed_42 = set(row_map_42.keys())

    def _name_of(code: str) -> str:
        """新增物料名称：按需求从“产品累计数据”获取，找不到再兜底用祼税价源sheet的名称。"""
        if product_ytd_name_map:
            n = product_ytd_name_map.get(code)
            if n:
                return n
        r_src = src_row_map.get(code)
        if r_src:
            return ws_bare_src.cell(r_src, SRC["name"]).value or ""
        return ""

    # ---- 4-2 补齐缺失行（以“祼税价”源 sheet 的物料为准）----
    new_codes = [c for c in codes if c not in existed_42]
    if new_codes:
        last_row = _find_last_material_row(ws_42, start_row=start_row_42, code_col=COL["code"])
        ptr = last_row + 1 if last_row >= start_row_42 else start_row_42
        for code in new_codes:
            ws_42.insert_rows(ptr)
            ws_42.cell(ptr, COL["code"], value=code)
            ws_42.cell(ptr, COL["name"], value=_name_of(code))
            r_src = src_row_map.get(code)
            if r_src:
                cons = _num_or_none(ws_bare_src.cell(r_src, SRC["cons"]).value)
                if cons is not None:
                    ws_42.cell(ptr, COL["cons"], value=cons)
            ptr += 1
        row_map_42 = _build_row_map(ws_42, start_row=start_row_42, code_col=COL["code"])

    # 若模板已存在物料但名称为空，也用产品累计数据补齐一次（不影响已有正确名称）
    if product_ytd_name_map:
        for code, r in row_map_42.items():
            cur = ws_42.cell(r, COL["name"]).value
            if cur is None or str(cur).strip() == "":
                n = product_ytd_name_map.get(code)
                if n:
                    ws_42.cell(r, COL["name"], value=n)

    def in_price(code: str, which: str) -> float | None:
        r = income_row_map.get(code)
        if not r:
            return None
        if which == "m":
            return _num_or_none(ws_income.cell(r, IN_COL["m_price"]).value)
        if which == "y":
            return _num_or_none(ws_income.cell(r, IN_COL["y_price"]).value)
        if which == "p":
            return _num_or_none(ws_income.cell(r, prev_price_col).value)
        return None

    def in_qty(code: str, which: str) -> float | None:
        r = income_row_map.get(code)
        if not r:
            return None
        if which == "m":
            return _num_or_none(ws_income.cell(r, IN_COL["m_qty_wan"]).value)  # 万吨
        if which == "y":
            return _num_or_none(ws_income.cell(r, IN_COL["y_qty_wan"]).value)  # 万吨
        if which == "p":
            return _num_or_none(ws_income.cell(r, prev_qty_col).value)  # 上月数量（按收入sheet月块：通常为吨）
        return None

    # ---- 先算每个物料裸税价（本月/累计/上月）----
    bare_map: dict[str, tuple[float | None, float | None, float | None]] = {}

    for code in codes:
        r_src = src_row_map[code]
        vat = _num_or_none(ws_bare_src.cell(r_src, SRC["vat"]).value)
        cons = _num_or_none(ws_bare_src.cell(r_src, SRC["cons"]).value)

        p_m = in_price(code, "m")
        p_y = in_price(code, "y")
        p_p = in_price(code, "p")

        # ✅最稳口径：用数量约束单价有效性（数量为0则不算裸税价，避免出现纯税项负数常量）
        q_m = in_qty(code, "m")
        q_y = in_qty(code, "y")
        q_p = in_qty(code, "p")

        b_m = _calc_bare_tax_price(p_m, cons, vat, qty=q_m)
        b_y = _calc_bare_tax_price(p_y, cons, vat, qty=q_y)
        b_p = _calc_bare_tax_price(p_p, cons, vat, qty=q_p)
        bare_map[code] = (b_m, b_y, b_p)

    # ---- 柴油参考（可能缺失）----
    diesel_bm, diesel_by, diesel_bp = bare_map.get(DIESEL_CODE, (None, None, None))

    # ---- 明细写入 ----
    for code in codes:
        r_42 = row_map_42.get(code)
        if not r_42:
            continue

        # 同步消费税标准到 D（以源sheet为准）
        r_src = src_row_map.get(code)
        if r_src:
            cons = _num_or_none(ws_bare_src.cell(r_src, SRC["cons"]).value)
            if cons is not None:
                ws_42.cell(r_42, COL["cons"], value=cons)

        # 清 E..K（输出区）
        for c in range(COL["bare_m"], COL["end_col"] + 1):
            ws_42.cell(r_42, c, value=0)

        b_m, b_y, b_p = bare_map.get(code, (None, None, None))

        _set(ws_42, r_42, COL["bare_m"], b_m)
        _set(ws_42, r_42, COL["bare_y"], b_y)
        _set(ws_42, r_42, COL["bare_p"], b_p)

        # 环比：若本月和上月都缺失 → “-”，否则缺失按0参与
        if b_m is None and b_p is None:
            _set(ws_42, r_42, COL["mom"], None)
        else:
            bm0 = b_m if b_m is not None else 0.0
            bp0 = b_p if b_p is not None else 0.0
            _set(ws_42, r_42, COL["mom"], bm0 - bp0)

        # ✅ 比国六柴油差价：本物料裸税价为空/“-” => 差价直接“-”（不当0减柴油）
        # 本月差价
        if b_m is None or diesel_bm is None:
            _set(ws_42, r_42, COL["diff_m"], None)
        else:
            _set(ws_42, r_42, COL["diff_m"], b_m - diesel_bm)

        # 累计差价
        if b_y is None or diesel_by is None:
            _set(ws_42, r_42, COL["diff_y"], None)
        else:
            _set(ws_42, r_42, COL["diff_y"], b_y - diesel_by)

        # 上月差价
        if b_p is None or diesel_bp is None:
            _set(ws_42, r_42, COL["diff_p"], None)
        else:
            _set(ws_42, r_42, COL["diff_p"], b_p - diesel_bp)

    # ---- 合计行（第5行）：加权平均填入 ----
    # 清 E..K
    for c in range(COL["bare_m"], COL["end_col"] + 1):
        ws_42.cell(total_row_42, c, value=0)

    # 读取收入表合计数量（优先用第3行；若无则自己汇总）
    income_total_row = 3
    total_qm = _num_or_none(ws_income.cell(income_total_row, IN_COL["m_qty_wan"]).value)
    total_qy = _num_or_none(ws_income.cell(income_total_row, IN_COL["y_qty_wan"]).value)

    if total_qm is None:
        s = 0.0
        for code in codes:
            q = in_qty(code, "m")
            if q is not None:
                s += q
        total_qm = s if s != 0 else None

    if total_qy is None:
        s = 0.0
        for code in codes:
            q = in_qty(code, "y")
            if q is not None:
                s += q
        total_qy = s if s != 0 else None

    # 本月裸税价合计（加权平均，权重=收入本月数量）
    num_m = 0.0
    has_m = False
    if total_qm is not None and total_qm != 0:
        for code in codes:
            b_m, _, _ = bare_map.get(code, (None, None, None))
            q_m = in_qty(code, "m")
            if b_m is None or q_m is None:
                continue
            has_m = True
            num_m += b_m * q_m
        bare_m_tot = (num_m / total_qm) if has_m else None
    else:
        bare_m_tot = None

    # 累计裸税价合计（加权平均，权重=收入累计数量）
    num_y = 0.0
    has_y = False
    if total_qy is not None and total_qy != 0:
        for code in codes:
            _, b_y, _ = bare_map.get(code, (None, None, None))
            q_y = in_qty(code, "y")
            if b_y is None or q_y is None:
                continue
            has_y = True
            num_y += b_y * q_y
        bare_y_tot = (num_y / total_qy) if has_y else None
    else:
        bare_y_tot = None

    # 上月裸税价合计：沿用“本月合计裸税价”的方法，但权重=上月数量、分母=上月合计数量

    def in_prev_qty_ton(code: str) -> float | None:
        r = income_row_map.get(code)
        if not r:
            return None
        return _num_or_none(ws_income.cell(r, prev_qty_col).value)  # 吨

    # 上月合计数量（吨）：优先取收入合计行；取不到则自己汇总
    total_qp_ton = _num_or_none(ws_income.cell(income_total_row, prev_qty_col).value)
    if total_qp_ton is None:
        s = 0.0
        for code in codes:
            q = in_prev_qty_ton(code)
            if q is not None:
                s += q
        total_qp_ton = s if s != 0 else None

    num_p = 0.0
    has_p = False
    if total_qp_ton is not None and total_qp_ton != 0:
        for code in codes:
            _, _, b_p = bare_map.get(code, (None, None, None))
            q_p = in_prev_qty_ton(code)  # ✅上月数量（吨）
            if b_p is None or q_p is None:
                continue
            has_p = True
            num_p += b_p * q_p
        bare_p_tot = (num_p / total_qp_ton) if has_p else None
    else:
        bare_p_tot = None

    _set(ws_42, total_row_42, COL["bare_m"], bare_m_tot)
    _set(ws_42, total_row_42, COL["bare_y"], bare_y_tot)
    _set(ws_42, total_row_42, COL["bare_p"], bare_p_tot)

    # 合计环比：两者都有值才算
    if bare_m_tot is None or bare_p_tot is None:
        _set(ws_42, total_row_42, COL["mom"], None)
    else:
        _set(ws_42, total_row_42, COL["mom"], bare_m_tot - bare_p_tot)

    # 合计差价：用合计裸税价直接减柴油裸税价
    if bare_m_tot is None or diesel_bm is None:
        _set(ws_42, total_row_42, COL["diff_m"], None)
    else:
        _set(ws_42, total_row_42, COL["diff_m"], bare_m_tot - diesel_bm)

    if bare_y_tot is None or diesel_by is None:
        _set(ws_42, total_row_42, COL["diff_y"], None)
    else:
        _set(ws_42, total_row_42, COL["diff_y"], bare_y_tot - diesel_by)

    if bare_p_tot is None or diesel_bp is None:
        _set(ws_42, total_row_42, COL["diff_p"], None)
    else:
        _set(ws_42, total_row_42, COL["diff_p"], bare_p_tot - diesel_bp)

    _renumber_seq(ws_42, start_row=start_row_42, seq_col=COL["seq"], code_col=COL["code"])