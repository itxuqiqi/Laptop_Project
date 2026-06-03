# revenue_fill.py
from __future__ import annotations
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from utils import to_float, norm


def _set(ws: Worksheet, r: int, c: int, v):
    ws.cell(r, c, value="-" if v is None else v)


def _write_qty_price_amt_ton(ws: Worksheet, r: int, c_qty: int, qty_ton, amt_yuan):
    if qty_ton is None and amt_yuan is None:
        _set(ws, r, c_qty, None)
        _set(ws, r, c_qty + 1, None)
        _set(ws, r, c_qty + 2, None)
        return

    q = None if qty_ton is None else float(qty_ton)
    a = None if amt_yuan is None else float(amt_yuan)

    _set(ws, r, c_qty, q)
    _set(ws, r, c_qty + 2, a)

    if q is None or q == 0 or a is None:
        _set(ws, r, c_qty + 1, None)
    else:
        _set(ws, r, c_qty + 1, a / q)


def _write_qty_price_amt_wan(ws: Worksheet, r: int, c_qty: int, qty_ton, amt_yuan):
    """
    万口径：
      qty = 吨/10000 -> 万吨
      amt = 元/10000 -> 万元
      price = 元/吨（不缩放）
    """
    if qty_ton is None and amt_yuan is None:
        _set(ws, r, c_qty, None)
        _set(ws, r, c_qty + 1, None)
        _set(ws, r, c_qty + 2, None)
        return

    q_ton = None if qty_ton is None else float(qty_ton)
    a_y = None if amt_yuan is None else float(amt_yuan)

    q_wan = None if q_ton is None else q_ton / 10000.0
    a_wan = None if a_y is None else a_y / 10000.0

    _set(ws, r, c_qty, q_wan)
    _set(ws, r, c_qty + 2, a_wan)

    if q_ton is None or q_ton == 0 or a_y is None:
        _set(ws, r, c_qty + 1, None)
    else:
        _set(ws, r, c_qty + 1, a_y / q_ton)


def _find_last_material_row(ws: Worksheet, start_row: int, code_col: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, code_col).value
        if v is None or str(v).strip() == "":
            continue
        last = r
    return last


def _build_code_row_map(ws: Worksheet, start_row: int, code_col: int) -> dict[str, int]:
    m = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, code_col).value
        if v is None or str(v).strip() == "":
            continue
        code = str(v).strip().lstrip("0")
        m[norm(code)] = r
    return m


def _reindex_seq(ws: Worksheet, start_row: int, seq_col: int, code_col: int):
    seq = 1
    for r in range(start_row, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if code is None or str(code).strip() == "":
            continue
        ws.cell(r, seq_col, value=seq)
        seq += 1


def fill_revenue_income_sheet(ws_income: Worksheet, sales_pack: dict):
    """
    产销存-收入表 sheet="收入" 填充

    表结构（你描述）：
      行：第3行合计；第4行开始物料
      列：
        A 序号
        B 物料号
        C 物料描述
        D-F 当月（万口径：D数量万吨 / E单价元 / F金额万元）
        G-I 1-12月（万口径：G数量万吨 / H单价元 / I金额万元）
        J-L 1-12月（吨口径：J数量吨 / K单价元 / L金额元）
        M-O 1月（吨/元）
        P-R 2月 ...
        ...
        AT-AV 12月（吨/元）
        AW-AY 上年12月（吨/元）
    """
    year = int(sales_pack["year"])
    cur_m = int(sales_pack["current_month"])
    prev_y = int(sales_pack["prev_year"])
    prev_m = int(sales_pack["prev_month"])

    month_map = sales_pack.get("month_map", {}) or {}
    ytd_map = sales_pack.get("ytd_map", {}) or {}
    cur_map = sales_pack.get("cur_map", {}) or {}

    START_ROW = 4
    COL_SEQ = 1
    COL_CODE = 2
    COL_DESC = 3

    # 固定块列
    # 当月：D-F
    COL_CUR_WAN = 4
    # 1-12月（万口径）：G-I
    COL_YTD_WAN = 7
    # 1-12月（吨口径）：J-L
    COL_YTD_TON = 10

    # 月度块：1月 M-O 起始列=13，每月+3列；12月起始=46(AT)
    def month_block_start(m: int) -> int:
        return 13 + (m - 1) * 3

    COL_PREV_DEC = 49  # AW(49) AY(51)

    # 现有物料映射
    code_row = _build_code_row_map(ws_income, START_ROW, COL_CODE)
    last_row = _find_last_material_row(ws_income, START_ROW, COL_CODE)

    # 插入新增物料
    materials = sorted(list(sales_pack.get("materials", set())))
    for code_norm in materials:
        if code_norm in code_row:
            continue
        ins = last_row + 1 if last_row >= START_ROW else START_ROW
        ws_income.insert_rows(ins)
        ws_income.cell(ins, COL_CODE, value=code_norm)
        ws_income.cell(ins, COL_DESC, value="")  # 底表无描述，留空
        code_row[code_norm] = ins
        last_row += 1

    # 填充每个物料
    for code_norm, r in code_row.items():
        # 当月（万口径）
        cur = cur_map.get(code_norm, {"qty_ton": None, "amt_yuan": None, "price": None})
        _write_qty_price_amt_wan(ws_income, r, COL_CUR_WAN, cur.get("qty_ton"), cur.get("amt_yuan"))

        # 1..cur_m 累计（万口径 & 吨口径）
        ytd = ytd_map.get(code_norm, {"qty_ton": None, "amt_yuan": None, "price": None})
        _write_qty_price_amt_wan(ws_income, r, COL_YTD_WAN, ytd.get("qty_ton"), ytd.get("amt_yuan"))
        _write_qty_price_amt_ton(ws_income, r, COL_YTD_TON, ytd.get("qty_ton"), ytd.get("amt_yuan"))

        # 1~12 月块（吨口径）
        for m in range(1, 13):
            rec = month_map.get((code_norm, year, m), {"qty_ton": None, "amt_yuan": None})
            c0 = month_block_start(m)
            _write_qty_price_amt_ton(ws_income, r, c0, rec.get("qty_ton"), rec.get("amt_yuan"))

        # 上年12月（吨口径）
        rec_prev_dec = month_map.get((code_norm, prev_y, 12), {"qty_ton": None, "amt_yuan": None})
        _write_qty_price_amt_ton(ws_income, r, COL_PREV_DEC, rec_prev_dec.get("qty_ton"), rec_prev_dec.get("amt_yuan"))

    # 重排序号
    _reindex_seq(ws_income, START_ROW, COL_SEQ, COL_CODE)
