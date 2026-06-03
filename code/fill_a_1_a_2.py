# fill_a_1_a_2.py
from __future__ import annotations

from typing import Any, Dict, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet


# ==========================================================
# 基础工具
# ==========================================================
def to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def norm(x: Any) -> str:
    """
    ✅用于锚点/项目名统一匹配：
    - 去空格/全角空格/制表符
    - 括号统一：中文（） -> ()
    - 各种横线统一
    """
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\r", "").replace("\n", "").replace("\t", "")
    s = s.replace("\u3000", "").replace(" ", "").strip()
    s = s.replace("（", "(").replace("）", ")")
    s = (
        s.replace("－", "-")
        .replace("—", "-")
        .replace("–", "-")
        .replace("−", "-")
    )
    return s


def safe_div(a: float, b: float) -> Optional[float]:
    if b == 0:
        return None
    return a / b


# ==========================================================
# A表：锚点索引（A列可能为空，跳过）
# ==========================================================
def build_anchor_index(ws: Worksheet, start_row: int = 5, anchor_col: int = 1) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, anchor_col).value
        if v is None:
            continue
        key = norm(v)
        if not key:
            continue
        if key not in idx:
            idx[key] = r
    return idx


def find_row_by_anchor(ws: Worksheet, anchor: str, start_row: int = 1, anchor_col: int = 1) -> Optional[int]:
    target = norm(anchor)
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, anchor_col).value
        if v is None:
            continue
        if norm(v) == target:
            return r
    return None


# ==========================================================
# 读取指定锚点行的 qty/price/amt
# ==========================================================
def read_triplet(ws: Worksheet, anchor: str,
                col_qty: int, col_price: int, col_amt: int,
                anchor_col: int = 1, start_row: int = 1) -> Tuple[float, float, float]:
    r = find_row_by_anchor(ws, anchor, start_row=start_row, anchor_col=anchor_col)
    if r is None:
        return 0.0, 0.0, 0.0
    q = to_float(ws.cell(r, col_qty).value)
    p = to_float(ws.cell(r, col_price).value)
    a = to_float(ws.cell(r, col_amt).value)
    return q, p, a


# ==========================================================
# 特殊：按物料号定位行（2-1外购原料油，C列=物料号）
# ==========================================================
def find_row_by_material_no(ws: Worksheet, material_no: str, col_material: int = 3, start_row: int = 1) -> Optional[int]:
    target = str(material_no).strip()
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col_material).value
        if v is None:
            continue
        if str(v).strip() == target:
            return r
    return None


def read_triplet_by_material(ws: Worksheet, material_no: str,
                            col_qty: int, col_price: int, col_amt: int,
                            col_material: int = 3, start_row: int = 1) -> Tuple[float, float, float]:
    r = find_row_by_material_no(ws, material_no, col_material=col_material, start_row=start_row)
    if r is None:
        return 0.0, 0.0, 0.0
    q = to_float(ws.cell(r, col_qty).value)
    p = to_float(ws.cell(r, col_price).value)
    a = to_float(ws.cell(r, col_amt).value)
    return q, p, a


# ==========================================================
# 写入 A表：锚点A列，D/E/F = 数量/单价/金额
# ✅覆盖空：qty/price/amt 如果传 None 也会写入 None
# ==========================================================
def write_a(ws: Worksheet, idx: Dict[str, int], anchor: str,
            qty: Optional[float], price: Optional[float], amt: Optional[float]):
    key = norm(anchor)
    if key not in idx:
        print(f"DEBUG ❌ A表未找到锚点行: {anchor}")
        return
    r = idx[key]

    ws.cell(r, 4).value = qty
    ws.cell(r, 5).value = price
    ws.cell(r, 6).value = amt


# ==========================================================
# 读取 1-2：A列锚点，E=本月，F=累计，G=本月含芳烃II
# ==========================================================
def read_1_2_vals(ws12: Worksheet, anchor: str) -> Tuple[float, float, float]:
    r = find_row_by_anchor(ws12, anchor, start_row=5, anchor_col=1)
    if r is None:
        return 0.0, 0.0, 0.0
    e = to_float(ws12.cell(r, 5).value)  # E
    f = to_float(ws12.cell(r, 6).value)  # F
    g = to_float(ws12.cell(r, 7).value)  # G
    return e, f, g


# ==========================================================
# 读取 6 sheet：利润总额
# 结构：B列=项目名，D=本月，E=本年金额
# ==========================================================
def read_profit_sheet_6_amt(ws6: Worksheet, target_name: str,
                            name_col: int = 2, col_m: int = 4, col_y: int = 5,
                            start_row: int = 1) -> Tuple[float, float]:
    target_key = norm(target_name)
    for r in range(start_row, ws6.max_row + 1):
        v = ws6.cell(r, name_col).value
        if not v:
            continue
        if norm(v) == target_key:
            m = to_float(ws6.cell(r, col_m).value)
            y = to_float(ws6.cell(r, col_y).value)
            return m, y
    return 0.0, 0.0


# ==========================================================
# 主填充：A-1（本月） + A-2（累计）
# ==========================================================
def fill_cost_sheet_a_1_a_2(
    ws_a1: Worksheet,
    ws_a2: Worksheet,
    ws_21: Worksheet,
    ws_22: Worksheet,
    ws_31: Worksheet,
    ws_32: Worksheet,
    ws_41: Worksheet,
    ws_12: Worksheet,
    ws_6: Worksheet,
    ws_7: Worksheet,
    start_row_a: int = 5,
):
    """
    ✅A-1 本月口径
    ✅A-2 累计口径
    """

    idx_a1 = build_anchor_index(ws_a1, start_row=start_row_a, anchor_col=1)
    idx_a2 = build_anchor_index(ws_a2, start_row=start_row_a, anchor_col=1)

    # 2-1/2-2 列位：采购 H/I/J，加工 K/L/M，库存 Q/R/S
    COL_BUY_Q, COL_BUY_P, COL_BUY_A = 8, 9, 10        # H I J
    COL_PROC_Q, COL_PROC_P, COL_PROC_A = 11, 12, 13   # K L M
    COL_STK_Q, COL_STK_P, COL_STK_A = 17, 18, 19      # Q R S

    # 3-1/3-2：生产 H/I/J，销售 K/L/M，库存 Q/R/S
    COL3_PROD_Q, COL3_PROD_P, COL3_PROD_A = 8, 9, 10
    COL3_SALE_Q, COL3_SALE_P, COL3_SALE_A = 11, 12, 13
    COL3_STK_Q, COL3_STK_P, COL3_STK_A = 17, 18, 19

    # ==========================================================
    # ① 采购部分（2-1/2-2 的采购 H-I-J）
    # ==========================================================
    def fill_buy(anchor_a: str, anchor_2: str):
        q1, p1, a1 = read_triplet(ws_21, anchor_2, COL_BUY_Q, COL_BUY_P, COL_BUY_A)
        q2, p2, a2 = read_triplet(ws_22, anchor_2, COL_BUY_Q, COL_BUY_P, COL_BUY_A)
        write_a(ws_a1, idx_a1, anchor_a, q1, p1, a1)
        write_a(ws_a2, idx_a2, anchor_a, q2, p2, a2)

    fill_buy("采购-原油采购(一般贸易)", "原油(一般贸易)")
    fill_buy("采购-原油采购(一般贸易)-进口原油", "原油(一般贸易)-进口原油")
    fill_buy("采购-原油采购(一般贸易)-海洋原油", "原油(一般贸易)-海洋原油")
    fill_buy("采购-原油采购(一般贸易)-来料加工原油", "来料加工-合计")

    # 天然气采购：2-1 外购原料油 C列物料号=81062610 行 采购 H-I-J
    q1, p1, a1 = read_triplet_by_material(ws_21, "81062610", COL_BUY_Q, COL_BUY_P, COL_BUY_A, col_material=3)
    q2, p2, a2 = read_triplet_by_material(ws_22, "81062610", COL_BUY_Q, COL_BUY_P, COL_BUY_A, col_material=3)
    write_a(ws_a1, idx_a1, "采购-原油采购(一般贸易)-天然气采购", q1, p1, a1)
    write_a(ws_a2, idx_a2, "采购-原油采购(一般贸易)-天然气采购", q2, p2, a2)

    # ==========================================================
    # ② 生产部分（2-1/2-2 的加工 K-L-M）
    # ==========================================================
    def fill_proc(anchor_a: str, anchor_2: str):
        q1, p1, a1 = read_triplet(ws_21, anchor_2, COL_PROC_Q, COL_PROC_P, COL_PROC_A)
        q2, p2, a2 = read_triplet(ws_22, anchor_2, COL_PROC_Q, COL_PROC_P, COL_PROC_A)
        write_a(ws_a1, idx_a1, anchor_a, q1, p1, a1)
        write_a(ws_a2, idx_a2, anchor_a, q2, p2, a2)

    fill_proc("生产-原油加工(一般贸易)", "原油(一般贸易)")
    fill_proc("生产-来料加工原油", "来料加工-合计")

    # 吨油指标：只填单价
    e_full, f_full, _ = read_1_2_vals(ws_12, "完全费用")
    write_a(ws_a1, idx_a1, "生产-吨油完全费用", None, e_full, None)
    write_a(ws_a2, idx_a2, "生产-吨油完全费用", None, f_full, None)

    e_proc, f_proc, _ = read_1_2_vals(ws_12, "指标-吨油加工成本")
    write_a(ws_a1, idx_a1, "生产-吨油加工成本", None, e_proc, None)
    write_a(ws_a2, idx_a2, "生产-吨油加工成本", None, f_proc, None)

    # 吨油变动费用 = 1-2 变动费用小计（E/F）
    e_var, f_var, _ = read_1_2_vals(ws_12, "变动费用小计")
    write_a(ws_a1, idx_a1, "生产-其中：吨油变动费用", None, e_var, None)
    write_a(ws_a2, idx_a2, "生产-其中：吨油变动费用", None, f_var, None)

    # 吨油期间费用 = 1-2 指标-吨油期间费用 + 指标-研发费用
    e_pf, f_pf, _ = read_1_2_vals(ws_12, "指标-吨油期间费用")
    e_rd, f_rd, _ = read_1_2_vals(ws_12, "指标-研发费用")
    write_a(ws_a1, idx_a1, "生产-吨油期间费用", None, e_pf + e_rd, None)
    write_a(ws_a2, idx_a2, "生产-吨油期间费用", None, f_pf + f_rd, None)

    # 产成品（成本）= 3-1 一般贸易-合计 本期生产 H/I/J；累计=3-2
    def fill_3_prod(anchor_a: str, anchor_3: str):
        q1, p1, a1 = read_triplet(ws_31, anchor_3, COL3_PROD_Q, COL3_PROD_P, COL3_PROD_A)
        q2, p2, a2 = read_triplet(ws_32, anchor_3, COL3_PROD_Q, COL3_PROD_P, COL3_PROD_A)
        write_a(ws_a1, idx_a1, anchor_a, q1, p1, a1)
        write_a(ws_a2, idx_a2, anchor_a, q2, p2, a2)

    fill_3_prod("生产-产成品（成本）", "一般贸易-合计")
    fill_3_prod("生产-来料加工产品", "来料加工-合计")

    # ==========================================================
    # ③ 销售部分
    # ==========================================================
    # 销售收入：4-1 第五行 D-F=本月，G-I=累计
    r5 = 5
    q1 = to_float(ws_41.cell(r5, 4).value)
    p1 = to_float(ws_41.cell(r5, 5).value)
    a1 = to_float(ws_41.cell(r5, 6).value)
    q2 = to_float(ws_41.cell(r5, 7).value)
    p2 = to_float(ws_41.cell(r5, 8).value)
    a2 = to_float(ws_41.cell(r5, 9).value)
    write_a(ws_a1, idx_a1, "销售-产品销售收入", q1, p1, a1)
    write_a(ws_a2, idx_a2, "销售-产品销售收入", q2, p2, a2)

    # 销售成本：3-1/3-2 一般贸易-合计 本期销售 K/L/M
    def fill_3_sale(anchor_a: str, anchor_3: str):
        q1, p1, a1 = read_triplet(ws_31, anchor_3, COL3_SALE_Q, COL3_SALE_P, COL3_SALE_A)
        q2, p2, a2 = read_triplet(ws_32, anchor_3, COL3_SALE_Q, COL3_SALE_P, COL3_SALE_A)
        write_a(ws_a1, idx_a1, anchor_a, q1, p1, a1)
        write_a(ws_a2, idx_a2, anchor_a, q2, p2, a2)

    fill_3_sale("销售-产品销售成本", "一般贸易-合计")

    # 利润总额：金额来自6sheet；单价=金额/(两种加工数量之和)
    profit_name = "****四、利润总额（亏损总额以“-”号填列）"
    m_amt, y_amt = read_profit_sheet_6_amt(ws_6, profit_name, name_col=2, col_m=4, col_y=5)

    prod1_q_m, _, _ = read_triplet(ws_21, "原油(一般贸易)", COL_PROC_Q, COL_PROC_P, COL_PROC_A)
    prod2_q_m, _, _ = read_triplet(ws_21, "来料加工-合计", COL_PROC_Q, COL_PROC_P, COL_PROC_A)
    denom_m = prod1_q_m + prod2_q_m

    prod1_q_y, _, _ = read_triplet(ws_22, "原油(一般贸易)", COL_PROC_Q, COL_PROC_P, COL_PROC_A)
    prod2_q_y, _, _ = read_triplet(ws_22, "来料加工-合计", COL_PROC_Q, COL_PROC_P, COL_PROC_A)
    denom_y = prod1_q_y + prod2_q_y

    m_price = safe_div(m_amt, denom_m)
    y_price = safe_div(y_amt, denom_y)

    write_a(ws_a1, idx_a1, "销售-利润总额", None, m_price, m_amt)
    write_a(ws_a2, idx_a2, "销售-利润总额", None, y_price, y_amt)

    # ==========================================================
    # ④ 库存情况
    # ==========================================================
    # 原油库存：2-1/2-2 原油(一般贸易) Q/R/S
    def fill_stock(anchor_a: str, anchor_2: str):
        q1, p1, a1 = read_triplet(ws_21, anchor_2, COL_STK_Q, COL_STK_P, COL_STK_A)
        q2, p2, a2 = read_triplet(ws_22, anchor_2, COL_STK_Q, COL_STK_P, COL_STK_A)
        write_a(ws_a1, idx_a1, anchor_a, q1, p1, a1)
        write_a(ws_a2, idx_a2, anchor_a, q2, p2, a2)

    fill_stock("库存情况-原油", "原油(一般贸易)")

    # 产成品库存：3-1/3-2 一般贸易-合计 Q/R/S
    def fill_stock_3(anchor_a: str, anchor_3: str):
        q1, p1, a1 = read_triplet(ws_31, anchor_3, COL3_STK_Q, COL3_STK_P, COL3_STK_A)
        q2, p2, a2 = read_triplet(ws_32, anchor_3, COL3_STK_Q, COL3_STK_P, COL3_STK_A)
        write_a(ws_a1, idx_a1, anchor_a, q1, p1, a1)
        write_a(ws_a2, idx_a2, anchor_a, q2, p2, a2)

    fill_stock_3("库存情况-产成品", "一般贸易-合计")

    # 半成品库存：sheet7 第五行 D/E/F，本月+累计都取这三个值
    semi_q = to_float(ws_7.cell(5, 4).value)
    semi_p = to_float(ws_7.cell(5, 5).value)
    semi_a = to_float(ws_7.cell(5, 6).value)
    write_a(ws_a1, idx_a1, "库存情况-半成品", semi_q, semi_p, semi_a)
    write_a(ws_a2, idx_a2, "库存情况-半成品", semi_q, semi_p, semi_a)

    # 其中：来料加工产品库存：取 3-2 来料加工-合计 Q/R/S，A-1/A-2一致
    q_lp, p_lp, a_lp = read_triplet(ws_32, "来料加工-合计", COL3_STK_Q, COL3_STK_P, COL3_STK_A)
    write_a(ws_a1, idx_a1, "库存情况-其中：来料加工产品", q_lp, p_lp, a_lp)
    write_a(ws_a2, idx_a2, "库存情况-其中：来料加工产品", q_lp, p_lp, a_lp)

    # 来料加工原油（含产品）
    # 数量口径：2-2 来料加工-合计 库存数量(Q) + (来料加工产品库存数量 / 0.84)
    # 金额口径：只取 2-2 来料加工-合计 库存金额(S)（来料加工产品金额单独在“其中：来料加工产品”反映）
    q_lc, _, a_lc = read_triplet(ws_22, "来料加工-合计", COL_STK_Q, COL_STK_P, COL_STK_A)
    q_total = q_lc + (q_lp / 0.84 if q_lp else 0.0)
    p_total = safe_div(a_lc, q_total)
    write_a(ws_a1, idx_a1, "库存情况-来料加工原油（含产品）", q_total, p_total, a_lc)
    write_a(ws_a2, idx_a2, "库存情况-来料加工原油（含产品）", q_total, p_total, a_lc)

    # 在途原油：取 2-2 在途原油-合计 Q/R/S，A-1/A-2一致
    q_it, p_it, a_it = read_triplet(ws_22, "在途原油-合计", COL_STK_Q, COL_STK_P, COL_STK_A)
    write_a(ws_a1, idx_a1, "库存情况-在途原油", q_it, p_it, a_it)
    write_a(ws_a2, idx_a2, "库存情况-在途原油", q_it, p_it, a_it)

    # 库存情况（汇总行）
    # 口径：
    #   数量 = 库存情况-原油、库存情况-产成品、库存情况-半成品、库存情况-来料加工原油（含产品）、库存情况-在途原油 数量之和
    #        -（库存情况-其中：来料加工产品数量 / 0.84）
    #   金额 = 库存情况-原油、库存情况-产成品、库存情况-半成品、库存情况-来料加工原油（含产品）、库存情况-在途原油、库存情况-其中：来料加工产品 金额之和
    #   单价 = 金额 / 数量

    # A-1（本月）
    q_o1, _, a_o1 = read_triplet(ws_21, "原油(一般贸易)", COL_STK_Q, COL_STK_P, COL_STK_A)
    q_p1, _, a_p1 = read_triplet(ws_31, "一般贸易-合计", COL3_STK_Q, COL3_STK_P, COL3_STK_A)
    qty_a1 = q_o1 + q_p1 + semi_q + q_total + q_it - (q_lp / 0.84 if q_lp else 0.0)
    amt_a1 = a_o1 + a_p1 + semi_a + a_lc + a_it + a_lp
    price_a1 = safe_div(amt_a1, qty_a1)
    write_a(ws_a1, idx_a1, "库存情况", qty_a1, price_a1, amt_a1)

    # A-2（累计）
    q_o2, _, a_o2 = read_triplet(ws_22, "原油(一般贸易)", COL_STK_Q, COL_STK_P, COL_STK_A)
    q_p2, _, a_p2 = read_triplet(ws_32, "一般贸易-合计", COL3_STK_Q, COL3_STK_P, COL3_STK_A)
    qty_a2 = q_o2 + q_p2 + semi_q + q_total + q_it - (q_lp / 0.84 if q_lp else 0.0)
    amt_a2 = a_o2 + a_p2 + semi_a + a_lc + a_it + a_lp
    price_a2 = safe_div(amt_a2, qty_a2)
    write_a(ws_a2, idx_a2, "库存情况", qty_a2, price_a2, amt_a2)

    print("✅ fill_cost_sheet_a_1_a_2 完成")
