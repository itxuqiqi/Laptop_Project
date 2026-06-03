
from registry import SHEETS
from income_fill import fill_income_ws, ensure_new_codes_in_bare_tax_ws, guess_bare_tax_sheet_name

def expand_sheet_deps(selected_sheets):
    """递归展开 sheet 的 depends_on 依赖。"""
    expanded = set(selected_sheets)
    changed = True
    while changed:
        changed = False
        for sh in list(expanded):
            deps = SHEETS.get(sh, {}).get("depends_on", []) or []
            for d in deps:
                if d not in expanded:
                    expanded.add(d)
                    changed = True
    return list(expanded)

# runner.py (完整版本，包含所有修改)

import os
import pandas as pd
from openpyxl import load_workbook

from utils import find_row_by_a_contains
from oil_insert import check_and_update_oil_data
from financial_fill import fill_financial_data
from anchor_fill import fill_anchor_rows
from intransit_read import read_intransit_estimate, read_intransit_usd_price_pack
from intransit_write import write_intransit_to_cost
from intransit_summaries import write_intransit_summaries
from ton_barrel import read_ton_barrel_ratio_map, apply_ton_barrel_and_usd_per_bbl
from intransit_price_pack import build_intransit_price_pack
from usd_anchor_backfill import backfill_usd_anchor_from_23
from fill_4_1_from_income import fill_cost_sheet_4_1_from_income  # ✅新增

from sap_profit_read import read_sap_profit_table
from profit_sheet6_fill import fill_sheet6_profit
from profit_utils import normalize_item_name
from audit_check import (
    read_balance_table, audit_sheet,
    read_product_balance_table, audit_sheet_product,
    save_audit_excel,
    audit_sheet_sales_41,   # ✅新增
)

from sheet_2_3_fill import fill_sheet_2_3

from product_read import read_product_sheet1_pack
from product_insert import insert_new_products_into_general_trade
from product_fill import fill_product_sheet_3x
from product_summaries import write_product_summaries

try:
    from general_trade_summaries import write_general_trade_summaries
except Exception:
    write_general_trade_summaries = None

# ✅新增：销售收入 pack + 两个填充器
from sales_revenue_read import build_sales_revenue_pack
from fill_4_1 import fill_cost_sheet_4_1
from income_fill import fill_income_ws
from semi_inventory_fill import (
    read_semi_inventory_sap,
    fill_cost_sheet_7_semi_inventory,
)

# ✅新增：1-1 生产经营费用表
from fill_1_1 import fill_cost_sheet_1_1

def _read_oil_df(file_path: str) -> pd.DataFrame:
    df = pd.read_excel(file_path, sheet_name="Sheet1", dtype={"物料号": str})
    if "物料描述" in df.columns:
        df["物料描述"] = df["物料描述"].astype(str).str.strip().str.lower()
    if "物料号" in df.columns:
        df["物料号"] = df["物料号"].apply(lambda x: str(x).strip().lstrip("0") if x is not None else "")
    return df


def _clean_code_for_map(v) -> str:
    """对齐各表的物料号清洗口径：去 .0 / 去前导0 / strip。"""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0").strip()


def _pipeline(ws_cost, wb_cost, df_oil, fx_rate: float, ratio_map: dict, intransit_items: list):
    check_and_update_oil_data(df_oil, ws_cost, wb_cost)
    fill_financial_data(df_oil, ws_cost)

    if write_general_trade_summaries is not None:
        write_general_trade_summaries(ws_cost)

    fill_anchor_rows(ws_cost, fx_rate)
    apply_ton_barrel_and_usd_per_bbl(ws_cost, ratio_map, fx_rate)

    write_intransit_to_cost(ws_cost, intransit_items)
    write_intransit_summaries(ws_cost)


def _auto_audit_path(cost_out_path: str) -> str:
    base, _ = os.path.splitext(cost_out_path)
    return f"{base}_审查日志.xlsx"


def _auto_income_out_path(cost_out_path: str) -> str:
    base, _ = os.path.splitext(cost_out_path)
    return f"{base}_产销存收入表.xlsx"


def run_workbook(
    selected_sheets: list,
    sources: dict,
    cost_out_path: str,
    audit_out_path: str | None = None,
    progress_cb=None,
):
    """
    返回：(wb_cost, audit_path, income_out_path)
    income_out_path 仅当选择了 "收入" 时才会生成，否则为 None
    """

    def _ensure_order_and_dependencies(sel: list):
        s = list(dict.fromkeys(sel))

        # 2-3 依赖 2-1/2-2
        if "2-3" in s:
            if "2-1" not in s:
                s.insert(0, "2-1")
            if "2-2" not in s:
                idx = s.index("2-1")
                s.insert(idx + 1, "2-2")

        order = {
            "2-1": 1, "2-2": 2, "2-3": 3,
            "3-1": 4, "3-2": 5,
            "收入": 6,
            "4-1": 7,
            "4-1(按收入表回填)": 8,
            "4-1（按收入表回填）": 8,
            "4-2": 9,
            "6": 10,  # ✅新增：利润表顺序
            "7": 11,  # ✅新增：半成品库存表顺序
            "1-1": 12,  # ✅新增：生产经营费用表顺序
            "1-2": 13,  # ✅新增：生产经营费用表/原料数量顺序
            "A-1":14,
            "A-2": 15,
        }
        if "A-1" in s and "A-2" not in s:
            s.append("A-2")
        if "A-2" in s and "A-1" not in s:
            s.append("A-1")

        s.sort(key=lambda x: order.get(x, 999))
        return s

    selected_sheets = _ensure_order_and_dependencies(selected_sheets)

    def _emit(step: int, total: int, sheet: str, phase: str, detail: str = ""):
        """向UI发送进度事件（可选）。"""
        if progress_cb is None:
            return
        try:
            progress_cb({
                "step": step,
                "total": total,
                "sheet": sheet,
                "phase": phase,
                "detail": detail,
            })
        except Exception:
            # 进度回调不应影响主流程
            pass

    # ✅成本表（只有当 selected 里包含成本相关 sheet 才需要打开）
    cost_related = any(sh != "收入" for sh in selected_sheets)
    wb_cost = load_workbook(sources["cost_file"]) if cost_related else None

    # 原油吨桶比（只有用到时才需要，但这里读不到也不致命）
    ratio_map = read_ton_barrel_ratio_map(sources["expense"]) if sources.get("expense") else {}

    # ---------------- ✅仅当需要审查(balance_table)时才读取 ----------------
    need_balance = any(
        sh in {
            "2-1", "2-2", "3-1", "3-2", "4-1",
            "4-1(按收入表回填)", "4-1（按收入表回填）"
        }
        for sh in selected_sheets
    )

    balance_map_oil = balance_meta_oil = balance_sheet_name_oil = None
    product_balance_map = product_balance_meta = product_balance_sheet_name = None

    if need_balance:
        if not sources.get("balance_table"):
            raise ValueError("本次勾选的sheet包含审查项，但未提供：外购原(燃)料收拨存平衡表（balance_table）")

        # 原油审查平衡表（2-1/2-2）
        balance_map_oil, balance_meta_oil = read_balance_table(
            sources["balance_table"], sheet_name="原燃料-合并石化"
        )
        balance_sheet_name_oil = balance_meta_oil["sheet_name"]

        # 产品审查平衡表（3-1/3-2 + 4-1）
        product_balance_map, product_balance_meta = read_product_balance_table(
            sources["balance_table"], sheet_name="产品表-合并石化部"
        )
        product_balance_sheet_name = product_balance_meta["sheet_name"]

    # 2-3：上月成本表
    ws_prev_23 = None
    if sources.get("prev_cost_file"):
        wb_prev = load_workbook(sources["prev_cost_file"], data_only=True)
        if "2-3" not in wb_prev.sheetnames:
            raise ValueError("上月成本报表缺少sheet：2-3")
        ws_prev_23 = wb_prev["2-3"]

    # 2-3：提前读暂估表（USD/桶口径）
    intransit_price_pack = None
    if sources.get("intransit"):
        intransit_price_pack = read_intransit_usd_price_pack(sources["intransit"])

    # ✅销售收入 pack（给 4-1 & 收入表）
    sales_pack = None
    if sources.get("sales_revenue") and sources.get("current_month"):
        sales_pack = build_sales_revenue_pack(sources["sales_revenue"], int(sources["current_month"]))

    # ✅新增：产品当月/累计名称映射（用于 4-1 / 收入表 / 4-2 新增物料时补名称）
    product_month_name_map = None
    product_ytd_name_map = None

    # 只要用户提供了对应文件，就提前构建（避免后面多次读文件）
    try:
        if sources.get("product_month"):
            ppack_m = read_product_sheet1_pack(sources["product_month"], sheet_name="Sheet1")
            product_month_name_map = {
                _clean_code_for_map(code): (rec.get("name_raw") or "")
                for code, rec in (ppack_m.get("by_code") or {}).items()
                if _clean_code_for_map(code)
            }
    except Exception:
        product_month_name_map = None

    try:
        if sources.get("product_ytd"):
            ppack_y = read_product_sheet1_pack(sources["product_ytd"], sheet_name="Sheet1")
            product_ytd_name_map = {
                _clean_code_for_map(code): (rec.get("name_raw") or "")
                for code, rec in (ppack_y.get("by_code") or {}).items()
                if _clean_code_for_map(code)
            }
    except Exception:
        product_ytd_name_map = None

    all_issues = []
    all_adjust_records = []
    income_out_path = None

    total_steps = len(selected_sheets)
    step_idx = 0

    for sheet_name in selected_sheets:
        step_idx += 1
        _emit(step_idx, total_steps, sheet_name, phase="start")
        # ---------------- 收入表（非成本表） ----------------
        if sheet_name == "收入":
            if not sources.get("income_file"):
                raise ValueError("未提供：产销存-收入表（income_file）")
            if sales_pack is None:
                raise ValueError("未提供：累计销售收入表/当前月份，无法填充收入表")

            wb_income = load_workbook(sources["income_file"])
            if "收入" not in wb_income.sheetnames:
                raise ValueError("产销存-收入表缺少sheet：收入")

            ws_income = wb_income["收入"]

            new_codes = fill_income_ws(ws_income, sales_pack, product_ytd_name_map)
            sales_desc_name_map = {
                str(code).strip().lstrip("0"): (rec.get("desc") or "")
                for code, rec in (sales_pack.get("by_code") or {}).items()
            }
            # 同步新增物料到裸税价/裸价税/祼税价 sheet
            bare_sheet_name = guess_bare_tax_sheet_name(wb_income.sheetnames)
            ws_bare = wb_income[bare_sheet_name]
            ensure_new_codes_in_bare_tax_ws(
                ws_bare,
                new_codes,
                start_row=4,
                code_col=2,  # B
                name_col=3,  # C
                vat_col=4,  # D
                cons_col=5,  # E
                name_map=product_ytd_name_map,
            )

            income_out_path = _auto_income_out_path(cost_out_path)
            wb_income.save(income_out_path)
            _emit(step_idx, total_steps, sheet_name, phase="done")
            continue

        # ---------------- 以下为成本表内sheet ----------------
        if wb_cost is None:
            raise ValueError("成本报表未加载（cost_file）")

        # ✅虚拟sheet：4-1（按收入表回填） —— 实际操作成本表里的 "4-1"
        virtual_41 = {"4-1（按收入表回填）", "4-1(按收入表回填)"}
        real_sheet_name = "4-1" if sheet_name in virtual_41 else sheet_name

        if real_sheet_name not in wb_cost.sheetnames:
            raise ValueError(f"成本报表中不存在工作表：{real_sheet_name}（由选项 {sheet_name} 映射）")

        ws_cost = wb_cost[real_sheet_name]

        if sheet_name == "2-1":
            fx_rate = float(sources["fx_rate_month"])
            df_oil = _read_oil_df(sources["oil_month"])
            intransit_items = read_intransit_estimate(sources["intransit"], fx_rate)
            _pipeline(ws_cost, wb_cost, df_oil, fx_rate, ratio_map, intransit_items)

            if need_balance:
                issues = audit_sheet(ws_cost, sheet_name, balance_map_oil, balance_sheet_name_oil, tol=0.0)
                all_issues.extend(issues)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "2-2":
            fx_rate = float(sources["fx_rate_ytd"])
            df_oil = _read_oil_df(sources["oil_ytd"])
            intransit_items = read_intransit_estimate(sources["intransit"], fx_rate)
            _pipeline(ws_cost, wb_cost, df_oil, fx_rate, ratio_map, intransit_items)

            # 2-2 的美元/桶锚点 N/O/P：与 2-1 保持一致
            try:
                ws_21_ref = wb_cost["2-1"]
                row_usd_21 = find_row_by_a_contains(ws_21_ref, "美元/桶")
                row_usd_22 = find_row_by_a_contains(ws_cost, "美元/桶")
                if row_usd_21 and row_usd_22:
                    for col in (14, 15, 16):
                        ws_cost.cell(row_usd_22, col, value=ws_21_ref.cell(row_usd_21, col).value)
            except Exception:
                pass

            if need_balance:
                issues = audit_sheet(ws_cost, sheet_name, balance_map_oil, balance_sheet_name_oil, tol=0.0)
                all_issues.extend(issues)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "2-3":
            ws_23 = wb_cost["2-3"]
            ws_21 = wb_cost["2-1"]
            ws_22 = wb_cost["2-2"]

            fx_month = float(sources["fx_rate_month"])
            fx_ytd = float(sources["fx_rate_ytd"])

            pack = build_intransit_price_pack(sources["intransit"])
            fill_sheet_2_3(ws_23, ws_21, ws_22, ws_prev_23, ratio_map, fx_month, fx_ytd, int(sources.get('current_month') or 0), pack)
            backfill_usd_anchor_from_23(ws_21, ws_22, ws_23)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "3-1":
            ppack = read_product_sheet1_pack(sources["product_month"], sheet_name="Sheet1")
            insert_new_products_into_general_trade(ws_cost, ppack, start_row=6)
            adjust_records = fill_product_sheet_3x(ws_cost, ppack, start_row=6)
            all_adjust_records.extend(adjust_records)
            write_product_summaries(ws_cost, start_row=6)

            if need_balance:
                issues = audit_sheet_product(ws_cost, "3-1", product_balance_map, product_balance_sheet_name, tol=0.0)
                all_issues.extend(issues)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "3-2":
            ppack = read_product_sheet1_pack(sources["product_ytd"], sheet_name="Sheet1")
            insert_new_products_into_general_trade(ws_cost, ppack, start_row=6)
            adjust_records = fill_product_sheet_3x(ws_cost, ppack, start_row=6)
            all_adjust_records.extend(adjust_records)
            write_product_summaries(ws_cost, start_row=6)

            if need_balance:
                issues = audit_sheet_product(ws_cost, "3-2", product_balance_map, product_balance_sheet_name, tol=0.0)
                all_issues.extend(issues)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "4-1":
            if sales_pack is None:
                raise ValueError("未提供：累计销售收入表/当前月份，无法填充4-1")

            fill_cost_sheet_4_1(ws_cost, sales_pack, product_ytd_name_map)

            if need_balance:
                issues = audit_sheet_sales_41(ws_cost, "4-1", product_balance_map, product_balance_sheet_name, tol=0.0)
                all_issues.extend(issues)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name in {"4-1(按收入表回填)", "4-1（按收入表回填）"}:
            if not sources.get("income_file"):
                raise ValueError("未提供：产销存-收入表（income_file），无法按收入表回填4-1")

            wb_income = load_workbook(sources["income_file"], data_only=False)
            if "收入" not in wb_income.sheetnames:
                raise ValueError("产销存-收入表缺少sheet：收入")
            ws_income = wb_income["收入"]

            fill_cost_sheet_4_1_from_income(ws_cost, ws_income, int(sources["current_month"]))

            if need_balance:
                issues = audit_sheet_sales_41(ws_cost, "4-1", product_balance_map, product_balance_sheet_name, tol=0.0)
                all_issues.extend(issues)

            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "4-2":
            if not sources.get("income_file"):
                raise ValueError("未提供：产销存-收入表（income_file），无法填充4-2")

            cm = int(sources.get("current_month") or 0)
            if not (1 <= cm <= 12):
                raise ValueError("current_month 无效（应为1~12）")

            from fill_4_2 import fill_cost_sheet_4_2, _guess_bare_sheet

            wb_income = load_workbook(sources["income_file"], data_only=True)

            if "收入" not in wb_income.sheetnames:
                raise ValueError("产销存-收入表缺少sheet：收入")
            ws_income = wb_income["收入"]

            ws_bare_src = _guess_bare_sheet(wb_income)  # 自动找"裸价税/裸税价"
            fill_cost_sheet_4_2(ws_cost, ws_income, ws_bare_src, cm, product_ytd_name_map)
            _emit(step_idx, total_steps, sheet_name, phase="done")


        elif sheet_name == "6":

            if "6" not in wb_cost.sheetnames:
                raise ValueError("成本报表中不存在工作表：6")

            ws6 = wb_cost["6"]

            if not sources.get("profit_new"):
                raise ValueError("未提供：SAP利润表-利新（profit_new）")

            if not sources.get("profit_old"):
                raise ValueError("未提供：SAP利润表-利旧（profit_old）")

            old_all = read_sap_profit_table(sources["profit_old"], header_row=12)

            new_all = read_sap_profit_table(sources["profit_new"], header_row=12)

            exclude_items = sources.get("exclude_inv_items") or []

            if isinstance(exclude_items, str):
                exclude_items = [x.strip() for x in exclude_items.split("\n") if x.strip()]

                        # ✅6表吨油利润需要读取 2-1/2-2/1-1 的数量分母；若函数签名不支持则自动降级
            ws_21 = wb_cost["2-1"] if "2-1" in wb_cost.sheetnames else None
            ws_22 = wb_cost["2-2"] if "2-2" in wb_cost.sheetnames else None
            ws_11 = wb_cost["1-1"] if "1-1" in wb_cost.sheetnames else None

            import inspect as _inspect
            try:
                _sig = _inspect.signature(fill_sheet6_profit)
                if "ws_21" in _sig.parameters:
                    fill_sheet6_profit(
                        ws6, old_all, new_all,
                        ui_exclude_items=exclude_items,
                        ws_21=ws_21, ws_22=ws_22, ws_11=ws_11,
                    )
                else:
                    fill_sheet6_profit(ws6, old_all, new_all, ui_exclude_items=exclude_items)
            except Exception:
                # 兜底：保持原调用方式，避免中断主流程
                fill_sheet6_profit(ws6, old_all, new_all, ui_exclude_items=exclude_items)

            _emit(step_idx, total_steps, sheet_name, phase="done")
            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "7":
            if "7" not in wb_cost.sheetnames:
                raise ValueError("成本报表中不存在工作表：7（半成品库存表）")
            ws7 = wb_cost["7"]

            if not sources.get("semi_cur"):
                raise ValueError("未提供：SAP半成品库存-当月底表（semi_cur）")
            if not sources.get("semi_prev"):
                raise ValueError("未提供：SAP半成品库存-上月底表（semi_prev）")

            cur_map = read_semi_inventory_sap(sources["semi_cur"], header_row=1, data_start_row=2)
            prev_map = read_semi_inventory_sap(sources["semi_prev"], header_row=1, data_start_row=2)

            fill_cost_sheet_7_semi_inventory(ws7, cur_map, prev_map)
            _emit(step_idx, total_steps, sheet_name, phase="done")
            _emit(step_idx, total_steps, sheet_name, phase="done")


        elif sheet_name == "1-1":

            if "1-1" not in wb_cost.sheetnames:
                raise ValueError("成本报表中不存在工作表：1-1")

            # 依赖sheet必须存在

            for req_sh in ("2-1", "2-2", "3-1", "3-2"):

                if req_sh not in wb_cost.sheetnames:
                    raise ValueError(f"成本报表缺少依赖sheet：{req_sh}（用于1-1取数）")

            if not sources.get("expense"):
                raise ValueError("未提供：生产经营费用表（expense），1-1需要读取其‘导’和‘万元’sheet")

            if not sources.get("profit_new"):
                raise ValueError("未提供：SAP利润表-利新（profit_new），1-1需要取管理/财务/销售费用")

            if not sources.get("cost_elem_month"):
                raise ValueError("未提供：SAP成本要素表-当月（cost_elem_month）")

            if not sources.get("cost_elem_ytd"):
                raise ValueError("未提供：SAP成本要素表-累计（cost_elem_ytd）")

            # runner.py（替换原 512~518 附近这段）
            from cost_element_reader import read_cost_element_table

            # ✅ header_row=None：自动扫描找到“成本要素/实际成本”的标题行
            cost_elem_month_map = read_cost_element_table(sources["cost_elem_month"], header_row=None)
            cost_elem_ytd_map = read_cost_element_table(sources["cost_elem_ytd"], header_row=None)

            ws_11 = wb_cost["1-1"]

            ws_21 = wb_cost["2-1"]

            ws_22 = wb_cost["2-2"]

            ws_31 = wb_cost["3-1"]

            ws_32 = wb_cost["3-2"]
            ws_6 = wb_cost["6"]
            ws_7 = wb_cost["7"] if "7" in wb_cost.sheetnames else None

            # ✅利新读取成 table

            profit_new = read_sap_profit_table(sources["profit_new"], header_row=12)

            manual_table = sources.get("manual_1_1_table")

            fill_cost_sheet_1_1(

                ws_11,

                ws_21=ws_21,

                ws_22=ws_22,

                ws_31=ws_31,

                ws_32=ws_32,
                ws_6=ws_6,  # ✅一定要加

                ws_7=ws_7,

                expense_file=sources["expense"],

                # ✅传入 map（推荐方式）

                cost_elem_month_map=cost_elem_month_map,

                cost_elem_ytd_map=cost_elem_ytd_map,

                profit_new_table=profit_new,

                manual_1_1_table=manual_table,

                dao_sheet_name="导",

                sheet_wanyuan_name="万元",

            )
            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name == "1-2":
            if "1-2" not in wb_cost.sheetnames:
                raise ValueError("成本报表中不存在工作表：1-2")
            if "1-1" not in wb_cost.sheetnames:
                raise ValueError("成本报表中不存在工作表：1-1（1-2依赖它）")

            ws_12 = wb_cost["1-2"]
            ws_11 = wb_cost["1-1"]

            from fill_1_2 import fill_cost_sheet_1_2
            fill_cost_sheet_1_2(ws_12, ws_11)
            _emit(step_idx, total_steps, sheet_name, phase="done")

        elif sheet_name in {"A-1", "A-2"}:
            for req in ("A-1", "A-2", "2-1", "2-2", "3-1", "3-2", "4-1", "1-2", "6", "7"):
                if req not in wb_cost.sheetnames:
                    raise ValueError(f"成本报表缺少依赖sheet：{req}（用于A-1/A-2）")

            ws_a1 = wb_cost["A-1"]
            ws_a2 = wb_cost["A-2"]
            ws_21 = wb_cost["2-1"]
            ws_22 = wb_cost["2-2"]
            ws_31 = wb_cost["3-1"]
            ws_32 = wb_cost["3-2"]
            ws_41 = wb_cost["4-1"]
            ws_12 = wb_cost["1-2"]
            ws_6 = wb_cost["6"]
            ws_7 = wb_cost["7"]

            from fill_a_1_a_2 import fill_cost_sheet_a_1_a_2
            fill_cost_sheet_a_1_a_2(
                ws_a1, ws_a2,
                ws_21, ws_22,
                ws_31, ws_32,
                ws_41, ws_12,
                ws_6, ws_7,
            )

            _emit(step_idx, total_steps, sheet_name, phase="done")


        else:
            raise ValueError(f"尚未实现该sheet的runner：{sheet_name}")

    audit_path = audit_out_path or _auto_audit_path(cost_out_path)
    _emit(total_steps, total_steps, "审查日志", phase="start", detail="生成审查日志")
    save_audit_excel(all_issues, audit_path, adjust_records=all_adjust_records)
    _emit(total_steps, total_steps, "审查日志", phase="done")

    return wb_cost, audit_path, income_out_path