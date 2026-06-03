# fill_4_1.py
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


def _clean_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0").strip()


def _set(ws: Worksheet, r: int, c: int, v):
    ws.cell(r, c, value="-" if v is None else v)


def _to_float_or_none(v):
    """把 None/''/'-' 等视为 None，其它尽量转 float。"""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
    try:
        return float(v)
    except Exception:
        return None


def _unit_price(amt_yuan: float | None, qty_ton: float | None):
    if amt_yuan is None or qty_ton is None or qty_ton == 0:
        return None
    return amt_yuan / qty_ton


def _neg_pair(q: float | None, a: float | None):
    if q is None and a is None:
        return (None, None)
    q2 = None if q is None else -float(q)
    a2 = None if a is None else -float(a)
    return (q2, a2)


def _normalize_sales_pack(pack: dict) -> dict:
    """
    统一清洗 by_code 的 key（去前导0、去 .0、strip），并把 monthly 汇总到同一 code。
    用于解决：模板里是 83005952，但 pack 里可能是 '083005952'/'83005952.0'/int 等。
    """
    by_code = pack.get("by_code") or {}
    if not by_code:
        return pack

    new_by: dict[str, dict] = {}
    for raw_code, rec in by_code.items():
        if not rec:
            continue
        code = _clean_code(raw_code)
        if not code:
            continue

        dst = new_by.get(code)
        if dst is None:
            dst = {"desc": rec.get("desc", ""), "monthly": {}}
            new_by[code] = dst

        if not dst.get("desc") and rec.get("desc"):
            dst["desc"] = rec.get("desc", "")

        m_dst = dst.setdefault("monthly", {})
        m_src = rec.get("monthly") or {}
        for ym, mm in m_src.items():
            if not mm:
                continue
            q = _to_float_or_none(mm.get("qty_ton"))
            a = _to_float_or_none(mm.get("amt_yuan"))

            cur = m_dst.get(ym)
            if cur is None:
                m_dst[ym] = {"qty_ton": q, "amt_yuan": a}
            else:
                cq = _to_float_or_none(cur.get("qty_ton"))
                ca = _to_float_or_none(cur.get("amt_yuan"))
                cur["qty_ton"] = (cq or 0.0) + (q or 0.0) if (cq is not None or q is not None) else None
                cur["amt_yuan"] = (ca or 0.0) + (a or 0.0) if (ca is not None or a is not None) else None

    new_pack = dict(pack)
    new_pack["by_code"] = new_by
    return new_pack


def _get_month(pack: dict, code: str, y: int, m: int):
    rec = (pack.get("by_code") or {}).get(code)
    if not rec:
        return (None, None)
    mm = rec.get("monthly", {}).get((y, m))
    if not mm:
        return (None, None)

    q = _to_float_or_none(mm.get("qty_ton"))
    a = _to_float_or_none(mm.get("amt_yuan"))
    if q is None and a is None:
        return (None, None)
    return (q, a)


def _get_ytd(pack: dict, code: str, y: int, month_to: int):
    qty = 0.0
    amt = 0.0
    has = False
    for m in range(1, month_to + 1):
        q, a = _get_month(pack, code, y, m)
        if q is None and a is None:
            continue
        has = True
        qty += (q or 0.0)
        amt += (a or 0.0)
    return (qty, amt) if has else (None, None)


def _build_row_map(ws: Worksheet, start_row: int) -> dict[str, int]:
    m = {}
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)  # B
        if code:
            m[code] = r
    return m


def _find_last_material_row(ws: Worksheet, start_row: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)
        if code:
            last = r
    return last


def _renumber_seq(ws: Worksheet, start_row: int):
    seq = 1
    for r in range(start_row, ws.max_row + 1):
        code = _clean_code(ws.cell(r, 2).value)
        if not code:
            continue
        ws.cell(r, 1, value=seq)  # A
        seq += 1


def _prev_month_of(year: int, month: int, prev_year: int | None):
    """严格定义“上月”= 当前月-1；若当前月=1，则上月=上年12月。"""
    if month == 1:
        py = prev_year if prev_year is not None else year - 1
        return py, 12
    return year, month - 1


def _has_price(q_raw: float | None, a_raw: float | None) -> bool:
    """
    是否“有价格”：
    - qty 和 amt 都不为空
    - qty != 0（取 abs 判断）
    """
    if q_raw is None or a_raw is None:
        return False
    return abs(float(q_raw)) != 0.0


def _calc_price_abs(q_raw: float | None, a_raw: float | None) -> float:
    """按比较口径算单价（qty/amt 取 abs），调用前确保 _has_price 为 True。"""
    q = abs(float(q_raw))
    a = abs(float(a_raw))
    p = _unit_price(a, q)
    return float(p) if p is not None else 0.0


def fill_cost_sheet_4_1(
    ws_41: Worksheet,
    sales_pack: dict,
    product_month_name_map: dict[str, str] | None = None,
):
    """
    4-1 成本报表口径：
    - D/E/F/G/H/I：写正数（abs）
    - J：量差（万吨口径）：保留（缺 qty 当 0）
    - K：价差：只有当月&上月都“有价格”才做差，否则=0
    - L：增利 = 价差 * 本月量（万吨口径，abs）；若 K=0 则 L=0
    - 合计行：模板第5行；明细从第6行开始
    - 合计行 L：逐物料增利求和
    - 合计行 K：合计行当月价 - 合计行上月价（缺价则0）
    """
    sales_pack = _normalize_sales_pack(sales_pack)

    year = sales_pack.get("year")
    prev_year = sales_pack.get("prev_year")
    cm = int(sales_pack.get("current_month") or 0)
    if not year or not (1 <= cm <= 12):
        return

    by_code = sales_pack.get("by_code") or {}
    if not by_code:
        return

    total_row = 5
    start_row = total_row + 1

    row_map = _build_row_map(ws_41, start_row=start_row)
    existed_codes = set(row_map.keys())
    all_codes = sorted(by_code.keys())

    def _name_of(code: str) -> str:
        """新增物料名称：优先用产品当月数据，兜底用 sales_pack 的 desc。"""
        if product_month_name_map:
            n = product_month_name_map.get(code)
            if n:
                return n
        return by_code.get(code, {}).get("desc", "")

    # 插入新增物料
    new_codes = [c for c in all_codes if c not in existed_codes]
    if new_codes:
        last_row = _find_last_material_row(ws_41, start_row=start_row)
        ptr = last_row + 1 if last_row >= start_row else start_row
        for code in new_codes:
            ws_41.insert_rows(ptr)
            ws_41.cell(ptr, 2, value=code)  # B
            ws_41.cell(ptr, 3, value=_name_of(code))  # C
            ptr += 1
        row_map = _build_row_map(ws_41, start_row=start_row)

    # 若模板已存在物料但名称为空，也用产品当月数据补齐一次
    if product_month_name_map:
        for code, r in row_map.items():
            cur = ws_41.cell(r, 3).value
            if cur is None or str(cur).strip() == "":
                n = product_month_name_map.get(code)
                if n:
                    ws_41.cell(r, 3, value=n)

    COL = {
        "m_qty_wan": 4, "m_price": 5, "m_amt_wan": 6,       # D/E/F
        "y_qty_wan": 7, "y_price": 8, "y_amt_wan": 9,       # G/H/I
        "mom_qty_diff": 10, "mom_price_diff": 11, "profit": 12,  # J/K/L
        "end_col": 12,
    }

    py, pm = _prev_month_of(year, cm, prev_year)

    profit_sum = 0.0

    # -------- 明细填充 --------
    for code, r in row_map.items():
        for c in range(4, COL["end_col"] + 1):
            ws_41.cell(r, c, value="-")

        # 当月 raw
        q_m_raw, a_m_raw = _get_month(sales_pack, code, year, cm)

        # 当月 display（D/E/F）
        q_m_disp, a_m_disp = _neg_pair(q_m_raw, a_m_raw)
        if q_m_disp is not None:
            _set(ws_41, r, COL["m_qty_wan"], q_m_disp / 10000.0)
            _set(ws_41, r, COL["m_amt_wan"], a_m_disp / 10000.0)
            _set(ws_41, r, COL["m_price"], _unit_price(a_m_disp, q_m_disp))

        # 累计 raw
        q_y_raw, a_y_raw = _get_ytd(sales_pack, code, year, cm)

        # 累计 display（G/H/I）
        q_y_disp, a_y_disp = _neg_pair(q_y_raw, a_y_raw)
        if q_y_disp is not None:
            _set(ws_41, r, COL["y_qty_wan"], q_y_disp / 10000.0)
            _set(ws_41, r, COL["y_amt_wan"], a_y_disp / 10000.0)
            _set(ws_41, r, COL["y_price"], _unit_price(a_y_disp, q_y_disp))

        # 上月 raw
        q_p_raw, a_p_raw = _get_month(sales_pack, code, py, pm)

        # 量差(J)：保留（缺 qty 当0），单位万吨
        q_m_cmp = abs(float(q_m_raw)) if q_m_raw is not None else 0.0
        q_p_cmp = abs(float(q_p_raw)) if q_p_raw is not None else 0.0
        _set(ws_41, r, COL["mom_qty_diff"], (q_m_cmp - q_p_cmp) / 10000.0)

        # 价差(K)：缺任一月单价信息 => 0
        has_price_m = _has_price(q_m_raw, a_m_raw)
        has_price_p = _has_price(q_p_raw, a_p_raw)

        if (not has_price_m) or (not has_price_p):
            price_diff = 0.0
            _set(ws_41, r, COL["mom_price_diff"], 0.0)
            profit_wan = 0.0
            _set(ws_41, r, COL["profit"], 0.0)
        else:
            p_m_cmp = _calc_price_abs(q_m_raw, a_m_raw)
            p_p_cmp = _calc_price_abs(q_p_raw, a_p_raw)
            price_diff = p_m_cmp - p_p_cmp
            _set(ws_41, r, COL["mom_price_diff"], price_diff)

            profit_wan = price_diff * (q_m_cmp / 10000.0)
            _set(ws_41, r, COL["profit"], profit_wan)

        profit_sum += profit_wan

    # -------- 合计行（第5行）--------
    for c in range(4, COL["end_col"] + 1):
        ws_41.cell(total_row, c, value="-")

    def sum_codes_month(y: int, m: int):
        tq = 0.0
        ta = 0.0
        has = False
        for code in all_codes:
            q, a = _get_month(sales_pack, code, y, m)
            if q is None and a is None:
                continue
            has = True
            tq += (q or 0.0)
            ta += (a or 0.0)
        return tq, ta, has

    # 当月合计 display（D/E/F）
    tq_m_raw, ta_m_raw, has_m = sum_codes_month(year, cm)
    if has_m:
        tq_m_disp, ta_m_disp = _neg_pair(tq_m_raw, ta_m_raw)
        _set(ws_41, total_row, COL["m_qty_wan"], tq_m_disp / 10000.0)
        _set(ws_41, total_row, COL["m_amt_wan"], ta_m_disp / 10000.0)
        _set(ws_41, total_row, COL["m_price"], _unit_price(ta_m_disp, tq_m_disp))

    # 累计合计 display（G/H/I）
    tqy_raw = 0.0
    tay_raw = 0.0
    hasy = False
    for mm in range(1, cm + 1):
        q, a, has = sum_codes_month(year, mm)
        if not has:
            continue
        hasy = True
        tqy_raw += q
        tay_raw += a
    if hasy:
        tqy_disp, tay_disp = _neg_pair(tqy_raw, tay_raw)
        _set(ws_41, total_row, COL["y_qty_wan"], tqy_disp / 10000.0)
        _set(ws_41, total_row, COL["y_amt_wan"], tay_disp / 10000.0)
        _set(ws_41, total_row, COL["y_price"], _unit_price(tay_disp, tqy_disp))

    # 上月合计 raw
    tq_p_raw, ta_p_raw, has_p = sum_codes_month(py, pm)

    # 合计量差(J)：保留（缺 qty 当0），单位万吨
    tq_m_cmp = abs(float(tq_m_raw)) if has_m and tq_m_raw is not None else 0.0
    tq_p_cmp = abs(float(tq_p_raw)) if has_p and tq_p_raw is not None else 0.0
    _set(ws_41, total_row, COL["mom_qty_diff"], (tq_m_cmp - tq_p_cmp) / 10000.0)

    # 合计价差(K)：缺任一月单价信息 => 0
    has_price_m_tot = has_m and _has_price(tq_m_raw, ta_m_raw)
    has_price_p_tot = has_p and _has_price(tq_p_raw, ta_p_raw)

    if (not has_price_m_tot) or (not has_price_p_tot):
        _set(ws_41, total_row, COL["mom_price_diff"], 0.0)
    else:
        p_m_tot = _calc_price_abs(tq_m_raw, ta_m_raw)
        p_p_tot = _calc_price_abs(tq_p_raw, ta_p_raw)
        pdiff_tot = p_m_tot - p_p_tot
        _set(ws_41, total_row, COL["mom_price_diff"], pdiff_tot)

    # 合计增利 L：逐物料增利求和
    _set(ws_41, total_row, COL["profit"], profit_sum)

    _renumber_seq(ws_41, start_row=start_row)
