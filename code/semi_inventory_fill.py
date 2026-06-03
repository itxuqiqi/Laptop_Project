# semi_inventory_fill.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Tuple, Optional, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, PatternFill, Protection
from copy import copy


@dataclass
class InvItem:
    mat: str                 # 物料号
    desc: str                # 物料描述
    qty_ton: float           # 库存量（吨）
    val_yuan: float          # 库存价值（元）

    @property
    def qty_wanton(self) -> float:
        # 吨 -> 万吨
        return self.qty_ton / 10000.0

    @property
    def val_wanyuan(self) -> float:
        # 元 -> 万元
        return self.val_yuan / 10000.0

    @property
    def unit_price_yuan_per_ton(self) -> Optional[float]:
        # 元/吨（保持“元”）
        if self.qty_ton == 0:
            return 0.0
        return self.val_yuan / self.qty_ton


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "" or s == "-":
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def read_semi_inventory_sap(
    file_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    data_start_row: int = 2,
    col_mat: int = 1,   # A
    col_desc: int = 3,  # C
    col_val: int = 4,   # D 总库存价值(元)
    col_qty: int = 6,   # F 总库存量(吨)
) -> Dict[str, InvItem]:
    """
    读取 SAP 半成品库存底表（当月/上月通用）。
    若同一物料号多行，按物料号聚合：qty/val 求和，desc 取第一次非空。
    返回: mat -> InvItem
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    out: Dict[str, InvItem] = {}

    max_row = ws.max_row
    for r in range(data_start_row, max_row + 1):
        mat = ws.cell(r, col_mat).value
        if mat is None or str(mat).strip() == "":
            continue
        mat = str(mat).strip()

        desc = ws.cell(r, col_desc).value
        desc = "" if desc is None else str(desc).strip()

        val = _to_float(ws.cell(r, col_val).value)
        qty = _to_float(ws.cell(r, col_qty).value)

        if mat not in out:
            out[mat] = InvItem(mat=mat, desc=desc, qty_ton=qty, val_yuan=val)
        else:
            # 聚合
            old = out[mat]
            if (not old.desc) and desc:
                old.desc = desc
            old.qty_ton += qty
            old.val_yuan += val

    return out


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 12):
    """
    复制一行的样式（字体/边框/填充/对齐/数字格式/保护），用于新增物料行保持模板样式。
    """
    for c in range(1, max_col + 1):
        src_cell = ws.cell(src_row, c)
        dst_cell = ws.cell(dst_row, c)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection) if src_cell.protection else Protection()
        dst_cell.alignment = copy(src_cell.alignment) if src_cell.alignment else Alignment()
        dst_cell.font = copy(src_cell.font) if src_cell.font else Font()
        dst_cell.border = copy(src_cell.border) if src_cell.border else Border()
        dst_cell.fill = copy(src_cell.fill) if src_cell.fill else PatternFill()


def _find_last_material_row(ws: Worksheet, start_row: int = 6, key_col: int = 2) -> int:
    """
    从 start_row 开始，找 B 列(物料号)最后一个非空行。
    """
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, key_col).value
        if v is None or str(v).strip() == "":
            continue
        last = r
    return last


def _build_sheet7_index(ws: Worksheet, start_row: int = 6) -> Dict[str, int]:
    """
    建立 sheet7 物料号 -> 行号索引（B列）。
    """
    idx: Dict[str, int] = {}
    for r in range(start_row, ws.max_row + 1):
        mat = ws.cell(r, 2).value  # B
        if mat is None:
            continue
        mat = str(mat).strip()
        if not mat:
            continue
        if mat not in idx:
            idx[mat] = r
    return idx


def _write_item_to_row(
    ws: Worksheet,
    row: int,
    seq: int,
    mat: str,
    desc: str,
    cur: Optional[InvItem],
    prev: Optional[InvItem],
):
    """
    写入一行：A序号 B物料号 C名称
    D/E/F 当月数量(万吨)/单价(元)/金额(万元)
    G/H/I 上月数量(万吨)/单价(元)/金额(万元)
    J/K/L 差异=当月-上月
    """
    ws.cell(row, 1).value = seq
    ws.cell(row, 2).value = mat
    ws.cell(row, 3).value = desc

    def q(cur_item: Optional[InvItem]) -> float:
        return 0.0 if cur_item is None else cur_item.qty_wanton

    def v(cur_item: Optional[InvItem]) -> float:
        return 0.0 if cur_item is None else cur_item.val_wanyuan

    def p(cur_item: Optional[InvItem]) -> float:
        return 0.0 if cur_item is None else (cur_item.unit_price_yuan_per_ton or 0.0)

    cur_qty = q(cur)
    cur_val = v(cur)
    cur_price = p(cur)

    prev_qty = q(prev)
    prev_val = v(prev)
    prev_price = p(prev)

    # 当月 D/E/F
    ws.cell(row, 4).value = cur_qty
    ws.cell(row, 5).value = cur_price
    ws.cell(row, 6).value = cur_val

    # 上月 G/H/I
    ws.cell(row, 7).value = prev_qty
    ws.cell(row, 8).value = prev_price
    ws.cell(row, 9).value = prev_val

    # 差异 J/K/L
    ws.cell(row, 10).value = cur_qty - prev_qty
    # 单价差异：按口径统一为数值；当两侧都为0时结果为0
    ws.cell(row, 11).value = (cur_price or 0.0) - (prev_price or 0.0)
    ws.cell(row, 12).value = cur_val - prev_val


def _recalc_total_row(ws: Worksheet, total_row: int = 5, start_row: int = 6):
    """
    合计行：
    - 数量列：D/G/J 求和
    - 金额列：F/I/L 求和
    - 单价列：
        E = 合计F / 合计D
        H = 合计I / 合计G
        K = E - H   （✅当月单价-上月单价）
    """
    last = _find_last_material_row(ws, start_row=start_row, key_col=2)
    if last < start_row:
        return

    def sum_col(col: int) -> float:
        s = 0.0
        for r in range(start_row, last + 1):
            s += _to_float(ws.cell(r, col).value)
        return s

    # 明细求和
    sum_d = sum_col(4)   # 当月数量
    sum_f = sum_col(6)   # 当月金额
    sum_g = sum_col(7)   # 上月数量
    sum_i = sum_col(9)   # 上月金额

    # 差异合计：更稳用“合计差”，避免明细单价为空时影响K（K用E-H）
    sum_j = sum_d - sum_g
    sum_l = sum_f - sum_i

    ws.cell(total_row, 4).value = sum_d
    ws.cell(total_row, 6).value = sum_f
    ws.cell(total_row, 7).value = sum_g
    ws.cell(total_row, 9).value = sum_i
    ws.cell(total_row, 10).value = sum_j
    ws.cell(total_row, 12).value = sum_l

    # 合计单价
    e = (sum_f / sum_d) if sum_d != 0 else 0.0
    h = (sum_i / sum_g) if sum_g != 0 else 0.0

    ws.cell(total_row, 5).value = e
    ws.cell(total_row, 8).value = h

    # ✅K：当月单价 - 上月单价
    if e is None and h is None:
        ws.cell(total_row, 11).value = None
    else:
        ws.cell(total_row, 11).value = (e or 0.0) - (h or 0.0)



def fill_cost_sheet_7_semi_inventory(
    ws7: Worksheet,
    cur_map: Dict[str, InvItem],
    prev_map: Dict[str, InvItem],
):
    """
    主填充函数：填成本报表 sheet=7 半成品库存表
    - 第5行合计
    - 第6行起明细
    - B列物料号为主键；新增物料追加到最后
    """
    start_row = 6
    total_row = 5

    idx = _build_sheet7_index(ws7, start_row=start_row)

    # 所有物料：当月 ∪ 上月（以便即使当月没了也能显示；你若不想显示“仅上月存在”，可改成只用当月）
    all_mats = sorted(set(cur_map.keys()) | set(prev_map.keys()))

    # 当前表最后一行物料
    last_row = _find_last_material_row(ws7, start_row=start_row, key_col=2)
    # 用于复制样式的模板行：优先用 last_row；如果表为空，用 start_row
    template_row = last_row if last_row >= start_row else start_row

    # 序号从现有表最大序号+1 或从1开始
    max_seq = 0
    for r in range(start_row, ws7.max_row + 1):
        v = ws7.cell(r, 1).value
        if isinstance(v, (int, float)):
            max_seq = max(max_seq, int(v))

    seq_next = max_seq + 1 if max_seq > 0 else 1

    for mat in all_mats:
        cur = cur_map.get(mat)
        prev = prev_map.get(mat)
        desc = (cur.desc if cur and cur.desc else (prev.desc if prev else ""))

        if mat in idx:
            row = idx[mat]
            # 序号：若已有则保留，否则补
            existing_seq = ws7.cell(row, 1).value
            if not isinstance(existing_seq, (int, float)):
                ws7.cell(row, 1).value = seq_next
                seq_next += 1
            seq = int(ws7.cell(row, 1).value)
            _write_item_to_row(ws7, row, seq, mat, desc, cur, prev)
        else:
            # 新增物料：追加到最后
            last_row = _find_last_material_row(ws7, start_row=start_row, key_col=2)
            new_row = (last_row + 1) if last_row >= start_row else start_row

            ws7.insert_rows(new_row, 1)
            _copy_row_style(ws7, template_row, new_row, max_col=12)

            _write_item_to_row(ws7, new_row, seq_next, mat, desc, cur, prev)
            idx[mat] = new_row
            seq_next += 1

    # 最后重算合计行
    _recalc_total_row(ws7, total_row=total_row, start_row=start_row)
