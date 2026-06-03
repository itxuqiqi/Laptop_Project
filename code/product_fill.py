# product_fill.py
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


def to_float(x):
    try:
        if x is None or x == "":
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "" or s == "-":
            return None
        return float(s.replace(",", ""))
    except Exception:
        return None


def fill_product_sheet_3x(ws, product_pack: dict, start_row=6):
    """
    3-1 / 3-2 填充：
    - 匹配：物料号优先；无号按名称
    - 数量/金额 ÷10000 写“万”；单价不换算
    - T-V（期末账面）= Q-S（期末库存）
    - 其他减少调整：
        * 若“其他减少”数量/金额非0 且 名称不含“自用”：
            - 将“其他减少”的数量/金额计入“本期生产”（保持原口径：生产重算单价）
            - 【新需求】写表展示：将“其他减少”清零（NOP=0）
            - 但审查日志保持“旧逻辑”的翻符号值（用于审查勾稽）
        * 若名称包含“自用”：
            - 不参与“挪到生产”的计算（保持原逻辑）
            - 且仍将“其他减少”的数量/金额改写为相反数（保持原逻辑）
    """
    adjust_records = []

    COL = {
        "begin_qty": 5, "begin_price": 6, "begin_amt": 7,
        "prod_qty": 8,  "prod_price": 9,  "prod_amt": 10,
        "sales_qty": 11, "sales_price": 12, "sales_amt": 13,
        "other_qty": 14, "other_price": 15, "other_amt": 16,
        "end_qty": 17,  "end_price": 18,  "end_amt": 19,
        "book_qty": 20, "book_price": 21, "book_amt": 22,
    }

    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(r, 4).value
        if name is None or str(name).strip() == "":
            continue

        code = ws.cell(r, 3).value
        code_s = "" if code is None else str(code).strip()
        name_raw = str(name).strip()

        # ====== 取底表数据（优先物料号，否则按名称）======
        rec = None
        if code_s and code_s in product_pack.get("by_code", {}):
            rec = product_pack["by_code"][code_s]
        else:
            rec = product_pack.get("by_name", {}).get(name_raw)

        if rec is None:
            # 即便没匹配到，也要把账面=库存做一下
            ws.cell(r, COL["book_qty"], value=ws.cell(r, COL["end_qty"]).value)
            ws.cell(r, COL["book_price"], value=ws.cell(r, COL["end_price"]).value)
            ws.cell(r, COL["book_amt"], value=ws.cell(r, COL["end_amt"]).value)
            continue

        # ====== 写 begin / prod / sales / other / end ======
        def w_qty_amt(q_key, a_key, q_col, p_col, a_col):
            q = rec.get(q_key)
            a = rec.get(a_key)
            q = 0.0 if q is None else float(q) / 10000.0
            a = 0.0 if a is None else float(a) / 10000.0
            p = 0.0 if q == 0 else a / q
            ws.cell(r, q_col, value=q)
            ws.cell(r, a_col, value=a)
            ws.cell(r, p_col, value=p)

        w_qty_amt("begin_qty", "begin_amt", COL["begin_qty"], COL["begin_price"], COL["begin_amt"])
        w_qty_amt("prod_qty", "prod_amt", COL["prod_qty"], COL["prod_price"], COL["prod_amt"])
        w_qty_amt("sales_qty", "sales_amt", COL["sales_qty"], COL["sales_price"], COL["sales_amt"])
        w_qty_amt("other_qty", "other_amt", COL["other_qty"], COL["other_price"], COL["other_amt"])
        w_qty_amt("end_qty", "end_amt", COL["end_qty"], COL["end_price"], COL["end_amt"])

        # ====== 其他减少调整 ======
        oq0 = to_float(ws.cell(r, COL["other_qty"]).value) or 0.0
        oa0 = to_float(ws.cell(r, COL["other_amt"]).value) or 0.0
        has_other = (abs(oq0) > 0.0) or (abs(oa0) > 0.0)

        if has_other:
            pq0 = to_float(ws.cell(r, COL["prod_qty"]).value) or 0.0
            pa0 = to_float(ws.cell(r, COL["prod_amt"]).value) or 0.0

            # 1) 非自用：挪到生产（数量/金额相加）并重算生产单价
            if "自用" not in name_raw:
                pq1 = pq0 + oq0
                pa1 = pa0 + oa0
                ws.cell(r, COL["prod_qty"], value=pq1)
                ws.cell(r, COL["prod_amt"], value=pa1)
                ws.cell(r, COL["prod_price"], value=0.0 if pq1 == 0 else pa1 / pq1)
            else:
                pq1 = pq0
                pa1 = pa0

            # 2) 其他减少：
            #    - 非自用：写表清零（展示口径），但审查日志仍保留“旧逻辑”的翻符号值
            #    - 自用：维持原逻辑（翻符号展示）
            if "自用" not in name_raw:
                # 旧逻辑（用于审查）：翻符号
                oq_audit = -oq0
                oa_audit = -oa0

                # 新逻辑（写表展示）：清零
                oq1 = 0.0
                oa1 = 0.0
                ws.cell(r, COL["other_qty"], value=oq1)
                ws.cell(r, COL["other_amt"], value=oa1)
                ws.cell(r, COL["other_price"], value=0.0)
            else:
                # 自用：保持原逻辑（翻符号写表）
                oq1 = -oq0
                oa1 = -oa0
                oq_audit = oq1
                oa_audit = oa1
                ws.cell(r, COL["other_qty"], value=oq1)
                ws.cell(r, COL["other_amt"], value=oa1)
                ws.cell(r, COL["other_price"], value=0.0 if oq1 == 0 else oa1 / oq1)

            # 3) 校验记录
            # 真实审查口径：比较“调整其他减少到生产之前”的数据是否平衡
            # - 非自用：其他减少虽会并到生产并在表上清零，但校验必须回到调整前
            #            => 期初 + 生产(调整前) - 销售 + 其他减少(未清零前) = 期末
            # - 自用：不做“挪到生产”，同样按原始口径校验
            #         => 期初 + 生产(调整前) - 销售 + 其他减少(未清零前) = 期末
            bq_chk = to_float(ws.cell(r, COL["begin_qty"]).value) or 0.0
            sq_chk = to_float(ws.cell(r, COL["sales_qty"]).value) or 0.0
            eq_chk = to_float(ws.cell(r, COL["end_qty"]).value) or 0.0

            lhs_before_adjust = bq_chk + pq0 - sq_chk + oq0
            ok_before_adjust = (round(lhs_before_adjust, 7) == round(eq_chk, 7))

            # 兼容保留：写表展示后的口径，便于排查
            pq_chk = to_float(ws.cell(r, COL["prod_qty"]).value) or 0.0
            oq_sheet = to_float(ws.cell(r, COL["other_qty"]).value) or 0.0
            lhs_sheet_after = bq_chk + pq_chk - sq_chk + oq_sheet
            ok_sheet_after = (round(lhs_sheet_after, 7) == round(eq_chk, 7))

            adjust_records.append({
                "sheet": ws.title,
                "row": r,
                "category": "" if ws.cell(r, 1).value is None else str(ws.cell(r, 1).value).strip(),
                "code": code_s,
                "name": name_raw,
                "is_self_use": ("自用" in name_raw),

                "begin_qty": bq_chk,
                "sales_qty": sq_chk,
                "prod_qty_before": pq0,
                "prod_amt_before": pa0,
                "other_qty_before": oq0,
                "other_amt_before": oa0,

                "prod_qty_after": pq1,
                "prod_amt_after": pa1,

                # 写表后的“展示口径”
                "other_qty_after": oq1,
                "other_amt_after": oa1,
                "other_qty_after_sheet": oq1,
                "other_amt_after_sheet": oa1,

                # 审查日志保留旧字段，便于兼容历史代码
                "other_qty_after_audit": oq_audit,
                "other_amt_after_audit": oa_audit,

                # 核心审查口径：调整前
                "check(before_adjust: begin+prod_before-sales+other_before=end)": "✅" if ok_before_adjust else "❌",
                "lhs(before_adjust: begin+prod_before-sales+other_before)": lhs_before_adjust,

                # 辅助排查口径：写表后
                "check(sheet_after: begin+prod_after-sales+other_after=end)": "✅" if ok_sheet_after else "❌",
                "lhs(sheet_after: begin+prod_after-sales+other_after)": lhs_sheet_after,

                # 兼容旧命名，供 audit_check 回退使用
                "check(begin+prod-sales=end)": "✅" if ok_sheet_after else "❌",
                "lhs(begin+prod-sales)": (bq_chk + pq_chk - sq_chk),

                "end_qty": eq_chk,
            })

        # 期末账面 = 期末库存
        ws.cell(r, COL["book_qty"], value=ws.cell(r, COL["end_qty"]).value)
        ws.cell(r, COL["book_price"], value=ws.cell(r, COL["end_price"]).value)
        ws.cell(r, COL["book_amt"], value=ws.cell(r, COL["end_amt"]).value)

    return adjust_records