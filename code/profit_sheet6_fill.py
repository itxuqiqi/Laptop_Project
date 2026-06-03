# profit_sheet6_fill.py
from typing import Dict, List, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet

from profit_utils import normalize_item_name, negate, sub, to_wanyuan


# --------------------------
# 1) 利旧：只这6个项目用利旧取数
#    - 仅前三个“收入类”翻符号（取相反数）
#    - 后三个“成本类”不翻符号
# --------------------------
OLD_6_ITEMS = [
    "其中:主营业务收入",
    "其中:内部收入",
    "其他业务收入",
    "其中：主营业务成本",
    "其中：内部成本",
    "其他业务成本",
]

OLD_FLIP_SIGN_ITEMS = [
    "其中:主营业务收入",
    "其中:内部收入",
    "其他业务收入",
]

# --------------------------
# 2) 利新：固定清单翻符号（取相反数），含联营（翻后才是真值）
# --------------------------
LIXIN_FLIP_SIGN_ITEMS = [
    "*一、营业总收入",
    "其中：营业收入",
    "※内部收入",
    "△利息收入",
    "△已赚保费",
    "△手续费及佣金收入",
    "加：其他收益",
    "投资收益（损失以“-”号填列）",
    "其中：对联营企业和合营企业",  # ✅也翻符号（翻后才是真值）
    "资产减值损失（损失以“-”号填列）",
    "资产处置收益（损失以“-”号填列）",
    "***三、营业利润（损失以“-”号填列）",
    "加：营业外收入",
    "****四、利润总额（亏损总额以“-”号填列）",
    "*****五、净利润（净亏损以“-”号填列）",
    "******1.持续经营净利润（净亏损以“－”号填列）",
    "2.终止经营净利润（净亏损以“－”号填列）",
    "减：*少数股东损益",
    "******六、归属于母公司所有者的净利润",
    "加：年初未分配利润",
    "其他综合收益结转留存收益",
    "其他转入",
    "*******七、可供分配的利润",
    "减：1.提取法定盈余公积",
    "2.提取任意盈余公积",
    "3.提取一般风险准备",
    "4.提取职工奖励及福利基金",
    "5.提取储备基金公积",
    "6.提取企业发展基金",
    "7.利润归还投资公积",
    "8.其他",
    "********八、可供投资者分配的利润",
    "减：1.应付优先股股利",
    "2.应付普通股股利",
    "3.转作股本的普通股股利",
    "*********九、期末未分配利润",
]

# --------------------------
# 3) 默认“需要扣除联营合营投资收益”的项目（内置，避免你手输）
#    UI口仍保留：用户输入的会在此基础上追加
# --------------------------
DEFAULT_EXCLUDE_INV_ITEMS = [
    "***三、营业利润（损失以“-”号填列）",
    "*****五、净利润（净亏损以“-”号填列）",
    "******1.持续经营净利润（净亏损以“－”号填列）",
    "****四、利润总额（亏损总额以“-”号填列）",
    "******六、归属于母公司所有者的净利润",
    "*******七、可供分配的利润",
    "********八、可供投资者分配的利润",
    "*********九、期末未分配利润",
]

# --------------------------
# 4) 别名映射（可选）：成本表项目名 -> SAP项目名
# --------------------------
ALIAS_MAP = {
    # 需要时你自己配置，例如：
    # "利息收入": "△利息收入",
}

# --------------------------
# 5) 吨油利润相关
# --------------------------
K_PROFIT_TOTAL = normalize_item_name("****四、利润总额（亏损总额以“-”号填列）")
K_TON_OIL_PROFIT = normalize_item_name("吨油利润")

K_21_ANCHOR_A = normalize_item_name("原油(一般贸易)")
K_21_ANCHOR_B = normalize_item_name("来料加工-合计")

K_11_RAW_A = normalize_item_name("原料-原油")
K_11_RAW_B = normalize_item_name("原料-来料加工原油")


def _norm_set(items: List[str]) -> set:
    return {normalize_item_name(x) for x in items if normalize_item_name(x)}


def _flip_sign_for_items(table: Dict[str, Dict], items: List[str]):
    """对指定项目清单：m/y/ly 全部取相反数（乘以 -1）。"""
    keys = _norm_set(items)
    for k in keys:
        if k not in table:
            continue
        for fld in ("m", "y", "ly"):
            table[k][fld] = negate(table[k].get(fld))


def _get_row_key_from_sheet6(cell_value) -> str:
    return normalize_item_name(cell_value)


def _apply_alias(key: str) -> List[str]:
    out = [key]
    for raw_k, raw_v in ALIAS_MAP.items():
        if normalize_item_name(raw_k) == key:
            out.append(normalize_item_name(raw_v))
    return out


def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def _safe_div(numer: float, denom: float) -> float:
    denom_f = float(denom)
    if abs(denom_f) < 1e-12:
        return 0.0
    return float(numer) / denom_f


def _find_row_by_anchor(ws: Worksheet, anchor_key: str, col_candidates=(1, 2, 3, 4)) -> Optional[int]:
    """
    在给定工作表中按“锚点文本”找行号。
    默认会在 A/B/C/D 列尝试匹配（normalize 后等于 anchor_key）。
    """
    if not ws:
        return None
    for r in range(1, ws.max_row + 1):
        for c in col_candidates:
            k = normalize_item_name(ws.cell(r, c).value)
            if k and k == anchor_key:
                return r
    return None


def _read_qty_from_ws(ws: Worksheet, anchor_key: str, qty_col: int) -> float:
    """
    从 ws 中按 anchor 找行，再读指定列 qty_col 的数值。
    找不到行或值非法 -> 0
    """
    r = _find_row_by_anchor(ws, anchor_key)
    if r is None:
        return 0.0
    return _to_float(ws.cell(r, qty_col).value)


def fill_sheet6_profit(
    ws6: Worksheet,
    old_all: Dict[str, Dict],
    new_all: Dict[str, Dict],
    ui_exclude_items: Optional[List[str]] = None,
    # ✅新增：用于“吨油利润”分母取数（不传也不报错，吨油利润写 0）
    ws_21: Optional[Worksheet] = None,   # 2-1
    ws_22: Optional[Worksheet] = None,   # 2-2
    ws_11: Optional[Worksheet] = None,   # 1-1
):
    """
    规则：
    - 利旧6项：从旧表取数；仅前三个收入类翻符号，后三个成本类不翻
    - 利新固定清单（含联营）：翻符号后才是真值
    - 投资收益差额/扣除联营：使用【翻符号后的联营值】做差（不取绝对值）
    - 清空联营行
    - 写入万元（空写 0）
    - ✅新增：吨油利润（基于“利润总额(本月/本年/上年同期)” ÷ 对应加工量）
    """

    k_inv = normalize_item_name("投资收益（损失以“-”号填列）")
    k_assoc = normalize_item_name("其中：对联营企业和合营企业")

    # 1) 利旧：仅前三项翻符号
    _flip_sign_for_items(old_all, OLD_FLIP_SIGN_ITEMS)

    # 2) 利新固定清单：翻符号（✅包含联营）
    _flip_sign_for_items(new_all, LIXIN_FLIP_SIGN_ITEMS)

    # 3) 取“翻符号后的联营值”（翻后才是真值）
    asc = new_all.get(k_assoc, {"m": None, "y": None, "ly": None})
    asc_m, asc_y, asc_ly = asc.get("m"), asc.get("y"), asc.get("ly")

    # 4) 投资收益差额：投资收益 = 投资收益 - 联营（两者都是翻符号后的真值）
    inv = new_all.get(k_inv, {"m": None, "y": None, "ly": None})
    new_all.setdefault(k_inv, {})["m"] = sub(inv.get("m"), asc_m)
    new_all[k_inv]["y"] = sub(inv.get("y"), asc_y)
    new_all[k_inv]["ly"] = sub(inv.get("ly"), asc_ly)

    # 5) 扣除联营合营投资收益的项目：默认清单 + UI清单
    exclude_items = list(DEFAULT_EXCLUDE_INV_ITEMS)
    if ui_exclude_items:
        exclude_items.extend([x for x in ui_exclude_items if str(x).strip()])
    exclude_keys = _norm_set(exclude_items)

    for k in exclude_keys:
        if k not in new_all:
            continue
        v = new_all[k]
        v["m"] = sub(v.get("m"), asc_m)
        v["y"] = sub(v.get("y"), asc_y)
        v["ly"] = sub(v.get("ly"), asc_ly)

    # 6) 清空联营行（三列）——按你“没值用0”的口径写 0（不再写 None）
    if k_assoc in new_all:
        new_all[k_assoc]["m"] = 0
        new_all[k_assoc]["y"] = 0
        new_all[k_assoc]["ly"] = 0

    # 7) 回填：逐行扫描成本表B列
    old_keys = _norm_set(OLD_6_ITEMS)

    for r in range(4, ws6.max_row + 1):
        raw_name = ws6.cell(r, 2).value
        key = _get_row_key_from_sheet6(raw_name)
        if not key:
            continue

        # 利旧6项优先用利旧，否则用利新
        src_table = old_all if key in old_keys else new_all

        # 别名映射（可选）
        cand_keys = _apply_alias(key)

        picked = None
        for ck in cand_keys:
            if ck in src_table:
                picked = src_table[ck]
                break

        # 默认写 0（不再写空/“-”）
        if picked is None:
            ws6.cell(r, 4).value = 0
            ws6.cell(r, 5).value = 0
            ws6.cell(r, 6).value = 0
            continue

        m_v = to_wanyuan(picked.get("m"))
        y_v = to_wanyuan(picked.get("y"))
        ly_v = to_wanyuan(picked.get("ly"))

        ws6.cell(r, 4).value = 0 if m_v is None else m_v
        ws6.cell(r, 5).value = 0 if y_v is None else y_v
        ws6.cell(r, 6).value = 0 if ly_v is None else ly_v

    # --------------------------
    # 8) ✅新增：吨油利润
    # 口径（你给的）：
    # - 本月：利润总额(本月数) / [2-1 原油(一般贸易)+来料加工-合计 的“本期加工数量”]
    # - 本年：利润总额(本年金额) / [2-2 同口径]
    # - 上年同期：利润总额(上年同期) / [1-1 原料-原油 + 原料-来料加工原油 的 M 列数量]
    #
    # 注：
    # - 利润总额在此时已完成翻符号与扣联营处理，取 new_all 的 “利润总额”
    # - 分母列假设：
    #   * 2-1/2-2 “本期加工数量”在 K 列（第 11 列）——与你之前口径一致
    #   * 1-1 “上年同期数量”在 M 列（第 13 列）——你明确指定 M 列
    # --------------------------
    # 8.1 取利润总额（万元）
    p = new_all.get(K_PROFIT_TOTAL, {"m": 0, "y": 0, "ly": 0})
    profit_m_wanyuan = _to_float(to_wanyuan(p.get("m")))
    profit_y_wanyuan = _to_float(to_wanyuan(p.get("y")))
    profit_ly_wanyuan = _to_float(to_wanyuan(p.get("ly")))

    # 8.2 取分母（数量）
    # 2-1 / 2-2：K 列（本期加工数量，通常为“万吨”）
    den_m = 0.0
    den_y = 0.0
    if ws_21 is not None:
        den_m = (
            _read_qty_from_ws(ws_21, K_21_ANCHOR_A, qty_col=11) +
            _read_qty_from_ws(ws_21, K_21_ANCHOR_B, qty_col=11)
        )
    if ws_22 is not None:
        den_y = (
            _read_qty_from_ws(ws_22, K_21_ANCHOR_A, qty_col=11) +
            _read_qty_from_ws(ws_22, K_21_ANCHOR_B, qty_col=11)
        )

    # 1-1：M 列（上年同期数量），你明确指定
    den_ly = 0.0
    if ws_11 is not None:
        den_ly_ton = (
            _read_qty_from_ws(ws_11, K_11_RAW_A, qty_col=13) +
            _read_qty_from_ws(ws_11, K_11_RAW_B, qty_col=13)
        )
        den_ly=den_ly_ton/10000.0

    # 8.3 计算吨油利润
    # 利润为“万元”，数量按你口径直接用对应列的数量（通常单位：万吨）。
    # 这里不做单位转换，沿用你指定的分母列；你若希望换算成“元/吨”，可以后续再明确单位要求。
    ton_oil_profit_m = _safe_div(profit_m_wanyuan, den_m)
    ton_oil_profit_y = _safe_div(profit_y_wanyuan, den_y)
    ton_oil_profit_ly = _safe_div(profit_ly_wanyuan, den_ly)

    # 8.4 写回 Sheet6 的“吨油利润”行（匹配 B 列项目名）
    r_ton = _find_row_by_anchor(ws6, K_TON_OIL_PROFIT, col_candidates=(2,))
    if r_ton is not None:
        ws6.cell(r_ton, 4).value = ton_oil_profit_m
        ws6.cell(r_ton, 5).value = ton_oil_profit_y
        ws6.cell(r_ton, 6).value = ton_oil_profit_ly
